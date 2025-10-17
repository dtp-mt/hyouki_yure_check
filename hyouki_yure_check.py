import sys
import os
import re
import time
import unicodedata
import difflib
from collections import Counter, defaultdict
from dataclasses import dataclass
from typing import List, Tuple, Dict, Optional, Set
import fitz  # PyMuPDF
import pandas as pd

from PySide6.QtCore import (
    Qt, QAbstractTableModel, QModelIndex, QSortFilterProxyModel, Signal, QObject, QThread, QRect,
    QTimer, QCollator, QPoint, QSize   # ← これを追加
)

from PySide6.QtGui import (
    QFont, QPalette, QColor, QPainter, QPen, QAction
)
# ===== [PATCH] import QSizePolicy =====
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QFileDialog, QVBoxLayout, QHBoxLayout, QLabel,
    QPushButton, QTableView, QTabWidget, QLineEdit, QDoubleSpinBox, QSpinBox,
    QGroupBox, QFormLayout, QMessageBox, QTextEdit, QTextBrowser, QProgressBar, QAbstractItemView,
    QListWidget, QCheckBox, QStyledItemDelegate, QStyle, QProxyStyle,
    QMenu, QDialog, QDialogButtonBox, QFrame, QHeaderView, QSizePolicy
)
# ===== [/PATCH] ======================

# ===== [ADD] CPU負荷抑制（結果を変えずに並列度と優先度だけ下げる） =====
import os
import sys

# 1) BLAS/OpenMP 等の並列スレッド数を起動時に制限（必ずプロセス再起動が必要）
#    これで NumPy/MKL/OpenBLAS/numexpr 等が内部で多数スレッドを立てるのを抑止します。
os.environ.setdefault("OMP_NUM_THREADS", "1")
os.environ.setdefault("OPENBLAS_NUM_THREADS", "1")
os.environ.setdefault("MKL_NUM_THREADS", "1")
os.environ.setdefault("VECLIB_MAXIMUM_THREADS", "1")
os.environ.setdefault("NUMEXPR_NUM_THREADS", "1")

# 2) Thinc / spaCy が CUDA を誤検出するのを防ぐ（GPU を使っていないなら冗長だが安全）
os.environ.setdefault("CUDA_VISIBLE_DEVICES", "")
os.environ.setdefault("USE_CUPY", "0")

# 3) プロセス優先度を下げて OS レベルで CPU を譲る（結果の順序や中身には影響なし）
try:
    if sys.platform.startswith("linux") or sys.platform.startswith("darwin"):
        # nice を +10 にして優先度を低くする（通常のユーザ権限で可能）
        try:
            os.nice(10)
        except Exception:
            pass
    elif sys.platform.startswith("win"):
        # Windows: pywin32 が無ければ ctypes 経由で優先度を BELOW_NORMAL に下げる
        try:
            import ctypes, ctypes.wintypes as wintypes
            pid = os.getpid()
            PROCESS_SET_INFORMATION = 0x0200
            handle = ctypes.windll.kernel32.OpenProcess(PROCESS_SET_INFORMATION, False, pid)
            if handle:
                BELOW_NORMAL_PRIORITY_CLASS = 0x00004000
                ctypes.windll.kernel32.SetPriorityClass(handle, BELOW_NORMAL_PRIORITY_CLASS)
                ctypes.windll.kernel32.CloseHandle(handle)
        except Exception:
            pass
except Exception:
    pass

# 4) もし可能なら「実行中プロセスを特定コアに制限（affinity）」して同時稼働コア数を減らす
#    Linux では os.sched_setaffinity が使えます。Windows では ctypes か psutil が必要（後述）。
try:
    # ここで使う core_count は 1..N の間で試し、効果と速度を見てください（例: 0..3 の4コアに限定）
    desired_cores = None  # None のままだと変更しない。例: set([0,1]) で 2 コアに限定
    if desired_cores:
        if hasattr(os, "sched_setaffinity"):
            os.sched_setaffinity(0, set(desired_cores))
        elif sys.platform.startswith("win"):
            # Windows の場合、psutil があれば簡単に設定可能
            try:
                import psutil
                p = psutil.Process(os.getpid())
                p.cpu_affinity(list(desired_cores))
            except Exception:
                pass
except Exception:
    pass

# 5) spaCy/GiNZA の内部で使われる .pipe(..., n_process=...) を 1 にしていることを前提にする
#    （このスクリプト内の ginza 呼び出しは既に n_process=1 にしている想定）
# ===== [/ADD] =====

# 先頭の import 群のどこかに追加
from PySide6.QtCore import QUrl
from PySide6.QtGui import QDesktopServices

# ===== [ADD] NLPテキスト保存（ON/OFF）と保存ユーティリティ =====
SAVE_NLP_TEXT: bool = False
SAVE_NLP_DIR: str = os.path.join(os.getcwd(), "nlp_inputs")

def set_save_nlp(enable: bool, out_dir: Optional[str] = None):
    """保存の ON/OFF を切り替え。out_dir を指定すると保存先を変更できます。"""
    global SAVE_NLP_TEXT, SAVE_NLP_DIR
    SAVE_NLP_TEXT = bool(enable)
    if out_dir:
        SAVE_NLP_DIR = str(out_dir)
    # ON 時のみ事前作成
    if SAVE_NLP_TEXT:
        try:
            os.makedirs(SAVE_NLP_DIR, exist_ok=True)
        except Exception:
            pass

def _safe_file_part(s: str) -> str:
    if not isinstance(s, str):
        s = "" if s is None else str(s)
    s = s.strip()
    # ファイル名用に無難化（日本語はそのまま、記号は _）
    return re.sub(r'[^A-Za-z0-9_.\-ぁ-んァ-ン一-龥]', '_', s)

def save_nlp_input(kind: str, text: str, *, suffix: str = "", src: Optional[str] = None) -> Optional[str]:
    """ON のとき kind/suffix 付きで .txt 保存。戻り値は保存パス（失敗時 None）。"""
    if not SAVE_NLP_TEXT:
        return None
    try:
        from datetime import datetime
        base = SAVE_NLP_DIR or os.getcwd()
        day = datetime.now().strftime("%Y%m%d")
        kind_dir = os.path.join(base, day, _safe_file_part(kind))
        os.makedirs(kind_dir, exist_ok=True)
        ts = datetime.now().strftime("%H%M%S_%f")
        suf = _safe_file_part(suffix) if suffix else "text"
        fn = f"{ts}_{suf}.txt"
        path = os.path.join(kind_dir, fn)
        meta = []
        if src:
            meta.append(f"[src] {src}")
        meta.append(f"[kind] {kind}")
        meta_line = ("# " + " | ".join(meta) + "\n") if meta else ""
        with open(path, "w", encoding="utf-8") as f:
            f.write(meta_line)
            f.write(text or "")
        return path
    except Exception:
        return None
# ===== [/ADD] =====


# === 帯付き Levenshtein & しきい値付き類似度判定（新規追加） ===
def levenshtein_band(a: str, b: str, k: int) -> Optional[int]:
    """
    帯幅 k の制限付き Levenshtein 距離。
    距離が k を超えることが確定したら None を返して早期終了。
    """
    la, lb = len(a), len(b)
    if a == b:
        return 0
    if abs(la - lb) > k:
        return None  # 距離 > k が確定
    if la < lb:
        a, b = b, a
        la, lb = lb, la
    INF = k + 1
    prev = [INF] * (lb + 1)
    for j in range(min(lb, k) + 1):
        prev[j] = j
    for i in range(1, la + 1):
        start = max(1, i - k)
        end   = min(lb, i + k)
        cur = [INF] * (lb + 1)
        if start == 1:
            cur[0] = i
        for j in range(start, end + 1):
            cost = 0 if a[i-1] == b[j-1] else 1
            cur[j] = min(prev[j] + 1,         # 削除
                         cur[j-1] + 1,         # 挿入
                         prev[j-1] + cost)     # 置換/一致
        # 早期打切り：帯内すべてが k 超なら終了
        if min(cur[start:end+1]) > k:
            return None
        prev = cur
    return prev[lb] if prev[lb] <= k else None

def sim_with_threshold(a: str, b: str, th: float) -> Optional[float]:
    """
    正規化類似度 >= th のときその値を返す。未満または判定不能は None。
    """
    la, lb = len(a), len(b)
    if la == 0 and lb == 0:
        return 1.0
    L = max(la, lb)
    k = int((1.0 - th) * L)
    d = levenshtein_band(a, b, k)
    if d is None:
        return None
    return 1.0 - d / L

# ===== [ADD] 軽量な縦組み判定＋縦テキスト復元ヘルパー =====
def _v_is_vertical_line(ln) -> bool:
    """rawdictの1行が縦組みっぽいかを判定（wmode/dir優先＋簡易形状）"""
    try:
        wmode = ln.get("wmode")
        if isinstance(wmode, int) and wmode == 1:
            return True
        dirv = ln.get("dir")
        if isinstance(dirv, (list, tuple)) and len(dirv) >= 2:
            dx, dy = float(dirv[0]), float(dirv[1])
            if abs(dy) > abs(dx):
                return True
    except Exception:
        pass

    # chars の幾何から fallback 判定（xs/ys の広がり比較）
    def _chars_of_span(sp):
        size = float(sp.get("size", 0.0) or 0.0)
        chs = sp.get("chars")
        if isinstance(chs, list) and chs:
            for ch in chs:
                x0, y0, x1, y1 = map(float, ch.get("bbox", (0,0,0,0)))
                yield (x0+x1)/2.0, (y0+y1)/2.0
        else:
            txt = sp.get("text") or ""
            if not txt:
                return
            x0, y0, x1, y1 = map(float, sp.get("bbox", (0,0,0,0)))
            n = max(1, len(txt))
            w = (x1-x0)/n if n else 0.0
            for i in range(n):
                cx = x0 + (i+0.5)*w
                cy = (y0+y1)/2.0
                yield cx, cy

    xs, ys = [], []
    for sp in ln.get("spans", []):
        for cx, cy in _chars_of_span(sp):
            xs.append(cx); ys.append(cy)
    if len(xs) < 2:
        return False

    try:
        import statistics as stats
        qx = stats.quantiles(xs, n=4)
        qy = stats.quantiles(ys, n=4)
        iqr_x = qx[2]-qx[0]
        iqr_y = qy[2]-qy[0]
    except Exception:
        # 平均平方偏差っぽい簡易
        mx = sum(xs)/len(xs); my = sum(ys)/len(ys)
        iqr_x = sum((x-mx)*(x-mx) for x in xs)/len(xs)
        iqr_y = sum((y-my)*(y-my) for y in ys)/len(ys)

    # “縦っぽさ”のしきいは既存と近い 1.6 倍基準
    return (iqr_y > iqr_x * 1.6)


# ==== 読み類似スコア（数・記号は読み対象外／満点抑制の強化版） ==================
KANJI_RE = re.compile(r"[\u4E00-\u9FFF\u3400-\u4DBF\uF900-\uFAFF]")

from functools import lru_cache

@lru_cache(maxsize=4096)
def _has_kanji(s: str) -> bool:
    """文字列に漢字が含まれるか（キャッシュ付き）"""
    if not isinstance(s, str):
        s = "" if s is None else str(s)
    s = unicodedata.normalize("NFKC", s)
    return bool(KANJI_RE.search(s))

_NFKC_SPACE_RE = re.compile(r"\s+")

@lru_cache(maxsize=4096)
def _norm_len_for_surface(s: str) -> int:
    """表層長（比較用）。NFKC→空白除去。キャッシュ付き高速版。"""
    if s is None:
        return 0
    s = unicodedata.normalize("NFKC", str(s))
    s = _NFKC_SPACE_RE.sub("", s)
    return len(s)


# ===== [Patch] 読み一致（reading_eq）のスコアを 1.0 固定にしない軽量スコアラー =====
def _numeric_chunk_count(s: str) -> int:
    """
    NFKC 後の文字列に _NUMERIC_CHUNK_RE を当てて、数値“まとまり”の個数を数える。
    例: '630W[30°C]' -> 2（630, 30）
    """
    try:
        nf = unicodedata.normalize("NFKC", "" if s is None else str(s))
    except Exception:
        nf = "" if s is None else str(s)
    return sum(1 for _ in _NUMERIC_CHUNK_RE.finditer(nf))

def reading_eq_score(a: str, b: str) -> float:
    """
    'reading_eq' に分類されたペアの “類似度” を 1.0 固定にせず、
    a/b の差に応じた小さな減点を与えて 0.90〜0.995 の範囲に収める。

    減点規則（初期チューニング）：
      - P0: a != b                   …… -0.01
      - P1: 表層長差（NFKC/空白除去）…… -min(0.05, 0.01 * |len_a - len_b|)
      - P2: 数値チャンク個数差             …… -min(0.02, 0.005 * |numA - numB|)
      - P3: 片側だけ漢字を含む                …… -0.01

    最終クランプ：0.90 <= score <= 0.995
    """
    a_s = "" if a is None else str(a)
    b_s = "" if b is None else str(b)

    score = 1.0

    # P0: 表層が同一でない（reading_eq は基本 a!=b の想定）
    if a_s != b_s:
        score -= 0.01

    # P1: NFKC + 空白除去の長さ差
    try:
        la = _norm_len_for_surface(a_s)  # 既存：nfkc -> 空白除去 -> len
        lb = _norm_len_for_surface(b_s)
        diff_len = abs(la - lb)
        if diff_len > 0:
            score -= min(0.05, 0.01 * diff_len)
    except Exception:
        pass

    # P2: 数値“まとまり”個数差
    try:
        na = _numeric_chunk_count(a_s)
        nb = _numeric_chunk_count(b_s)
        diff_num = abs(na - nb)
        if diff_num > 0:
            score -= min(0.02, 0.005 * diff_num)
    except Exception:
        pass

    # P3: 片方だけ漢字を含む
    try:
        if _has_kanji(a_s) != _has_kanji(b_s):
            score -= 0.01
    except Exception:
        pass

    # 取りすぎない・取りなさすぎない
    score = max(0.90, min(0.995, score))

    # UI 側の丸めと整合（内部でも丸めておく）
    return round(float(score), 3)
# ===== [Patch end] ============================================================

def _temp_marker_core_after_number_strip(s: str) -> str:
    """
    NFKC → 数字まとまり除去（既存の _strip_numbers_like）後、
    温度記号に関係する最小核だけを取り出す。
    - '°' と 'C/c' がともに残っていれば '°C' を返す
    - それ以外は空文字
    """
    if s is None:
        return ""
    # 既存の数字まとまり除去（NFKC 内部で実施される）
    t = _strip_numbers_like(s)  # 例: "630W[30°C]" -> "W[°C]"
    if not t:
        return ""
    # 温度記号に関係ない文字は落とし、'°' と 'C/c' だけ残す
    keep = []
    for ch in t:
        if ch == "°" or ch in ("C", "c"):
            keep.append(ch)
    has_deg = "°" in keep
    has_c   = ("C" in keep) or ("c" in keep)
    return "°C" if has_deg and has_c else ""

def is_symbol_only_surface_diff(a: str, b: str) -> bool:
    """
    表層 a, b のうち、ひらがな/カタカナ/漢字だけを取り出した “和字コア” が
    一致するかどうか（= 記号・英数・単位の差だけか）を判定する。

    変更点（℃専用の救済）：
      - 和字コアが両方 空 の場合に限り、
        「数字を除いた後」に温度記号の最小核（° + C/c）が
        双方で満たされる（= '°C'）なら True を返す。
      - 単位一般は扱わず、あくまで「℃」相当のケースに限定。

    これにより、以下が True（= reading_eq, score=1.0 対象）になります：
      - "20℃" ↔ "30℃"
      - "30℃" ↔ "630W［30℃"
    反対に、和字コアが非対称（例: "℃" ↔ "たび"）は False のまま。
    """
    ca = _surface_core_for_reading(a)
    cb = _surface_core_for_reading(b)

    # 既存ルール：和字コアが非空で一致 → True
    if ca and cb:
        return ca == cb

    # 追加ルール（℃専用）：
    # 両方とも和字コアが空 かつ 「数字を除いた後」の温度核が双方 '°C'
    if not ca and not cb:
        ta = _temp_marker_core_after_number_strip(a)
        tb = _temp_marker_core_after_number_strip(b)
        if ta and tb and (ta == tb == "°C"):
            return True

    return False

def reading_sim_with_penalty(a_surface: str, b_surface: str, ra: str, rb: str, th: float) -> Optional[float]:
    """
    読み同士の類似度（帯付きLevenshtein）を返すが、以下の調整を行う：

    [A] “和字コア”が一致する（= 数字・記号・空白など読み対象外だけが違う）場合は
        読み類似を強制で 1.0（満点）にする。
        例: 「000円」vs「250円」→ コアはどちらも「円」 ⇒ 1.0

    [B] それ以外で満点(≈1.0)になった場合は、満点抑制を適用：
        [P0] 表層が異なるだけで 1.0 は避ける       → -0.01
        [P1] 片方だけ漢字を含む（かな/漢字表記差）→ さらに -0.01（計 -0.02）
        [P2] 表層長が異なる（読み対象文字数差）   → さらに -0.05

    例：
      「電気」vs「電機」  → P0: 1.00→0.99
      「とき」vs「冬季」  → P0+P1: 1.00→0.98
      「離し」vs「話」    → P0+P2: 1.00→0.94
    """
    ra = "" if ra is None else str(ra)
    rb = "" if rb is None else str(rb)

    # --- [A] 数字・記号・空白などを除いた“和字コア”が一致するなら、読み=1.0 扱い ---
    core_a = _surface_core_for_reading(a_surface)
    core_b = _surface_core_for_reading(b_surface)
    if core_a and core_b and core_a == core_b:
        # 読み生成（ra/rb）が数字読みなどで異なっても、読み対象はコアのみとみなす
        return 1.0

    # --- 読みが両方空なら、読み類似は使わない ---
    if len(ra) == 0 and len(rb) == 0:
        return None

    sim = sim_with_threshold(ra, rb, th)
    if sim is None:
        return None

    # --- [B] 満点(≈1.0)のときだけ抑制ロジックを適用 ---
    if sim >= 0.9999:
        penalized = 1.0

        # [P0] 表層差があるなら -0.01
        if str(a_surface) != str(b_surface):
            penalized -= 0.01

        # [P1] 片方だけ漢字を含むなら -0.01（合計 -0.02）
        if _has_kanji(a_surface) != _has_kanji(b_surface):
            penalized -= 0.01

        # [P2] 表層長（NFKC/空白無視）が異なるなら -0.05
        if _norm_len_for_surface(a_surface) != _norm_len_for_surface(b_surface):
            penalized -= 0.05

        sim = penalized

    return sim

# ==== 読み類似のスコア再採点（lemma/活用も“読み”基準で減点 + P4 追加） ============
READING_LIKE_REASONS = {"lemma", "lemma_read", "inflect", "reading"}

# 追加：読み不一致の追加減点（調整しやすいように定数化）
READ_MISMATCH_PENALTY = 0.03  # ←「0.99より下げたい」ニュアンスに最適化

from functools import lru_cache

@lru_cache(maxsize=4096)
def _reading_for_surface_cached(s: str) -> str:
    """
    表層 s の“読み（骨格）”。MeCab があればそれ、無ければ簡易フォールバック。
    キャッシュ強化済み（maxsize=4096）。
    """
    ok, _ = ensure_mecab()
    if ok:
        return phrase_reading_norm_cached(s or "")
    # フォールバック（NFKC→ひら->カナ→カナ以外除去→長音除去）
    return normalize_kana(hira_to_kata(nfkc(s or "")), drop_choon=True)

def _lemma_joined_for_surface(s: str) -> str:
    """
    表層 s の lemma を連結して返す（活用差の検出に使う）。
    例: 「あわせ」「合わせる」→ ともに '合わせる'（期待）
    """
    ok, _ = ensure_mecab()
    if not ok or MECAB_TAGGER is None:
        # MeCab 無しなら NFKC 表層を代用（厳密でなくてOK）
        return nfkc(s or "")
    toks = tokenize_mecab(s or "", MECAB_TAGGER)
    # lemma が無ければ surface を採用
    return "".join((lemma if (lemma and isinstance(lemma, str)) else surf)
                   for (surf, pos1, lemma, reading, ct, cf) in toks)

@lru_cache(maxsize=4096)
def _surface_core_for_reading(s: str) -> str:
    """読み対象の“和字コア”（漢字・ひらがな・カタカナのみ）を抽出。キャッシュ付き高速版。"""
    nf = unicodedata.normalize("NFKC", str(s or ""))
    return "".join(ch for ch in nf if (
        0x3040 <= ord(ch) <= 0x309F or  # ひらがな
        0x30A0 <= ord(ch) <= 0x30FF or  # カタカナ
        0x3400 <= ord(ch) <= 0x4DBF or  # CJK統合漢字拡張A
        0x4E00 <= ord(ch) <= 0x9FFF or  # CJK統合漢字
        0xF900 <= ord(ch) <= 0xFAFF     # 互換漢字
    ))

# ===== 追加: 囲み数字などの「リストマーカー」を判定 =====
def _is_enclosed_list_marker(ch: str) -> bool:
    """
    '①' '②' ... '㉑' や '❶' '❷' ... などの囲み数字系は
    「リストマーカー」とみなし、ノイズ扱いしないための判定。
    """
    if not ch:
        return False
    o = ord(ch)
    # Enclosed Alphanumerics (一部): ⓪(0x24EA), ①..⑳(0x2460..0x2473), ⑴..⒇ 等
    if 0x2460 <= o <= 0x24FF or o == 0x24EA:
        return True
    # Dingbats (Negative circled digits 等): ❶..➓(0x2776..0x2793)
    if 0x2776 <= o <= 0x2793:
        return True
    # Enclosed Alphanumeric Supplement (拡張の囲み文字): U+1F100..U+1F1FF
    if 0x1F100 <= o <= 0x1F1FF:
        return True

# ===== 置換: 非日本語(英字・数字・記号)ノイズの有無を判定 =====
def _has_non_japanese_noise(s: str) -> bool:
    """
    '英数字や記号によるノイズ' が含まれるか。
    - 日本語スクリプト(漢字/ひら/カタカナ)と空白は無視
    - 囲み数字などの「リストマーカー」(①, ❶ など)はノイズに数えない
    - それ以外の ASCII/全角英字・数字・各種記号はノイズとみなす
    """
    if s is None:
        return False
    # 原字のまま評価（NFKCにすると ①→"1" に潰れてしまうため）
    for ch in str(s):
        if not ch or ch.isspace():
            continue
        o = ord(ch)
        # 日本語の主要スクリプト → ノイズではない
        if (
            0x3040 <= o <= 0x309F   # ひらがな
            or 0x30A0 <= o <= 0x30FF  # カタカナ
            or 0x3400 <= o <= 0x4DBF  # CJK統合漢字拡張A
            or 0x4E00 <= o <= 0x9FFF  # CJK統合漢字
            or 0xF900 <= o <= 0xFAFF  # 互換漢字
        ):
            continue
        # 囲み数字などの「リストマーカー」 → ノイズに数えない
        if _is_enclosed_list_marker(ch):
            continue
        # ここからはノイズ候補
        # ・英数字（ASCII/全角問わず）
        # ・各種記号（Unicode Category 'S*'）や句読点等（'P*'）
        cat = unicodedata.category(ch)  # 'Nd' 数字, 'Ll' 小文字, 'Lu' 大文字, 'S*' 記号, 'P*' 句読点など
        if ch.isdigit() or ('A' <= ch <= 'Z') or ('a' <= ch <= 'z'):
            return True
        if cat.startswith('S') or cat.startswith('P'):
            return True
        # 全角英字（NFKCせずともカテゴリで拾えない場合があるため補助）
        ch_nfkc = nfkc(ch)
        if ch_nfkc != ch:
            if ch_nfkc.isdigit() or ('A' <= ch_nfkc <= 'Z') or ('a' <= ch_nfkc <= 'z'):
                return True
            cat2 = unicodedata.category(ch_nfkc)
            if cat2.startswith('S') or cat2.startswith('P'):
                return True


def _script_pattern(s: str) -> str:
    """
    文字ごとのスクリプト種別列（K:漢字/H:ひら/C:カナ/O:その他）を返す。
    スペースは除外。NFKC 後のコードポイントに基づく。
    """
    def tag(ch: str) -> str:
        o = ord(ch)
        if 0x3400 <= o <= 0x4DBF or 0x4E00 <= o <= 0x9FFF or 0xF900 <= o <= 0xFAFF:
            return "K"
        if 0x3040 <= o <= 0x309F:
            return "H"
        if 0x30A0 <= o <= 0x30FF:
            return "C"
        return "O"
    nf = nfkc(str(s or ""))
    return "".join(tag(ch) for ch in nf if not ch.isspace())

def recalibrate_reading_like_scores(df: pd.DataFrame, read_th: float) -> pd.DataFrame:
    """
    {lemma, lemma_read, inflect, reading} を読みルールで再採点し reason='reading' に揃える。
    和字コア一致は強制1.0維持。満点時の抑制・P3/P4は現状ロジックを踏襲。
    """
    # --- 早期スキップ条件 ---
    if df is None or df.empty:
        return df
    if not {"a", "b", "reason", "score"}.issubset(df.columns):
        return df

    mask = df["reason"].isin(READING_LIKE_REASONS)
    if not mask.any():
        return df

    df = df.copy()

    # ---- 前計算（surface単位）
    surfaces = set(df.loc[mask, "a"].astype(str)) | set(df.loc[mask, "b"].astype(str))
    if not surfaces:
        return df  # 早期スキップ：対象表層が空

    def _batch_reading_and_lemma(surfaces: Set[str]) -> Tuple[Dict[str, str], Dict[str, str]]:
        """
        表層群に対して一括で読みと lemma を取得（MeCabバッチ化）。
        戻り値: (表層→読み, 表層→lemma連結)
        """
        ok, _ = ensure_mecab()
        if not ok or MECAB_TAGGER is None:
            # フォールバック：キャッシュ付き個別処理
            return (
                {s: _reading_for_surface_cached(s) for s in surfaces},
                {s: _lemma_joined_for_surface(s) for s in surfaces}
            )

        # --- 一括解析 ---
        # 各表層を改行区切りで連結し、一度に MeCab へ渡す
        joined = "\n".join(surfaces)
        toks = tokenize_mecab(joined, MECAB_TAGGER)

        result_read: Dict[str, str] = {}
        result_lemma: Dict[str, str] = {}

        cur_surface_parts: List[str] = []
        cur_lemma_parts: List[str] = []
        cur_idx = 0
        surf_list = list(surfaces)

        for surf in surf_list:
            cur_surface_parts.clear()
            cur_lemma_parts.clear()
            while cur_idx < len(toks):
                tsurf, _, lemma, reading, _, _ = toks[cur_idx]
                cur_idx += 1
                if tsurf == "\n":
                    break
                cur_surface_parts.append(reading or "")
                cur_lemma_parts.append(lemma if lemma else tsurf)
            result_read[surf] = "".join(cur_surface_parts)
            result_lemma[surf] = "".join(cur_lemma_parts)

        return result_read, result_lemma

    surf2read, surf2lemma = _batch_reading_and_lemma(surfaces)

    surf2core  = {s: _surface_core_for_reading(s) for s in surfaces}
    surf2script = {s: _script_pattern(s) for s in surfaces}

    def _calc(row):
        a = str(row["a"]); b = str(row["b"])

        # 1) 和字コア一致 → 読み 1.0 維持
        ca, cb = surf2core.get(a, ""), surf2core.get(b, "")
        if ca and cb and ca == cb:
            return 1.0

        ra, rb = surf2read.get(a, ""), surf2read.get(b, "")
        pattern_equal = (surf2script.get(a, "") == surf2script.get(b, ""))
        read_mismatch = (bool(ra) and bool(rb) and ra != rb)

        # 2) 読み類似（満点抑制は reading_sim_with_penalty 内で処理）
        sim = reading_sim_with_penalty(a, b, ra, rb, read_th)
        if sim is not None:
            val = float(sim)
            if val >= 0.9999 and a != b and (surf2lemma.get(a, "") == surf2lemma.get(b, "")):
                val -= 0.02
            if pattern_equal and read_mismatch:
                val -= READ_MISMATCH_PENALTY
            return max(val, 0.0)

        # 3) 読みが取れなかった場合でも、満点のままにしない（P0〜P4）
        penal = 1.0
        if a != b: penal -= 0.01
        if _has_kanji(a) != _has_kanji(b): penal -= 0.01
        if _norm_len_for_surface(a) != _norm_len_for_surface(b): penal -= 0.05
        if a != b and (surf2lemma.get(a, "") == surf2lemma.get(b, "")): penal -= 0.02
        if pattern_equal and read_mismatch: penal -= READ_MISMATCH_PENALTY
        return max(penal, 0.0)

    df.loc[mask, "score"] = df.loc[mask].apply(_calc, axis=1)
    df.loc[mask, "reason"] = "reading"
    return df


def _ngram_set(s: str, n: int = 2):
    return {s[i:i+n] for i in range(len(s)-n+1)} if len(s) >= n else {s}

# 事前コンパイル（モジュール先頭に一度だけ）
BULLET_RE = re.compile(
    "^[\u2022\u2023\u25E6\u2043\u2219\u25AA\u25CF\u25CB\u25A0\u25A1\u30FB\ufeff\\s\u3000]+"
)

def strip_leading_control_chars(s: str) -> str:
    if not isinstance(s, str):
        return s
    s = BULLET_RE.sub("", s)
    while s and (unicodedata.category(s[0]) in ("Cc", "Cf")):
        s = s[1:]
    return s

# ===== MeCab (fugashi) =====
try:
    from fugashi import Tagger
    HAS_MECAB = True
    MECAB_TAGGER: Optional["Tagger"] = None
except Exception:
    HAS_MECAB = False
    MECAB_TAGGER = None

# === 最終スコアを combined_char_similarity で強制（score_reasonは作らない既定） ===
def enforce_combined_similarity_score(
    df: pd.DataFrame,
    keep_backup: bool = False,   # 既定で退避しない
    drop_existing_backup: bool = True,  # 既にある 'score_reason' は落とす
) -> pd.DataFrame:
    """
    最終表示用に score を combined_char_similarity に差し替える。
    keep_backup=True のときだけ元スコアを score_reason に退避（.3f）。
    drop_existing_backup=True のとき、既にある score_reason を削除。
    """
    if df is None or df.empty:
        return df
    if not {"a", "b"}.issubset(df.columns):
        return df

    df = df.copy()

    # バックアップ列の扱い
    if drop_existing_backup:
        df.drop(columns=["score_reason"], errors="ignore", inplace=True)
    if keep_backup and "score" in df.columns and "score_reason" not in df.columns:
        df["score_reason"] = pd.to_numeric(df["score"], errors="coerce").round(3)

    # 表層キャッシュ（NFKC と 2-gram）
    try:
        uniq = pd.unique(pd.concat([df["a"], df["b"]]).astype("string"))
    except Exception:
        uniq = pd.unique(pd.concat([df["a"].astype(str), df["b"].astype(str)]))

    nf = {}
    grams = {}
    for s in uniq:
        s = "" if pd.isna(s) else str(s)
        ns = nfkc(s)
        nf[s] = ns
        grams[s] = frozenset(_ngram_set(ns, 2))

    def _calc(row):
        a = "" if pd.isna(row["a"]) else str(row["a"])
        b = "" if pd.isna(row["b"]) else str(row["b"])
        try:
            val = combined_char_similarity(
                a, b,
                sa=nf.get(a), sb=nf.get(b),
                grams_a=grams.get(a), grams_b=grams.get(b),
            )
        except Exception:
            # フォールバック：最悪でも 0.0〜1.0
            try:
                val = max(0.0, min(1.0, float(levenshtein_sim(nfkc(a), nfkc(b)))))
            except Exception:
                val = 0.0
        return round(float(val), 3)

    df["score"] = df.apply(_calc, axis=1)
    return df
# =====================================================================

# === 読み一致( reading_eq )で最終スコア1.0のものを 'basic' に付け替え ==========
def reclassify_basic_for_reading_eq(df: pd.DataFrame, eps: float = 0.0005) -> pd.DataFrame:
    """
    df に対し、reason=='reading_eq' かつ score が 1.0（丸め誤差±eps）を 'basic' へ変更。
    """
    if df is None or df.empty or "reason" not in df.columns or "score" not in df.columns:
        return df
    df = df.copy()
    # 丸め誤差込みで 1.0 判定（例: 0.9996 以上は 1.000 とみなす）
    m = (df["reason"] == "reading_eq") & (pd.to_numeric(df["score"], errors="coerce").fillna(0.0) >= 1.0 - eps)
    if m.any():
        df.loc[m, "reason"] = "basic"
        df.loc[m, "score"] = 1.0  # 表示上も 1.0 にそろえる
    return df
# =============================================================================

# ===== [ADD] phrase_reading_norm_keepchoon（長音あり読み生成） =====
def phrase_reading_norm_keepchoon(s: str) -> str:
    """
    表層 s から長音「ー」を保持した読みを生成。
    MeCabがあれば読みを使い、なければひら→カタカナ変換で代替。
    """
    if not isinstance(s, str) or not s:
        return ""
    ok, _ = ensure_mecab()
    if ok and MECAB_TAGGER is not None:
        toks = tokenize_mecab(s, MECAB_TAGGER)
        parts: List[str] = []
        for surf, pos1, lemma, reading, ct, cf in toks:
            r = reading or hira_to_kata(nfkc(surf))
            parts.append(r)
        joined = "".join(parts)
        return normalize_kana(joined, drop_choon=False)  # 長音保持
    else:
        return normalize_kana(hira_to_kata(nfkc(s or "")), drop_choon=False)

# =============================================================================
# 読み一致の厳密判定（前置/後置による包含を除外）
def _readings_equal_strict(a: str, b: str) -> bool:
    """
    ① 読み（長音保持）が完全一致（非空）であること
    ② ただし、和字コア（漢字・ひら・カナのみ）を比較して、
       一方がもう一方を strict な prefix/suffix として「包含」している場合は除外する
       （= 前置/後置で語が付加された可能性が高いケースを除く）
    例：還気温度センサ vs 温度センサ
        読み: （欠落により）どちらも 'オンドセンサ' になり得るが、
        和字コアは '還気温度センサ' と '温度センサ' で「後方一致」（suffix包含）→ NG
    """
    a_s = "" if a is None else str(a)
    b_s = "" if b is None else str(b)

    # 読み（長音保持）
    try:
        ra = phrase_reading_norm_keepchoon(a_s) or ""
        rb = phrase_reading_norm_keepchoon(b_s) or ""
    except Exception:
        ra, rb = "", ""

    if not (ra and rb):
        return False
    if ra != rb:
        return False

    # 和字コア（漢字・ひら・カナのみ）
    try:
        ca = _surface_core_for_reading(a_s) or ""
        cb = _surface_core_for_reading(b_s) or ""
    except Exception:
        ca, cb = "", ""

    if ca and cb and ca != cb:
        # strict な prefix/suffix 包含か？（完全一致は除外済み）
        if ca.endswith(cb) or cb.endswith(ca) or ca.startswith(cb) or cb.startswith(ca):
            # 前置/後置の付加（= 語の増減）とみなし、厳密一致から除外
            return False

    return True
# =============================================================================

# =============================================================================
# === [REPLACE] 読み一致（表記違い）: 厳格読み一致のみ昇格
#      ※ ただし「英数・記号ノイズ」が片側に含まれる場合は reading_eq を優先
# =============================================================================
def reclassify_reading_equal_formdiff(df: pd.DataFrame) -> pd.DataFrame:
    """
    'reading' の行から、以下の優先規則で再分類します：

    [優先1] 厳格読み一致(_readings_equal_strict) かつ
            片側にでも英数・記号ノイズ(_has_non_japanese_noise)が含まれる
            → reason='reading_eq', score=reading_eq_score(a,b)

    [優先2] それ以外で厳格読み一致(_readings_equal_strict) かつ a!=b
            → reason='reading_same'（従来昇格）

    [優先3] レマ一致 & 「名詞のみ」&「単一形態素」& 厳格読み一致
            → reason='reading_same'（従来昇格）

    なお、英数・記号だけの差は 'reading_eq' 領域（is_symbol_only_surface_diff True）
    とみなし本関数では扱いません（従来どおり）。
    """
    if (
        df is None or df.empty
        or "reason" not in df.columns
        or "a" not in df.columns or "b" not in df.columns
    ):
        return df

    # まず 'reading' 行のみ対象
    mask = df["reason"].eq("reading")
    if not mask.any():
        return df

    df = df.copy()

    # 表層キャッシュ
    sub = df.loc[mask, ["a", "b"]].astype("string")
    surfaces = pd.unique(pd.concat([sub["a"], sub["b"]], ignore_index=True))

    # 読み（長音保持）と レマ のキャッシュ
    surf2read_keep: Dict[str, str] = {}
    surf2lemma: Dict[str, str] = {}
    for s in surfaces:
        s_str = str(s)
        try:
            surf2read_keep[s_str] = phrase_reading_norm_keepchoon(s_str) or ""
        except Exception:
            surf2read_keep[s_str] = ""
        try:
            surf2lemma[s_str] = _lemma_joined_for_surface(s_str) or ""
        except Exception:
            surf2lemma[s_str] = ""

    # MeCab ユーティリティ
    ok_mecab, _ = ensure_mecab()
    tagger = MECAB_TAGGER if ok_mecab else None

    def _tokens(s: str):
        if not tagger:
            return []
        try:
            return tokenize_mecab(s, tagger)
        except Exception:
            return []

    def _is_noun_only(s: str) -> bool:
        toks = _tokens(s)
        if not toks:
            return False
        for _, pos1, *_ in toks:
            if not (isinstance(pos1, str) and pos1.startswith("名詞")):
                return False
        return True

    def _is_single_morpheme(s: str) -> bool:
        toks = _tokens(s)
        return len(toks) == 1

    # ここから再分類
    idx = df.index[mask]

    # 行ごとに適用
    for i in idx:
        a = "" if pd.isna(df.at[i, "a"]) else str(df.at[i, "a"])
        b = "" if pd.isna(df.at[i, "b"]) else str(df.at[i, "b"])
        if not a or not b or a == b:
            continue

        # 英数・記号だけの差は 'reading_eq' 領域（ここでは扱わない＝従来）
        try:
            if is_symbol_only_surface_diff(a, b):
                continue
        except Exception:
            pass

        # 厳格読み一致？
        try:
            eq_strict = _readings_equal_strict(a, b)
        except Exception:
            eq_strict = False
        if not eq_strict:
            continue

        # --- 優先1：片側にでも英数・記号ノイズが含まれるなら reading_eq 優先 ---
        try:
            if _has_non_japanese_noise(a) or _has_non_japanese_noise(b):
                df.at[i, "reason"] = "reading_eq"
                df.at[i, "score"] = reading_eq_score(a, b)
                continue
        except Exception:
            # 失敗時は従来どおりの昇格判断へフォールバック
            pass

        # --- 優先2：厳格読み一致のみでも 'reading_same' へ昇格 ---
        # （この分岐に来るのは、上のノイズ判定に該当しない純粋な表記差と考える）
        df.at[i, "reason"] = "reading_same"
        # score は既存のままでも良いが、読みの厳格一致なので 1.0 にしておく手もある
        # ただし最終 enforce_combined_similarity_score() が上書きする前提なら不要
        # df.at[i, "score"] = 1.0

        # --- 優先3：レマ一致 & 名詞のみ & 単一形態素 & 厳格読み一致 も reading_same ---
        # （上で 'reading_same' 済みなので実質そのまま。明示的に残しておくなら以下）
        # la, lb = surf2lemma.get(a, ""), surf2lemma.get(b, "")
        # if la and lb and (la == lb):
        #     if _is_noun_only(a) and _is_noun_only(b) and _is_single_morpheme(a) and _is_single_morpheme(b):
        #         df.at[i, "reason"] = "reading_same"

    return df


# =============================================================================
# 'reading_same' の最終サニタイズ（厳格読み一致を満たさなければ 'reading' に戻す）
def sanitize_reading_same(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty or "reason" not in df.columns:
        return df
    m = df["reason"].eq("reading_same")
    if not m.any():
        return df
    df = df.copy()
    bad = df.loc[m].apply(
        lambda r: not _readings_equal_strict(str(r.get("a", "")), str(r.get("b", ""))),
        axis=1
    )
    if bad.any():
        df.loc[bad[bad].index, "reason"] = "reading"  # 取りやめて 'reading' に戻す
    return df
# =============================================================================

# ------------------------------------------------------------
# 共通ユーティリティ
# ------------------------------------------------------------
PAIR_COLUMNS = ["a", "b", "a_count", "b_count", "reason", "score", "scope"]
def empty_pairs_df() -> pd.DataFrame:
    return pd.DataFrame(columns=PAIR_COLUMNS)

JP_BLOCK = r"[ぁ-んァ-ン一-龥々〆ヵヶA-Za-zＡ-Ｚａ-ｚ0-9０-９]+"
CONNECT = r"[ー\-－–—・/／\_＿]"  # 連結記号
TOKEN_RE = re.compile(rf"{JP_BLOCK}(?:{CONNECT}{JP_BLOCK})*")
CONNECT_CHARS = set("ー-－–—・/／_＿")

def nfkc(s: str) -> str:
    return unicodedata.normalize("NFKC", s)

def hira_to_kata(s: str) -> str:
    out = []
    for ch in s:
        o = ord(ch)
        out.append(chr(o + 0x60) if 0x3041 <= o <= 0x3096 else ch)
    return "".join(out)

def normalize_kana(s: str, drop_choon: bool = True) -> str:
    s = nfkc(s)
    s = hira_to_kata(s)
    s = "".join(ch for ch in s if 0x30A0 <= ord(ch) <= 0x30FF)
    if drop_choon:
        s = s.replace("ー", "")
    return s

def is_single_kana_char(s: str) -> bool:
    if not isinstance(s, str):
        return False
    s_norm = nfkc(s)
    if len(s_norm) != 1:
        return False
    o = ord(s_norm)
    return (0x3040 <= o <= 0x309F) or (0x30A0 <= o <= 0x30FF)

# ===== [REPLACE] 改行フラット化（英文は連結／英字1トークン・数字行は改行保持） =====
_ALPHA_TOKEN_RE     = re.compile(r"^[A-Za-z][A-Za-z0-9\-\._/]*$")   # ← 英字1トークン（空白なし）
_NUM_ONLY_LINE_RE   = re.compile(r"^[+\-]?\d[\d\s,./:\-–—]*$")      # 数字行（空白や桁区切りも許容）
_ASCII_EDGE_RE      = re.compile(r"[A-Za-z0-9]")                    # ASCII 境界判定用

# ===== [REPLACE] 改行処理：座標・文字サイズ・句点・タイトル考慮 =====
def soft_join_lines_lang_aware(s: str) -> str:
    """
    改行を賢く処理するが、以下の条件で改行を保持する:
    - 英字1トークン行（SSIDなど）
    - 数字だけの行（2025など）
    - 行末が「。」や「！」「？」で終わる場合（文章の別れ際）
    - 特殊マーカー [KEEP_NL] が含まれる場合（座標差・サイズ差ヒント）
    """
    import re

    KEEP = "[KEEP_NL]"

    # 既存の正規表現があれば利用、無ければ安全フォールバック
    try:
        ALPHA_RE = _ALPHA_TOKEN_RE
    except NameError:
        # 英字1トークン（英数/_-. / を許可）※ハイフンはクラス末尾に置くかエスケープで安全化
        ALPHA_RE = re.compile(r"^[A-Za-z][A-Za-z0-9._/\-]*$")

    try:
        NUM_RE = _NUM_ONLY_LINE_RE
    except NameError:
        # 数字だけの行（符号/桁区切り/コロン/スラッシュ/ハイフン等）
        NUM_RE = re.compile(r"^[\+\-]?\d[\d\s,./:\-–—]*$")

    if not isinstance(s, str):
        return s

    # 空白整形
    s = s.replace("\u00A0", " ").replace("\u3000", " ").replace("\r", "")
    lines = s.split("\n")
    if len(lines) <= 1:
        return re.sub(r"[ ]{2,}", " ", s)

    def _is_alpha_single_token(line: str) -> bool:
        t = (line or "").strip()
        return bool(t) and bool(ALPHA_RE.fullmatch(t))

    def _is_numeric_line(line: str) -> bool:
        t = (line or "").strip()
        return bool(t) and bool(NUM_RE.fullmatch(t))

    out: list[str] = []
    for i, cur in enumerate(lines):
        nxt = lines[i + 1] if i + 1 < len(lines) else ""
        cur_is_keep_only = (cur.strip() == KEEP)          # ★ 追加
        cur_has_keep = (KEEP in cur)
        nxt_has_keep = (KEEP in nxt)


        if cur_is_keep_only:
            continue


        cur_clean = cur.replace(KEEP, "").strip()
        out.append(cur_clean)

        if i == len(lines) - 1:
            break

        # 改行保持条件
        keep_nl = False
        if cur_has_keep or nxt_has_keep:
            keep_nl = True
        elif _is_alpha_single_token(cur_clean) or _is_alpha_single_token(nxt.replace(KEEP, "").strip()):
            keep_nl = True
        elif _is_numeric_line(cur_clean) or _is_numeric_line(nxt.replace(KEEP, "").strip()):
            keep_nl = True
        elif cur_clean.endswith(("。", "！", "?", "？")):
            keep_nl = True

        if keep_nl:
            out.append("\n")
            continue

        # 英単語の行末ハイフン → 次行の英単語へ連結（ハイフン除去）
        nxt_clean = nxt.replace(KEEP, "").strip()
        if cur_clean.endswith("-") and re.match(r"^[A-Za-z]", nxt_clean or ""):
            out[-1] = cur_clean[:-1]  # 末尾 '-' を落として完全連結
            continue

        # ASCIIどうしの境界は空白1個で連結
        if re.search(r"[A-Za-z0-9]$", cur_clean or "") and re.match(r"^[A-Za-z0-9]", nxt_clean or ""):
            out.append(" ")
            continue

        # その他は改行除去（= ベタ結合）
        continue

    joined = "".join(out)
    joined = re.sub(r"[ ]{2,}", " ", joined)
    return joined
# ===== [/REPLACE] =====

@dataclass
class PageData:
    pdf: str
    page: int
    text_join: str

# ===== [REPLACE] 完全対応版 _compose_text_from_rawdict（KEEP_NL挿入） =====
def _compose_text_from_rawdict(
    page,
    *,
    vertical_strategy: str = "auto",
    drop_marginal: bool = False,
    margin_ratio: float = 0.045,
    col_bin_ratio_h: float = 0.08,
    col_bin_ratio_v: float = 0.06,
    span_space_ratio_h: float = 0.006,
    cell_tab_ratio_h: float = 0.018,
    char_space_ratio_v: float = 0.004,
    char_tab_ratio_v: float = 0.012,
    drop_ruby: bool = True,
    ruby_size_ratio: float = 0.60,
    font_size_tolerance: float = 0.15,  # 15%に拡大
    debug_font_changes: bool = False,  # デバッグモード追加
) -> str:
    """
    PDF rawdictを座標ベースでテキスト復元
    縦組みテキストでフォントサイズが変化した箇所では改行を保持
    """
    import statistics
    from collections import defaultdict

    KEEP = "[KEEP_NL]"

    # 1) rawdict 取得
    try:
        rd = page.get_text("rawdict") or {}
    except Exception:
        rd = {}
    blocks = rd.get("blocks", []) if isinstance(rd, dict) else []
    if not blocks:
        return ""

    width = float(page.rect.width)
    height = float(page.rect.height)

    # しきい値 (絶対値化)
    span_space_h = max(1.5, width * float(span_space_ratio_h or 0.0))
    cell_tab_h = max(3.0, width * float(cell_tab_ratio_h or 0.0))
    char_space_v = max(1.0, height * float(char_space_ratio_v or 0.0))
    char_tab_v = max(2.0, height * float(char_tab_ratio_v or 0.0))
    col_bin_w_h = max(16.0, width * float(col_bin_ratio_h or 0.0))
    col_bin_w_v = max(12.0, width * float(col_bin_ratio_v or 0.0))
    top_margin = height * float(margin_ratio or 0.0)
    bot_margin = height * (1.0 - float(margin_ratio or 0.0))

    # ヘルパー: span→テキスト
    def _span_text(sp):
        t = sp.get("text") or ""
        if t.strip():
            return t
        chs = sp.get("chars")
        if isinstance(chs, list) and chs:
            return "".join(ch.get("c", "") for ch in chs)
        return ""

    # ヘルパー: 行の縦判定
    def _is_vertical_line(ln) -> bool:
        wmode = ln.get("wmode")
        if isinstance(wmode, int) and wmode == 1:
            return True
        dirv = ln.get("dir")
        if isinstance(dirv, (list, tuple)) and len(dirv) >= 2:
            try:
                dx, dy = float(dirv[0]), float(dirv[1])
                if abs(dy) > abs(dx):
                    return True
            except Exception:
                pass
        chars = []
        for sp in ln.get("spans", []):
            size = float(sp.get("size", 0.0) or 0.0)
            chs = sp.get("chars")
            if isinstance(chs, list) and chs:
                for ch in chs:
                    c = ch.get("c", "")
                    x0, y0, x1, y1 = map(float, ch.get("bbox", (0, 0, 0, 0)))
                    chars.append((c, x0, y0, x1, y1, size))
            else:
                t = sp.get("text") or ""
                if not t:
                    continue
                x0, y0, x1, y1 = map(float, sp.get("bbox", (0, 0, 0, 0)))
                n = max(1, len(t))
                w = (x1 - x0) / n if n else 0.0
                for i, c in enumerate(t):
                    cx0 = x0 + i * w
                    cx1 = cx0 + w
                    chars.append((c, cx0, y0, cx1, y1, size))
        if len(chars) < 2:
            return False
        xs = [(x0 + x1) / 2.0 for _, x0, _, x1, _, _ in chars]
        ys = [(y0 + y1) / 2.0 for _, _, y0, _, y1, _ in chars]
        try:
            import statistics as stats
            qx = stats.quantiles(xs, n=4)
            qy = stats.quantiles(ys, n=4)
            iqr_x = qx[2] - qx[0]
            iqr_y = qy[2] - qy[0]
        except Exception:
            mx = sum(xs) / len(xs)
            my = sum(ys) / len(ys)
            iqr_x = sum((x - mx) * (x - mx) for x in xs) / len(xs)
            iqr_y = sum((y - my) * (y - my) for y in ys) / len(ys)
        return (iqr_y > iqr_x * 1.6)

    def _line_x_range(ln):
        xs = []
        for sp in ln.get("spans", []):
            chs = sp.get("chars")
            if isinstance(chs, list) and chs:
                for ch in chs:
                    x0, _, x1, _ = map(float, ch.get("bbox", (0, 0, 0, 0)))
                    xs += [x0, x1]
            else:
                x0, _, x1, _ = map(float, sp.get("bbox", (0, 0, 0, 0)))
                xs += [x0, x1]
        if not xs:
            return None
        return (min(xs), max(xs))

    # ブロック収集
    blist = []
    for b in blocks:
        if b.get("type", 1) != 0:
            continue
        x0, y0, x1, y1 = map(float, b.get("bbox", (0, 0, 0, 0)))
        if drop_marginal and (y1 <= top_margin or y0 >= bot_margin):
            continue
        lines = b.get("lines", [])
        if not lines:
            continue
        ln_items = []
        v_cnt, h_cnt = 0, 0
        sizes = []
        for ln in lines:
            for sp in ln.get("spans", []):
                sz = float(sp.get("size", 0.0) or 0.0)
                if sz > 0:
                    sizes.append(sz)
            is_v = _is_vertical_line(ln)
            ln_items.append((ln, is_v))
            v_cnt += int(bool(is_v))
            h_cnt += int(not is_v)
        orient = "v" if v_cnt > h_cnt else "h"
        font_med = statistics.median(sizes) if sizes else 0.0
        blist.append({
            "bbox": (x0, y0, x1, y1),
            "x0": x0, "y0": y0, "x1": x1, "y1": y1,
            "lines_oriented": ln_items,
            "orient": orient,
            "font_med": font_med,
            "page_num": page.number,
        })
    if not blist:
        return ""

    # ページ優勢方向
    v_total = sum(1 for tb in blist if tb["orient"] == "v")
    h_total = len(blist) - v_total
    if vertical_strategy == "force_v":
        page_mode = "v"
    elif vertical_strategy == "force_h":
        page_mode = "h"
    else:
        page_mode = "v" if v_total > h_total else "h"

    # === 横ブロックの統合処理 ===
    def _line_info_horizontal(ln, span_space_h: float, cell_tab_h: float) -> dict:
        spans = ln.get("spans", [])
        if not spans:
            return {}
        spans = sorted(spans, key=lambda sp: float(sp.get("bbox", [0, 0, 0, 0])[0]))
        parts, font_sizes = [], []
        prev_x1 = None
        x_min = x_max = y_top = y_bottom = None
        for sp in spans:
            txt = (_span_text(sp) or "").replace("\u00A0", " ").strip()
            if not txt:
                continue
            x0, y0, x1, y1 = map(float, sp.get("bbox", (0, 0, 0, 0)))
            x_min = x0 if x_min is None else min(x_min, x0)
            x_max = x1 if x_max is None else max(x_max, x1)
            y_top = y0 if y_top is None else min(y_top, y0)
            y_bottom = y1 if y_bottom is None else max(y_bottom, y1)
            fs = float(sp.get("size", 0.0) or 0.0)
            if fs > 0:
                font_sizes.append(fs)
            if prev_x1 is None:
                parts.append(txt)
            else:
                gap = x0 - prev_x1
                if gap >= cell_tab_h:
                    parts.append("\t")
                    parts.append(txt)
                elif gap >= span_space_h:
                    parts.append(" ")
                    parts.append(txt)
                else:
                    parts.append(txt)
            prev_x1 = x1
        text = "".join(parts).strip()
        if not text:
            return {}
        font_size = font_sizes[0] if font_sizes else 0.0
        line_height = (y_bottom - y_top) if (y_bottom is not None and y_top is not None) else 0.0
        return {
            "text": text,
            "x0": x_min,
            "x1": x_max,
            "top": y_top,
            "height": line_height,
            "font_size": font_size,
        }

    all_h_lines = []
    for tb in blist:
        if tb["orient"] != "h":
            continue
        for ln, is_v in tb["lines_oriented"]:
            if is_v:
                continue
            line_info = _line_info_horizontal(ln, span_space_h, cell_tab_h)
            if line_info:
                all_h_lines.append(line_info)

    def _merge_all_horizontal_lines(line_dicts: list[dict]) -> str:
        import re
        KEEP = "[KEEP_NL]"

        def _is_bullet_line(text: str) -> bool:
            if not text:
                return False
            bullet_patterns = [
                r'^[・•●○◆◇■□▪▫★☆→⇒›»]',
                r'^[①-⑳]',
                r'^[⑴-⒇]',
                r'^[\d]+[.．)）]\s',
                r'^[a-zA-Z][.．)）]\s',
                r'^[-−‐–—]\s',
                r'^[*＊]\s',
                r'^[+＋]\s',
            ]
            for pattern in bullet_patterns:
                if re.match(pattern, text.strip()):
                    return True
            return False

        def _is_merge_exempt(text: str) -> bool:
            s = re.sub(r'\s+', '', text or '')
            if not s:
                return False
            if re.fullmatch(r'[0-9]+', s):
                return True
            if len(s) == 1 and re.fullmatch(r'[A-Za-z]|[ -/:-@\[-`{-~]', s):
                return True
            return False

        def _smart_join(a: str, b: str) -> str:
            if not a:
                return b or ''
            if not b:
                return a
            if re.search(r'[A-Za-z0-9]$', a) and re.search(r'^[A-Za-z0-9]', b):
                return a + ' ' + b
            return a + b

        if not line_dicts:
            return ""
        
        merged: list[str] = []
        buffer = line_dicts[0]["text"]
        
        for i in range(1, len(line_dicts)):
            prev_line = line_dicts[i - 1]
            curr_line = line_dicts[i]

            if _is_merge_exempt(prev_line["text"]) or _is_merge_exempt(curr_line["text"]):
                merged.append(buffer + KEEP)
                buffer = curr_line["text"]
                continue

            if _is_bullet_line(curr_line["text"]):
                merged.append(buffer + KEEP)
                buffer = curr_line["text"]
                continue

            x_overlap = not (prev_line["x1"] < curr_line["x0"] or curr_line["x1"] < prev_line["x0"])
            y_condition = (curr_line["top"] > prev_line["top"]) and (
                (curr_line["top"] - prev_line["top"]) <= 2.0 * max(1.0, prev_line["height"])
            )
            prev_fs = float(prev_line.get("font_size") or 0.0)
            curr_fs = float(curr_line.get("font_size") or 0.0)
            TOL = 0.10
            font_diff_ok = (
                (prev_fs == 0.0 and curr_fs == 0.0)
                or (prev_fs > 0.0 and curr_fs > 0.0 and (1 - TOL) <= (curr_fs / prev_fs) <= (1 + TOL))
            )

            if x_overlap and y_condition and font_diff_ok:
                buffer = _smart_join(buffer, curr_line["text"])
            else:
                merged.append(buffer + KEEP)
                buffer = curr_line["text"]

        merged.append(buffer)
        return "\n".join(merged)

    parts_h_text = _merge_all_horizontal_lines(all_h_lines) if all_h_lines else ""

    # === 縦行の復元（フォントサイズ変化で改行保持・強化版） ===
    def _build_line_text_vertical(ln) -> str:
        import statistics as stats

        chars = []
        for sp in ln.get("spans", []):
            size = float(sp.get("size", 0.0) or 0.0)
            chs = sp.get("chars")
            if isinstance(chs, list) and chs:
                for ch in chs:
                    c = ch.get("c", "")
                    x0, y0, x1, y1 = map(float, ch.get("bbox", (0, 0, 0, 0)))
                    chars.append((c, x0, y0, x1, y1, size))
            else:
                t = sp.get("text") or ""
                if not t:
                    continue
                x0, y0, x1, y1 = map(float, sp.get("bbox", (0, 0, 0, 0)))
                n = max(1, len(t))
                h = (y1 - y0) / n if n else 0.0
                for i, c in enumerate(t):
                    ly0 = y0 + i * h
                    ly1 = ly0 + h
                    chars.append((c, x0, ly0, x1, ly1, size))
        if not chars:
            return ""
        
        if debug_font_changes:
            print(f"[縦行] 収集した文字数: {len(chars)}")
            sizes_before = [sz for *_, sz in chars if sz > 0]
            if sizes_before:
                print(f"[縦行] サイズ範囲（ルビ除去前）: {min(sizes_before):.1f} - {max(sizes_before):.1f}")

        # ルビ抑制
        med = 0.0
        if drop_ruby:
            sizes = [sz for *_, sz in chars if sz > 0]
            med = (stats.median(sizes) if sizes else 0.0)
            if debug_font_changes and med > 0:
                print(f"[縦行] フォント中央値: {med:.1f}, 閾値: {med * float(ruby_size_ratio):.1f}")
            if med > 0:
                before_count = len(chars)
                chars = [t for t in chars if not (t[-1] and t[-1] < med * float(ruby_size_ratio or 0.60))]
                if debug_font_changes:
                    print(f"[縦行] ルビ除去: {before_count} -> {len(chars)} 文字")
        if not chars:
            return ""
        
        if debug_font_changes:
            sizes_after = [sz for *_, sz in chars if sz > 0]
            if sizes_after:
                print(f"[縦行] サイズ範囲（ルビ除去後）: {min(sizes_after):.1f} - {max(sizes_after):.1f}")
                # 最初の10文字のサイズを表示
                for i, (c, _, _, _, _, sz) in enumerate(chars[:10]):
                    print(f"  文字[{i}]: '{c}' サイズ={sz:.1f}")

        # Y昇順ソート
        chars.sort(key=lambda t: t[2])
        parts = []
        prev_y1 = None
        prev_size = None
        
        for c, x0, y0, x1, y1, sz in chars:
            if prev_y1 is None:
                parts.append(c)
                prev_size = sz
            else:
                gap = y0 - prev_y1
                
                # フォントサイズ変化チェック（両方が正の値の時のみ）
                size_changed = False
                if prev_size > 0.1 and sz > 0.1:  # 極小値を除外
                    size_ratio = sz / prev_size
                    if size_ratio < (1 - font_size_tolerance) or size_ratio > (1 + font_size_tolerance):
                        size_changed = True
                        if debug_font_changes:
                            print(f"[縦] サイズ変化検出: {prev_size:.1f} -> {sz:.1f} (比率: {size_ratio:.2f}) 文字: '{c}'")
                
                if size_changed:
                    parts.append("\n")
                    parts.append(c)
                elif gap >= char_tab_v:
                    parts.append("\t")
                    parts.append(c)
                elif gap >= char_space_v:
                    parts.append(" ")
                    parts.append(c)
                else:
                    parts.append(c)
                
                prev_y1 = y1
                prev_size = sz
        
        return "".join(parts).strip()

    # === strict_y モード ===
    if page_mode == "v" and vertical_strategy == "strict_y":
        vchars = []
        for tb in blist:
            for ln, is_v in tb["lines_oriented"]:
                if not is_v:
                    continue
                for sp in ln.get("spans", []):
                    size = float(sp.get("size", 0.0) or 0.0)
                    chs = sp.get("chars")
                    if isinstance(chs, list) and chs:
                        for ch in chs:
                            c = ch.get("c", "")
                            x0, y0, x1, y1 = map(float, ch.get("bbox", (0, 0, 0, 0)))
                            cx = (x0 + x1) / 2.0
                            vchars.append((c, cx, y0, y1, size))
                    else:
                        t = sp.get("text") or ""
                        if not t:
                            continue
                        x0, y0, x1, y1 = map(float, sp.get("bbox", (0, 0, 0, 0)))
                        n = max(1, len(t))
                        h = (y1 - y0) / n if n else 0.0
                        cx = (x0 + x1) / 2.0
                        for i, c in enumerate(t):
                            ly0 = y0 + i * h
                            ly1 = ly0 + h
                            vchars.append((c, cx, ly0, ly1, size))
        
        if not vchars:
            return parts_h_text or ""

        # ルビ除去
        if drop_ruby:
            sizes = [sz for *_, sz in vchars if sz > 0]
            med = statistics.median(sizes) if sizes else 0.0
            if debug_font_changes and med > 0:
                print(f"[strict_y] 全体フォント中央値: {med:.1f}, ルビ閾値: {med * float(ruby_size_ratio):.1f}")
                print(f"[strict_y] ルビ除去前の文字数: {len(vchars)}")
            if med > 0:
                vchars = [t for t in vchars if not (t[-1] and t[-1] < med * float(ruby_size_ratio or 0.60))]
            if debug_font_changes:
                print(f"[strict_y] ルビ除去後の文字数: {len(vchars)}")
                if vchars:
                    sizes_after = [sz for *_, sz in vchars if sz > 0]
                    if sizes_after:
                        print(f"[strict_y] サイズ範囲: {min(sizes_after):.1f} - {max(sizes_after):.1f}")
        if not vchars:
            return parts_h_text or ""

        # X で列クラスタ
        bin_w = max(8.0, width * 0.020)
        cols = defaultdict(list)
        for c, cx, y0, y1, sz in vchars:
            key = int(cx // bin_w)
            cols[key].append((c, cx, y0, y1, sz))

        # 各列を Y 昇順で連結（フォントサイズ変化で改行）
        out_cols = []
        prev_col_avg_size = None  # 前の列の平均フォントサイズ
        prev_col_x = None  # 前の列のX座標（列間の距離チェック用）
        
        sorted_col_keys = sorted(cols.keys(), reverse=True)
        
        for col_idx, k in enumerate(sorted_col_keys):
            col = cols[k]
            col.sort(key=lambda t: t[2])  # y0
            
            # この列の平均フォントサイズとX座標を計算
            col_sizes = [sz for *_, sz in col if sz > 0]
            current_col_avg_size = sum(col_sizes) / len(col_sizes) if col_sizes else 0.0
            
            # 列のX座標（中央値）
            col_x_coords = [cx for _, cx, *_ in col]
            current_col_x = sum(col_x_coords) / len(col_x_coords) if col_x_coords else 0.0
            
            if debug_font_changes and col_idx < 3:
                print(f"\n[strict_y列{col_idx}] 列の文字数: {len(col)}, X座標: {current_col_x:.1f}")
                if col_sizes:
                    print(f"[strict_y列{col_idx}] サイズ範囲: {min(col_sizes):.1f} - {max(col_sizes):.1f}")
                    print(f"[strict_y列{col_idx}] 平均サイズ: {current_col_avg_size:.1f}")
                    unique_sizes = sorted(set(col_sizes))
                    print(f"[strict_y列{col_idx}] ユニークなサイズ: {[f'{s:.1f}' for s in unique_sizes]}")
            
            parts = []
            prev_y1 = None
            prev_size = None
            size_change_count = 0
            
            for idx, (c, cx, y0, y1, sz) in enumerate(col):
                if prev_y1 is None:
                    parts.append(c)
                    prev_size = sz
                else:
                    gap = y0 - prev_y1
                    
                    # フォントサイズ変化チェック
                    size_changed = False
                    if prev_size > 0.1 and sz > 0.1:
                        size_ratio = sz / prev_size
                        if size_ratio < (1 - font_size_tolerance) or size_ratio > (1 + font_size_tolerance):
                            size_changed = True
                            size_change_count += 1
                            if debug_font_changes and col_idx < 3 and size_change_count <= 5:
                                print(f"[strict_y列{col_idx}] サイズ変化#{size_change_count}: {prev_size:.1f} -> {sz:.1f} (比率: {size_ratio:.2f}, {(size_ratio-1)*100:.1f}%) 文字[{idx}]: '{c}' gap={gap:.1f}")
                    
                    if size_changed:
                        parts.append("\n")
                        parts.append(c)
                    elif gap >= char_tab_v:
                        parts.append("\t")
                        parts.append(c)
                    elif gap >= char_space_v:
                        parts.append(" ")
                        parts.append(c)
                    else:
                        parts.append(c)
                    
                    prev_y1 = y1
                    prev_size = sz
            
            if debug_font_changes and col_idx < 3:
                print(f"[strict_y列{col_idx}] 検出したサイズ変化の回数: {size_change_count}")
            
            t = "".join(parts).strip()
            if t:
                should_add_keep = False
                
                # 1. 前の列との平均サイズ差をチェック
                if prev_col_avg_size is not None and current_col_avg_size > 0:
                    size_ratio = current_col_avg_size / prev_col_avg_size
                    if size_ratio < (1 - font_size_tolerance) or size_ratio > (1 + font_size_tolerance):
                        should_add_keep = True
                        if debug_font_changes and col_idx < 5:
                            print(f"[strict_y] 列{col_idx-1}→列{col_idx}間でサイズ変化: {prev_col_avg_size:.1f} -> {current_col_avg_size:.1f} (比率: {size_ratio:.2f})")
                
                # 2. 列が1つ以上空いているかチェック（X座標の差が大きい）
                if prev_col_x is not None:
                    col_gap = abs(prev_col_x - current_col_x)
                    # ビン幅の2倍以上離れている場合は列が空いていると判断
                    if col_gap > bin_w * 2.0:
                        should_add_keep = True
                        if debug_font_changes and col_idx < 5:
                            print(f"[strict_y] 列{col_idx-1}→列{col_idx}間で列アキ検出: X差={col_gap:.1f} (閾値={bin_w * 2.5:.1f})")
                
                # 3. 句点で終わっているかチェック
                if out_cols and (out_cols[-1].rstrip().endswith('。') or out_cols[-1].rstrip().endswith('.')):
                    # 前の列が句点で終わっている場合
                    if not out_cols[-1].endswith(KEEP):
                        out_cols[-1] = out_cols[-1] + KEEP
                        if debug_font_changes and col_idx < 5:
                            print(f"[strict_y] 列{col_idx-1}が句点で終了 -> KEEPマーカー追加")
                
                # KEEPマーカーを前の列に追加
                if should_add_keep and out_cols and not out_cols[-1].endswith(KEEP):
                    out_cols[-1] = out_cols[-1] + KEEP
                
                out_cols.append(t)
                prev_col_avg_size = current_col_avg_size
                prev_col_x = current_col_x

        chunks = []
        if out_cols:
            # 列間の結合（KEEPマーカーがある場合は改行を保持）
            if debug_font_changes:
                print(f"\n[strict_y] 出力列数: {len(out_cols)}")
                for i, col_text in enumerate(out_cols[:5]):
                    has_keep = KEEP in col_text
                    print(f"[strict_y] 列{i}: KEEPあり={has_keep}, 長さ={len(col_text)}, 内容={repr(col_text[:50] if len(col_text) > 50 else col_text)}...末尾={repr(col_text[-20:])}")
            
            # 列を連結
            result_parts = []
            for i, col_text in enumerate(out_cols):
                result_parts.append(col_text)
                # KEEPマーカーがある場合はその後に改行を追加（KEEPマーカー自体は残す）
                # KEEPマーカーがない場合は次の列と直接連結
                if i < len(out_cols) - 1:  # 最後の列でない場合
                    if col_text.endswith(KEEP):
                        # KEEPで終わる場合は改行を追加
                        result_parts.append("\n")
                    # KEEPで終わらない場合は何も追加しない（直接連結）
            
            combined_text = "".join(result_parts)
            chunks.append(combined_text)
            
            if debug_font_changes:
                print(f"\n[strict_y] 結合後のテキスト（最初の600文字）:")
                print(repr(combined_text[:600]))
        
        if parts_h_text:
            chunks.append(parts_h_text)
        
        final_text = "\n\n".join([c for c in chunks if c])
        
        return final_text

    # === 既存の縦クラスタ(列単位) ===
    cols_v = defaultdict(list)
    for tb in blist:
        if tb["orient"] == "v":
            cx = (tb["x0"] + tb["x1"]) / 2.0
            key = int(cx // col_bin_w_v)
            cols_v[key].append(tb)

    def _build_block_text_v(tb) -> str:
        texts = []
        for ln, is_v in tb["lines_oriented"]:
            if not is_v:
                continue
            t = _build_line_text_vertical(ln)
            if t:
                texts.append(t)
        return "\n".join(texts)

    parts_v = []
    for ck in sorted(cols_v.keys(), reverse=True):
        col_blocks = sorted(cols_v[ck], key=lambda b: (b["y0"], b["x0"]))
        col_texts = []
        for b in col_blocks:
            t = _build_block_text_v(b)
            if t:
                col_texts.append(t)
        if col_texts:
            parts_v.append("\n".join(col_texts))

    # ページ合成
    if page_mode == "v":
        chunks = []
        if parts_v:
            chunks.append("\n".join(parts_v))
        if parts_h_text:
            chunks.append(parts_h_text)
        return "\n\n".join([c for c in chunks if c])
    else:
        chunks = []
        if parts_h_text:
            chunks.append(parts_h_text)
        if parts_v:
            chunks.append("\n".join(parts_v))
        return "\n\n".join([c for c in chunks if c])


# ===== [PATCH] extract_pages: v_strategy 引数を追加して _compose に受け渡し =====
def extract_pages(pdf_path: str, *, flatten_for_nlp: bool = False, v_strategy: str = "auto") -> List[PageData]:


    def _remove_invisible_chars(s: str) -> str:
        if not isinstance(s, str):
            return s
        # 制御文字（Cc, Cf）を削除。ただし改行・タブは残す
        return "".join(ch for ch in s if not (
            unicodedata.category(ch) in ("Cc", "Cf") and ch not in ("\n", "\t")
        ))


    pages: List[PageData] = []
    doc = fitz.open(pdf_path)
    try:
        for i in range(len(doc)):
            page = doc[i]
            # ① rawdict（ここで縦横判定＋縦復元を実施）
            text_blocked = _compose_text_from_rawdict(
                page,
                vertical_strategy=v_strategy,  # ← ここを引数で指定可能に
                drop_marginal=False,
                margin_ratio=0.05,
                debug_font_changes=False
            )
            # （以下は従来どおりのフォールバック）...
            if not text_blocked.strip():
                try:
                    blks = page.get_text("blocks")
                except Exception:
                    blks = []
                if blks:
                    blks_sorted = sorted(blks, key=lambda t: (t[0], t[1]))
                    parts = []
                    for b in blks_sorted:
                        if len(b) >= 5:
                            txt = (b[4] or "").rstrip()
                            if txt:
                                parts.append(txt)
                    text_blocked = "\n".join(parts)
            if not text_blocked.strip():
                text_blocked = page.get_text("text") or ""

            text_blocked = _remove_invisible_chars(text_blocked)    

            text_final = soft_join_lines_lang_aware(text_blocked) if flatten_for_nlp else text_blocked
            pages.append(PageData(os.path.basename(pdf_path), i + 1, text_final))
    finally:
        doc.close()
    return pages
# ===== [/PATCH] =====


# ===== [REPLACE] レイアウト分割を無効化（常に全体1本） =====
def iter_nlp_segments(text: str):
    """
    以前は改行・タブで分割していたが、
    版面由来の分割を一切やめ、テキスト全体をそのまま返す。
    """
    if isinstance(text, str) and text:
        yield text
# ===== [/REPLACE] =====


# ===== GiNZA: 軽量モデル + 文節専用で初期化（差し替え） =====
HAS_GINZA = False
GINZA_NLP = None

# 文節抽出に必要な最小構成（parser + bunsetu_recognizer）
GINZA_ENABLED_COMPONENTS = ("parser", "bunsetu_recognizer")

try:
    import spacy, ginza
    # ▼ モデル選択：環境変数 GINZA_MODEL で上書き可（既定は軽量の ja_ginza）
    #   - ja_ginza: CPU向けの標準モデル（軽量）[1](https://pypi.org/project/ja-ginza/)
    #   - ja_ginza_electra: Transformerベースで重い（必要時のみ）[2](https://pypi.org/project/ja-ginza-electra/)[3](https://huggingface.co/megagonlabs/transformers-ud-japanese-electra-base-ginza)
    model_name = os.environ.get("GINZA_MODEL", "ja_ginza").strip()

    # ▼ 不要コンポーネントをロード時に無効化（文節には不要）
    #   - ner, attribute_ruler, compound_splitter をデフォルトで停止
    #     （parser と bunsetu_recognizer は動かす）[1](https://pypi.org/project/ja-ginza/)
    disable = os.environ.get(
        "GINZA_DISABLE", "ner,attribute_ruler,compound_splitter"
    ).split(",")
    disable = [d.strip() for d in disable if d.strip()]

    GINZA_NLP = spacy.load(model_name, disable=disable)

    # （任意）Sudachi の分割モードをCに固定：複合語優先
    try:
        ginza.set_split_mode(GINZA_NLP, "C")
    except Exception:
        pass

    HAS_GINZA = True
except Exception:
    HAS_GINZA = False
    GINZA_NLP = None
# ===== GiNZA 初期化ここまで =====

# ===== [NEW] GiNZA 文節抽出：文章リストから処理 =====
def ginza_bunsetsu_from_sentences(sentences: list[str]) -> list[str]:
    """
    GiNZAで文節を抽出するが、入力は文単位のリスト。
    - 各文をGiNZAに渡してbunsetsu_spansで抽出
    - 長さフィルタは従来通り（1..120）
    """
    if not sentences or not HAS_GINZA or GINZA_NLP is None:
        return []

    chunks: list[str] = []
    try:
        pipes_to_enable = [p for p in ["parser", "bunsetu_recognizer"] if p in GINZA_NLP.pipe_names]
        with GINZA_NLP.select_pipes(enable=pipes_to_enable):
            for doc in GINZA_NLP.pipe(sentences, n_process=1, batch_size=1):
                for sent in doc.sents:
                    for sp in ginza.bunsetu_spans(sent):
                        s = re.sub(r"[\t\r\n]+", " ", sp.text).strip()
                        if 1 <= len(s) <= 120:
                            chunks.append(s)
    except Exception:
        return []
    return chunks

# ===== [REPLACE] GiNZA 文節抽出：全文1本で処理 =====
def ginza_bunsetsu_chunks(text: str) -> list[str]:
    """
    GiNZA を用いた文節抽出。
    - レイアウト分割はせず、テキスト全体をそのまま1本で処理
    - 句の長さフィルタのみ従来踏襲（1..120）
    """
    ok = HAS_GINZA and (GINZA_NLP is not None)
    if not ok or not isinstance(text, str) or not text:
        return []
    chunks: list[str] = []
    try:
        pipes_to_enable = [p for p in GINZA_ENABLED_COMPONENTS if p in GINZA_NLP.pipe_names]
        with GINZA_NLP.select_pipes(enable=pipes_to_enable):
            for doc in GINZA_NLP.pipe([text], n_process=1, batch_size=1):
                for sent in doc.sents:
                    for sp in ginza.bunsetu_spans(sent):
                        s = re.sub(r"[\t\r\n]+", " ", sp.text).strip()
                        if 1 <= len(s) <= 120:
                            chunks.append(s)
    except Exception:
        return []
    return chunks
# ===== [/REPLACE] =====

def ginza_sentence_units(text: str, *, max_len: int = 300) -> list[str]:
    """
    GiNZAで文抽出。GiNZA結果をさらに「。」や「！」で強制分割。
    """
    import re
    if not isinstance(text, str) or not text:
        return []

    BULLET_SPLIT_RE = re.compile(r"[●•◆▶■□]+")
    HEADER_BRACKET_RE = re.compile(r"^【.+?】")
    UPPER_TITLE_RE = re.compile(r"^[A-Z][A-Z0-9 ()/_-]*$")
    FORCE_SPLIT_RE = re.compile(r"(?<=[。！？!?])")  # ← 強制分割用

    lines = text.splitlines(keepends=True)
    pre_chunks = []
    for ln in lines:
        if not ln.strip() and ln != "\n":
            continue
        had_nl = ln.endswith("\n")
        body = ln[:-1] if had_nl else ln
        if HEADER_BRACKET_RE.match(body) or UPPER_TITLE_RE.fullmatch(body):
            pre_chunks.append(body + ("\n" if had_nl else ""))
            continue
        if BULLET_SPLIT_RE.search(body):
            parts = [p.strip() for p in BULLET_SPLIT_RE.split(body) if p.strip()]
            for j, p in enumerate(parts):
                if j == len(parts) - 1 and had_nl:
                    pre_chunks.append(p + "\n")
                else:
                    pre_chunks.append(p)
        else:
            pre_chunks.append(body + ("\n" if had_nl else ""))

    out = []
    ok = HAS_GINZA and (GINZA_NLP is not None)
    if ok:
        try:
            pipes_to_enable = [p for p in ["sentencizer", "parser", "bunsetu_recognizer"] if p in GINZA_NLP.pipe_names]
            with GINZA_NLP.select_pipes(enable=pipes_to_enable):
                for doc in GINZA_NLP.pipe(pre_chunks, n_process=1, batch_size=1):
                    for sent in doc.sents:
                        # GiNZA結果をさらに「。」で分割
                        for part in FORCE_SPLIT_RE.split(sent.text):
                            part = part.replace("\n", "").strip()
                            if part and (1 <= len(part) <= max_len):
                                out.append(part)
        except Exception:
            # フォールバック
            for chunk in pre_chunks:
                for s in FORCE_SPLIT_RE.split(chunk):
                    s = s.replace("\n", "").strip()
                    if s and (1 <= len(s) <= max_len):
                        out.append(s)
    else:
        # GiNZAなしフォールバック
        for chunk in pre_chunks:
            for s in FORCE_SPLIT_RE.split(chunk):
                s = s.replace("\n", "").strip()
                if s and (1 <= len(s) <= max_len):
                    out.append(s)
    return out

# ===== [ADD] GiNZA 1-pass: 文 + 文節（細切れ）を同時抽出 =====
from typing import Tuple

def ginza_sentence_and_bunsetsu_onepass(
    text: str,
    *,
    sent_max_len: int = 300
) -> Tuple[list[str], list[str]]:
    """
    GiNZA を 1パスで実行し、同一の Doc から
      - 文（強制分割あり）
      - 文節（＋細切れの POS ルール適用）
    を同時に抽出します。
    """
    import re
    if not isinstance(text, str) or not text:
        return [], []

    # 既存 ginza_sentence_units の前処理ロジックを移植（ヘッダ/箇条書き対応）
    BULLET_SPLIT_RE = re.compile(r"[●•◆▶■□]+")
    HEADER_BRACKET_RE = re.compile(r"^【.+?】")
    UPPER_TITLE_RE = re.compile(r"^[A-Z][A-Z0-9 \(\)/_\-]*$")
    FORCE_SPLIT_RE = re.compile(r"(?<=[。！？!?])")  # ← GiNZA文の「。」等で強制分割

    # 行ごとの軽い前処理（既存と同じ）
    lines = text.splitlines(keepends=True)
    pre_chunks: list[str] = []
    for ln in lines:
        if not ln.strip() and ln != "\n":
            continue
        had_nl = ln.endswith("\n")
        body = ln[:-1] if had_nl else ln

        if HEADER_BRACKET_RE.match(body) or UPPER_TITLE_RE.fullmatch(body):
            pre_chunks.append(body + ("\n" if had_nl else ""))
            continue

        if BULLET_SPLIT_RE.search(body):
            parts = [p.strip() for p in BULLET_SPLIT_RE.split(body) if p.strip()]
            for j, p in enumerate(parts):
                if j == len(parts) - 1 and had_nl:
                    pre_chunks.append(p + "\n")
                else:
                    pre_chunks.append(p)
        else:
            pre_chunks.append(body + ("\n" if had_nl else ""))

    sentences_out: list[str] = []
    bunsetsu_out: list[str] = []

    ok = HAS_GINZA and (GINZA_NLP is not None)
    if ok:
        try:
            pipes_to_enable = [p for p in ["sentencizer", "parser", "bunsetu_recognizer"]
                               if p in GINZA_NLP.pipe_names]
            with GINZA_NLP.select_pipes(enable=pipes_to_enable):
                # ★ ここが“1パス”：pre_chunks をまとめて GiNZA に渡し、
                #   1回のパイプライン内で「文」と「文節」を両方回収します
                for doc in GINZA_NLP.pipe(pre_chunks, n_process=1, batch_size=1):
                    for sent in doc.sents:
                        # --- 文（強制分割） ---
                        for part in FORCE_SPLIT_RE.split(sent.text):
                            s = part.replace("\n", "").strip()
                            if s and (1 <= len(s) <= sent_max_len):
                                sentences_out.append(s)

                        # --- 文節（＋細切り POS ルール） ---
                        # 既存 ginza_bunsetsu_fine_pos の方針をインライン適用
                        for sp in ginza.bunsetu_spans(sent):
                            text_sp = "".join([t.text_with_ws for t in sp]).strip()
                            if len(text_sp) <= 15:
                                s_norm = re.sub(r"[\t\r\n]+", " ", text_sp)
                                s_norm = re.sub(r"[ ]{2,}", " ", s_norm).strip()
                                if 1 <= len(s_norm) <= 120:
                                    bunsetsu_out.append(s_norm)
                                continue

                            buf: list[str] = []

                            def flush_buf():
                                nonlocal buf
                                if not buf:
                                    return
                                part = "".join(buf).strip()
                                if len(part) >= 4:
                                    s_norm = re.sub(r"[\t\r\n]+", " ", part)
                                    s_norm = re.sub(r"[ ]{2,}", " ", s_norm).strip()
                                    if 1 <= len(s_norm) <= 120:
                                        bunsetsu_out.append(s_norm)
                                buf.clear()

                            for i, t in enumerate(sp):
                                buf.append(t.text_with_ws)
                                next_token = sp[i + 1] if i + 1 < len(sp) else None

                                # 分割条件（長めの文節のみ）
                                # 助詞/助動詞/PUNCT（、。）の扱いは既存ロジック踏襲
                                if t.pos_ in {"ADP", "PART"}:
                                    flush_buf()
                                elif t.pos_ == "AUX":
                                    if not next_token or next_token.pos_ != "AUX":
                                        flush_buf()
                                elif t.pos_ == "PUNCT" and t.text in {"、", "。"}:
                                    flush_buf()

                            if buf:
                                flush_buf()
        except Exception:
            # 失敗時は後述の MeCab フォールバックへ
            pass

    # GiNZAが無い/失敗時のフォールバック（従来どおり簡易）
    if not sentences_out and not bunsetsu_out:
        import re
        FORCE_SPLIT_RE = re.compile(r"(?<=[。！？!?])")
        # 文（簡易）
        for s in FORCE_SPLIT_RE.split(text):
            s2 = s.replace("\n", "").strip()
            if s2 and (1 <= len(s2) <= sent_max_len):
                sentences_out.append(s2)
        # 文節（MeCab既存）
        try:
            bunsetsu_out = mecab_bunsetsu_chunks(text)
        except Exception:
            bunsetsu_out = []

    return sentences_out, bunsetsu_out

# ===== [ADD] 抽出（1パス版）：ページごとに文 + 文節を同時カウント =====
from typing import Tuple

def extract_candidates_sentence_bunsetsu_onepass(
    pages: List[PageData],
    *,
    bun_min_len: int = 1,
    bun_max_len: int = 120,
    sent_min_len: int = 1,
    sent_max_len: int = 300,
    min_count: int = 1,
    top_k: int = 0
) -> Tuple[List[Tuple[str, int]], List[Tuple[str, int]]]:
    """
    ページ配列から、GiNZA 1パスで「文＋文節」を同時抽出してカウント。
    従来の「数字だけ除外（時間/日付は残す）」も適用して返します。
    """
    import re
    cnt_b = Counter()
    cnt_s = Counter()

    # 既存の「数字だけ」・時間/日付判定をそのまま利用
    try:
        NUM_RE = _NUM_ONLY_LINE_RE
    except NameError:
        NUM_DIG = r"[0-9\uFF10-\uFF19]"
        GROUP_SEP = r"[,\uFF0C \u00A0\u3000]"
        DEC_SEP = r"[.\uFF0E]"
        NUM_RE = re.compile(
            rf"""
            ^{NUM_DIG}+                         # 先頭の数字列
            (?:{GROUP_SEP}{NUM_DIG}{{3}})*      # 千区切りは3桁固定で繰り返し
            (?:{DEC_SEP}{NUM_DIG}+)?            # 小数部（任意）
            $""",
            re.VERBOSE
        )
    DATE_RE = re.compile(r"^\s*\d{4}[\-/]\d{1,2}[\-/]\d{1,2}\s*$")

    for p in pages:
        # 保存（解析ログ用）
        try:
            save_nlp_input("ginza_onepass", p.text_join, suffix=f"{p.pdf}_p{p.page}")
        except Exception:
            pass

        sents, buns = ginza_sentence_and_bunsetsu_onepass(
            p.text_join, sent_max_len=sent_max_len
        )

        # ★ 文節フィルタ用ヘルパー（関数内に追加）
        def _is_single_symbol_1char(s: str) -> bool:
            """記号/句読点だけ1文字（Unicodeカテゴリ 'P*' or 'S*'）なら True"""
            if not isinstance(s, str):
                return False
            t = s.strip()
            if len(t) != 1:
                return False
            cat = unicodedata.category(t[0])  # 'P*' 句読点, 'S*' 記号
            return cat.startswith('P') or cat.startswith('S')

        def _is_single_alpha_1char(s: str) -> bool:
            """アルファベットだけ1文字（全角含む。NFKCでASCII化して判定）なら True"""
            try:
                # プロジェクト内の nfkc がある前提（なければ unicodedata.normalize で代用可能）
                t = nfkc(s.strip())
            except NameError:
                from unicodedata import normalize
                t = normalize('NFKC', s.strip())
            return len(t) == 1 and (('A' <= t <= 'Z') or ('a' <= t <= 'z'))

        # --- 文：数字だけ除外 ---
        for s in sents:
            t = strip_leading_control_chars(s).strip()
            if not t:
                continue
            if NUM_RE.fullmatch(t):
                continue
            if _is_single_symbol_1char(t) or _is_single_alpha_1char(t):
                continue
            if sent_min_len <= len(t) <= sent_max_len:
                cnt_s[t] += 1

        # ーーー 文節の集計ループ（差し替え）ーーー
        for ch in buns:
            ch2 = strip_leading_control_chars(ch).strip()
            if not ch2:
                continue

            # ✅ sentence と同じ「数字だけ除外」ただし 時間/日付 は残す
            if NUM_RE.fullmatch(ch2):
                continue

            # ✅ 追加除外：記号だけ1文字／アルファベットだけ1文字
            if _is_single_symbol_1char(ch2) or _is_single_alpha_1char(ch2):
                continue

            # 既存の長さ条件
            if bun_min_len <= len(ch2) <= bun_max_len:
                cnt_b[ch2] += 1


    arr_b = [(w, c) for w, c in cnt_b.items() if c >= min_count]
    arr_s = [(w, c) for w, c in cnt_s.items() if c >= min_count]
    arr_b.sort(key=lambda x: (-x[1], x[0]))
    arr_s.sort(key=lambda x: (-x[1], x[0]))

    if top_k and top_k > 0:
        arr_b = arr_b[:top_k]
        arr_s = arr_s[:top_k]
    return arr_b, arr_s



# ------------------------------------------------------------
# MeCab helpers
# ------------------------------------------------------------
def ensure_mecab() -> Tuple[bool, Optional[str]]:
    global MECAB_TAGGER
    if not HAS_MECAB:
        return False, 'fugashi/unidic-lite が未インストールです。 pip install "fugashi[unidic-lite]"'
    if MECAB_TAGGER is None:
        try:
            MECAB_TAGGER = Tagger()
        except Exception as e:
            return False, f"MeCab 初期化失敗: {e}"
    return True, None

def feat_get(w, keys: List[str]) -> Optional[str]:
    f = getattr(w, "feature", None)
    if f is None:
        return None
    for k in keys:
        try:
            v = getattr(f, k)
            if v:
                return str(v)
        except Exception:
            pass
    return None

def tokenize_mecab(text: str, tagger: "Tagger"):
    out = []
    for w in tagger(text):
        surf = w.surface or ""
        pos1 = feat_get(w, ["pos1"]) or ""
        lemma = feat_get(w, ["lemma", "baseform"]) or surf
        reading = feat_get(w, ["reading", "pron", "pronBase", "kana", "kanaBase", "yomi"]) or ""
        ct = feat_get(w, ["cType", "conjType"]) or ""
        cf = feat_get(w, ["cForm", "conjForm"]) or ""
        if surf:
            out.append((surf, pos1, lemma, reading, ct, cf))
    return out

# ==== フレーズの読み正規化（複合語/文節用） ====
def phrase_reading_norm(s: str, tagger: "Tagger") -> str:
    """
    文字列 s を MeCab で再トークン化し、各語の読みを連結。
    その後 normalize_kana(drop_choon=True) で
    ・ひらがな→カタカナ
    ・カタカナ以外を除去
    ・長音「ー」を除去
    して返します（= 読みの“骨格”同士で比較できる）。
    """
    if not isinstance(s, str) or not s:
        return ""

    # 既存の tokenize_mecab を活用
    toks = tokenize_mecab(s, tagger)
    parts = []
    for surf, pos1, lemma, reading, ct, cf in toks:
        r = reading or ""
        if not r:
            # 読みが取れない場合は表層を NFKC→カタカナ化して代用
            r = hira_to_kata(nfkc(surf))
        parts.append(r)

    joined = "".join(parts)
    # ★ カタカナ以外は落ち、長音「ー」も削除される
    return normalize_kana(joined, drop_choon=True)

# ===== 読み正規化のキャッシュ（追加ブロック：貼り付け） =====
# phrase_reading_norm(s, tagger) を毎回呼ぶのは高コストなので、
# "語 → 読み(骨格)" をプロセス内でキャッシュします。
READ_NORM_CACHE: Dict[str, str] = {}

def phrase_reading_norm_cached(s: str) -> str:
    """語 s の読み(骨格)をキャッシュして返す。MeCab未使用時はフォールバック。"""
    if not isinstance(s, str) or not s:
        return ""
    v = READ_NORM_CACHE.get(s)
    if v is not None:
        return v
    ok, _ = ensure_mecab()
    if ok and MECAB_TAGGER is not None:
        v = phrase_reading_norm(s, MECAB_TAGGER)
    else:
        # フォールバック：NFKC→カタカナ→カタカナ以外除去→長音除去
        k = hira_to_kata(nfkc(s))
        v = normalize_kana(k, drop_choon=True)
    READ_NORM_CACHE[s] = v
    return v
# ===== 読みキャッシュ ここまで =====

# ===== [ADD] 長音あり読み（keepchoon）のキャッシュ =====
READ_NORM_KEEP_CACHE: Dict[str, str] = {}

def phrase_reading_norm_keepchoon_cached(s: str) -> str:
    if not isinstance(s, str) or not s:
        return ""
    v = READ_NORM_KEEP_CACHE.get(s)
    if v is not None:
        return v
    ok, _ = ensure_mecab()
    if ok and MECAB_TAGGER is not None:
        toks = tokenize_mecab(s, MECAB_TAGGER)
        parts: List[str] = []
        for surf, pos1, lemma, reading, ct, cf in toks:
            r = reading or hira_to_kata(nfkc(surf))
            parts.append(r)
        joined = "".join(parts)
        v = normalize_kana(joined, drop_choon=False)
    else:
        v = normalize_kana(hira_to_kata(nfkc(s or "")), drop_choon=False)
    READ_NORM_KEEP_CACHE[s] = v
    return v

# ===== [ADD] 読みキャッシュの一括プリウォーム =====
def prewarm_reading_caches(surfaces: List[str], *, keepchoon: bool = True):
    """MeCab を使う前に、表層群の '読み' を一括生成してキャッシュに載せる。"""
    if not surfaces:
        return
    uniq = sorted({s for s in surfaces if isinstance(s, str) and s})
    ok, _ = ensure_mecab()
    if not ok or MECAB_TAGGER is None:
        # MeCab 無しでもフォールバックを通す（コストは軽い）
        for s in uniq:
            _ = phrase_reading_norm_cached(s)
            if keepchoon:
                _ = phrase_reading_norm_keepchoon_cached(s)
        return
    # MeCab あり：1語ずつでOK（Taggerは内部でキャッシュされる）
    for s in uniq:
        _ = phrase_reading_norm_cached(s)              # 長音なし骨格
        if keepchoon:
            _ = phrase_reading_norm_keepchoon_cached(s)  # 長音あり骨格

# ===== [REPLACE] extract_candidates_regex: 候補抽出時にフィルタ =====
def extract_candidates_regex(pages: List[PageData], min_len=1, max_len=120, min_count=1):
    cnt = Counter()
    for p in pages:
        for m in TOKEN_RE.finditer(p.text_join):
            w = m.group(0)
            # ★ ここでフィルタ適用(早期除外)
            if _should_skip_for_pairing(w):
                continue
            if min_len <= len(w) <= max_len:
                cnt[w] += 1
    arr = [(w, c) for w, c in cnt.items() if c >= min_count]
    arr.sort(key=lambda x: (-x[1], x[0]))
    return arr
# ===== [/REPLACE] =====

# ------------------------------------------------------------
# Compound（連結記号でのみ結合）
# ------------------------------------------------------------

# ===== （任意）置き換え: 複合語もセグメント単位で =====
# ===== [REPLACE] MeCab 複合語（全文一発） =====
def mecab_compound_tokens_alljoin(text: str) -> List[str]:
    ok, msg = ensure_mecab()
    if not ok:
        raise RuntimeError(msg)
    if not isinstance(text, str) or not text:
        return []

    out: List[str] = []
    toks = tokenize_mecab(text, MECAB_TAGGER)

    parts: List[str] = []
    has_conn = False
    prev_type: Optional[str] = None

    def flush_local():
        nonlocal parts, has_conn, prev_type
        if parts and has_conn:
            token = "".join(parts)
            if 1 <= len(token) <= 120:
                out.append(token)
        parts = []
        has_conn = False
        prev_type = None

    for surf, pos1, lemma, reading, ct, cf in toks:
        if not surf:
            continue
        is_word = pos1.startswith(("名詞", "動詞", "形容詞", "助動詞"))
        is_conn = surf and all(ch in CONNECT_CHARS for ch in surf)

        if is_conn:
            if parts and prev_type == 'word':
                parts.append(surf)
                has_conn = True
                prev_type = 'conn'
            continue

        if is_word:
            if not parts:
                parts = [surf]
                prev_type = 'word'
            else:
                if prev_type == 'conn':
                    parts.append(surf)
                    prev_type = 'word'
                else:
                    flush_local()
                    parts = [surf]
                    prev_type = 'word'
        else:
            flush_local()

    flush_local()
    return out
# ===== [/REPLACE] =====

# ===== [REPLACE] extract_candidates_compound_alljoin: 候補抽出時にフィルタ =====
def extract_candidates_compound_alljoin(
    pages: List[PageData], min_len=1, max_len=120, min_count=1, top_k=0, use_mecab=True
):
    if use_mecab and HAS_MECAB:
        cnt = Counter()
        for p in pages:
            try:
                save_nlp_input("mecab_compound", p.text_join, suffix=f"{p.pdf}_p{p.page}")
            except Exception:
                pass

            for w in mecab_compound_tokens_alljoin(p.text_join):
                w = strip_leading_control_chars(w)
                if not w:
                    continue
                # ★ ここでフィルタ適用(早期除外)
                if _should_skip_for_pairing(w):
                    continue
                if min_len <= len(w) <= max_len:
                    cnt[w] += 1

        arr = [(w, c) for w, c in cnt.items() if c >= min_count]
        arr.sort(key=lambda x: (-x[1], x[0]))
        return arr if (not top_k or top_k <= 0) else arr[:top_k]
    else:
        # MeCab なしフォールバック
        return extract_candidates_regex(pages, min_len=min_len, max_len=max_len, min_count=min_count)
# ===== [/REPLACE] =====

# ------------------------------------------------------------
# 文節抽出（簡易ルール）
# ------------------------------------------------------------
PUNCT_SET = set('。、．，!.！？？」』〕）】]〉》"\\\'\')')

# ===== 置き換え: MeCab 文節（セグメント単位 + 改行/タブの空白化） =====
# ===== [REPLACE] MeCab 文節（全文一発） =====
def mecab_bunsetsu_chunks(text: str) -> List[str]:
    ok, msg = ensure_mecab()
    if not ok:
        raise RuntimeError(msg)
    if not isinstance(text, str) or not text:
        return []

    chunks: List[str] = []

    def flush(cur: List[str]):
        if not cur:
            return
        s = "".join(cur)
        # 改行・タブは空白に正規化
        s = re.sub(r"[\t\r\n]+", " ", s)
        s = re.sub(r"[ ]{2,}", " ", s).strip()
        if 1 <= len(s) <= 120:
            chunks.append(s)
        cur.clear()

    toks = tokenize_mecab(text, MECAB_TAGGER)
    cur: List[str] = []
    for surf, pos1, lemma, reading, ct, cf in toks:
        if not surf:
            continue
        # 句読点などの記号で文節を打ち切り
        if surf in PUNCT_SET or pos1.startswith("記号"):
            flush(cur)
            continue
        cur.append(surf)
        # 助詞・助動詞の直後で文節を区切る（従来ルール）
        if pos1 in ("助詞", "助動詞"):
            flush(cur)

    flush(cur)
    return chunks
# ===== [/REPLACE] =====

def ginza_bunsetsu_fine_pos(sentences: list[str]) -> list[str]:
    if not sentences or not HAS_GINZA or GINZA_NLP is None:
        return []
    chunks: list[str] = []
    try:
        pipes_to_enable = [p for p in ["parser", "bunsetu_recognizer"] if p in GINZA_NLP.pipe_names]
        with GINZA_NLP.select_pipes(enable=pipes_to_enable):
            for doc in GINZA_NLP.pipe(sentences, n_process=1, batch_size=1):
                for sent in doc.sents:
                    for sp in ginza.bunsetu_spans(sent):
                        # ★ ここ：空白を保持して文節テキスト化
                        text = "".join(t.text_with_ws for t in sp).strip()
                        if len(text) <= 15:
                            chunks.append(text)
                            continue

                        # 長めの文節のみ細切り
                        buf: list[str] = []
                        for i, t in enumerate(sp):
                            # ★ ここ：細切りバッファも空白保持で積む
                            buf.append(t.text_with_ws)
                            next_token = sp[i + 1] if i + 1 < len(sp) else None

                            if t.pos_ in {"ADP", "PART"}:
                                part = "".join(buf).strip()
                                if len(part) >= 4:
                                    chunks.append(part)
                                buf.clear()
                            elif t.pos_ == "AUX":
                                if not next_token or next_token.pos_ != "AUX":
                                    part = "".join(buf).strip()
                                    if len(part) >= 4:
                                        chunks.append(part)
                                    buf.clear()
                            elif t.pos_ == "PUNCT" and t.text in {"、", "。"}:
                                part = "".join(buf).strip()
                                if len(part) >= 4:
                                    chunks.append(part)
                                buf.clear()

                        if buf:
                            part = "".join(buf).strip()
                            if len(part) >= 4:
                                chunks.append(part)
    except Exception:
        return []
    return chunks


# ===== [PATCH] bunsetsu と sentence を同時に返せるように拡張 =====
from typing import Union

def extract_candidates_bunsetsu(
    pages: List[PageData],
    min_len=1,
    max_len=120,
    min_count=1,
    top_k=0,
    *,
    also_sentence: bool = False,       # ← 追加: 文も一緒にカウントするか
    sent_max_len: int = 300            # ← 追加: 文の最大長
) -> Union[List[Tuple[str,int]], Tuple[List[Tuple[str,int]], List[Tuple[str,int]]]]:
    cnt_b = Counter()
    cnt_s = Counter() if also_sentence else None

    import re
    try:
        NUM_RE = _NUM_ONLY_LINE_RE
    except NameError:
        NUM_RE = re.compile(r"^[\+\-]?\d[\d\s,./:\-–—]*$")

    TIME_RE  = re.compile(r"""
        ^\s*
        \d{1,2} [:：] \d{2} (?: [:：] \d{2} )?
        (?: \s* (?:〜|~|\- |–|—) \s*
            \d{1,2} [:：] \d{2} (?: [:：] \d{2} )?
        )?
        \s*$
    """, re.VERBOSE)
    DATE_RE  = re.compile(r"^\s*\d{4}[\-/]\d{1,2}[\-/]\d{1,2}\s*$")

    def is_time_or_date_like_line(t: str) -> bool:
        s = strip_leading_control_chars(t).strip()
        if not s:
            return False
        if any(u in s for u in ("時","分","秒")):
            return True
        return bool(TIME_RE.fullmatch(s) or DATE_RE.fullmatch(s))

    for p in pages:
        # 既存の保存（変更なし）
        try:
            kind = "ginza_bunsetsu_posfine" if HAS_GINZA else "mecab_bunsetsu"
            save_nlp_input(kind, p.text_join, suffix=f"{p.pdf}_p{p.page}")
        except Exception:
            pass

        if HAS_GINZA:
            sentences = ginza_sentence_units(p.text_join, max_len=sent_max_len)
            sents_filtered = []
            for s in sentences:
                t = strip_leading_control_chars(s).strip()
                if not t:
                    continue
                if NUM_RE.fullmatch(t) and not is_time_or_date_like_line(t):
                    continue
                sents_filtered.append(t)

            # ★ 追加：ここで「文」を同時にカウント
            if also_sentence:
                for s in sents_filtered:
                    if 1 <= len(s) <= sent_max_len:
                        cnt_s[s] += 1

            # 文節へ（既存）
            chunks = ginza_bunsetsu_fine_pos(sents_filtered)
        else:
            # MeCab フォールバック（既存）
            lines = p.text_join.splitlines()
            clean_lines = []
            for ln in lines:
                t = strip_leading_control_chars(ln).strip()
                if t and NUM_RE.fullmatch(t) and not is_time_or_date_like_line(t):
                    continue
                clean_lines.append(ln)
            clean_text = "\n".join(clean_lines)
            chunks = mecab_bunsetsu_chunks(clean_text)

        # --- ここから追加（関数内の集計ループ直前に置く） ---
        import unicodedata

        def _is_single_symbol_1char(s: str) -> bool:
            """記号/句読点だけ1文字（Unicodeカテゴリ 'S*' or 'P*'）なら True"""
            if not isinstance(s, str):
                return False
            t = s.strip()
            if len(t) != 1:
                return False
            cat = unicodedata.category(t[0])  # 'P*' 句読点, 'S*' 記号
            return cat.startswith('P') or cat.startswith('S')

        def _is_single_alpha_1char(s: str) -> bool:
            """アルファベットだけ1文字（全角含む/NFKCでASCII化して判定）なら True"""
            if not isinstance(s, str):
                return False
            t = nfkc(s.strip())  # ← 既存の nfkc を利用
            return len(t) == 1 and (('A' <= t <= 'Z') or ('a' <= t <= 'z'))
        # --- ここまで追加 ---

        # --- 既存の「集計（既存）」からこの for ループを差し替え ---
        for ch in chunks:
            # 既存の前処理（制御文字などを剥がす）
            ch2 = strip_leading_control_chars(ch).strip()
            if not ch2:
                continue

            # ✅ sentence と同様の「数字だけ行」の除外（ただし時間/日付は残す）
            #    NUM_RE と is_time_or_date_like_line は関数冒頭で既に用意済み
            if NUM_RE.fullmatch(ch2) and not is_time_or_date_like_line(ch2):
                continue

            # ✅ 追加フィルタ：
            #  - 記号だけ1文字 … 句読点/記号カテゴリ（'P*'/'S*'）を1文字で除外
            #  - アルファベットだけ1文字 … NFKC 正規化して ASCII 英字1文字を除外（全角も対象）
            if _is_single_symbol_1char(ch2) or _is_single_alpha_1char(ch2):
                continue

            # 文字数の範囲（既存）
            if min_len <= len(ch2) <= max_len:
                cnt_b[ch2] += 1
        # --- 差し替えここまで ---


    arr_b = [(w, c) for w, c in cnt_b.items() if c >= min_count]
    arr_b.sort(key=lambda x: (-x[1], x[0]))

    if also_sentence:
        arr_s = [(w, c) for w, c in cnt_s.items() if c >= min_count]
        # 文は 文字数の制約が別なので並べ替えだけ
        arr_s.sort(key=lambda x: (-x[1], x[0]))
        return (arr_b if (not top_k or top_k <= 0) else arr_b[:top_k],
                arr_s if (not top_k or top_k <= 0) else arr_s[:top_k])
    else:
        return arr_b if (not top_k or top_k <= 0) else arr_b[:top_k]


# ===== [REPLACE] 文章（文）候補抽出：投入直前のテキストを保存 =====
def extract_candidates_sentence(
    pages: List[PageData],
    min_len: int = 1,
    max_len: int = 300,
    min_count: int = 1,
    top_k: int = 0
) -> List[Tuple[str, int]]:
    cnt = Counter()
    for p in pages:
        # ★ 投げる直前のテキストを保存
        try:
            kind = "ginza_sentence" if HAS_GINZA else "simple_sentence"
            save_nlp_input(kind, p.text_join, suffix=f"{p.pdf}_p{p.page}")
        except Exception:
            pass

        sents = ginza_sentence_units(p.text_join, max_len=max_len) if HAS_GINZA else ginza_sentence_units(p.text_join, max_len=max_len)

        # --- 追加: 時間/日付っぽい行の例外判定 ---
        import re
        try:
            NUM_RE = _NUM_ONLY_LINE_RE  # 既存の「数字だけ」判定
        except NameError:
            NUM_RE = re.compile(r"^[\+\-]?\d[\d\s,./:\-–—]*$")

        # 時間表記: HH:MM, HH:MM:SS（半角/全角コロン）, レンジ（〜, -, –, —）
        TIME_RE = re.compile(
            r"""
            ^\s*
            \d{1,2} [:\：] \d{2}            # HH:MM
            (?: [:\：] \d{2} )?             # optional :SS
            (?: \s* (?:〜|~|-|–|—) \s*      # optional range delimiter
                \d{1,2} [:\：] \d{2}
                (?: [:\：] \d{2} )?
            )?
            \s*$
            """,
            re.VERBOSE,
        )
        # 日付表記: YYYY-MM-DD / YYYY/MM/DD
        DATE_RE = re.compile(r"^\s*\d{4}[-/]\d{1,2}[-/]\d{1,2}\s*$")

        def is_time_or_date_like_line(t: str) -> bool:
            s = strip_leading_control_chars(t).strip()
            if not s:
                return False
            # 日本語の時間単位が含まれるなら時間扱い（12時34分、30秒 等）
            if any(u in s for u in ("時", "分", "秒")):
                return True
            # HH:MM(:SS)/レンジ or YYYY-MM-DD(／)
            return bool(TIME_RE.fullmatch(s) or DATE_RE.fullmatch(s))

        # 数字だけは除外。ただし「時間・日付」は残す
        filtered = []
        for s in sents:
            t = strip_leading_control_chars(s).strip()
            if not t:
                continue
            if NUM_RE.fullmatch(t) and not is_time_or_date_like_line(t):
                # 「数字だけ」行で、時間・日付のどちらにも該当しない → 除外
                continue
            filtered.append(t)

        for s in filtered:
            if s and (min_len <= len(s) <= max_len):
                cnt[s] += 1

    arr = [(w, c) for w, c in cnt.items() if c >= min_count]
    arr.sort(key=lambda x: (-x[1], x[0]))
    return arr if (not top_k or top_k <= 0) else arr[:top_k]
# ===== [/REPLACE] =====


# ------------------------------------------------------------
# Similarity
# ------------------------------------------------------------
def levenshtein_sim(a: str, b: str) -> float:
    la, lb = len(a), len(b)
    if la == 0 and lb == 0:
        return 1.0
    if a == b:
        return 1.0
    if la == 0 or lb == 0:
        return 0.0
    if la < lb:
        a, b = b, a
        la, lb = lb, la
    prev = list(range(lb + 1))
    for i in range(1, la + 1):
        cur = [i] + [0] * lb
        ca = a[i - 1]
        for j in range(1, lb + 1):
            cb = b[j - 1]
            cost = 0 if ca == cb else 1
            cur[j] = min(prev[j] + 1, cur[j - 1] + 1, prev[j - 1] + cost)
        prev = cur
    return 1.0 - prev[lb] / max(la, lb)

# === 合成 2-gram Jaccard × Levenshtein 類似 + 軽ペナルティ ===================
# 重み（お好みで微調整）
CHAR_SIM_W_LEV  = 0.60   # Levenshtein 類似の寄与
CHAR_SIM_W_JACC = 0.40   # 2-gram Jaccard の寄与

# ペナルティ係数（“軽め”）
P_LEN_PER_CHAR  = 0.01   # 表層長差 1 文字あたりの減点
P_LEN_CAP       = 0.05   # 表層長差ペナルティの上限
P_NUM_PER_CHUNK = 0.01   # 数値“まとまり”個数差 1 つあたりの減点
P_NUM_CAP       = 0.03   # 数値まとまりペナルティの上限
P_CORE_ASYM     = 0.02   # 片側だけ和字コア（かな/カナ/漢字）が存在する時の一括減点

def _jaccard_2gram(sa: str, sb: str, A=None, B=None) -> float:
    """
    2-gram Jaccard（かぶり具合）。A/B に事前計算済みの集合を渡せます。
    """
    sa = nfkc(sa or "")
    sb = nfkc(sb or "")
    A = _ngram_set(sa, 2) if A is None else set(A)
    B = _ngram_set(sb, 2) if B is None else set(B)
    inter = len(A & B)
    union = len(A | B)
    return 0.0 if union == 0 else inter / union

def _lev_sim(sa: str, sb: str) -> float:
    """
    正規化 Levenshtein 類似（1 - dist/maxlen）。既存の levenshtein_sim を利用。
    """
    return levenshtein_sim(nfkc(sa or ""), nfkc(sb or ""))

def _char_penalty(a: str, b: str) -> float:
    """
    軽い機械的ペナルティの合算：
      - 表層長差（NFKC + 空白除去後の len 差）
      - 数値“まとまり”個数差（例: '630W[30°C]' -> 2）
      - 片側だけ和字コア（かな/カナ/漢字）がある
    """
    try:
        la = _norm_len_for_surface(a); lb = _norm_len_for_surface(b)
        p_len = min(P_LEN_CAP, P_LEN_PER_CHAR * abs(la - lb))
    except Exception:
        p_len = 0.0

    try:
        na = _numeric_chunk_count(a); nb = _numeric_chunk_count(b)
        p_num = min(P_NUM_CAP, P_NUM_PER_CHUNK * abs(na - nb))
    except Exception:
        p_num = 0.0

    try:
        core_a = _surface_core_for_reading(a)
        core_b = _surface_core_for_reading(b)
        p_core = P_CORE_ASYM if bool(core_a) != bool(core_b) else 0.0
    except Exception:
        p_core = 0.0

    return float(p_len + p_num + p_core)

def combined_char_similarity(
    a: str,
    b: str,
    *,
    sa: str | None = None,
    sb: str | None = None,
    grams_a=None,
    grams_b=None,
    w_lev: float = CHAR_SIM_W_LEV,
    w_jacc: float = CHAR_SIM_W_JACC,
    clamp: tuple[float, float] = (0.0, 1.0),
) -> float:
    """
    文字ベースの最終スコア：
        raw = w_lev * Levenshtein_sim  +  w_jacc * Jaccard2
        score = clamp(raw - penalty)

    - sa/sb: NFKC 済み表層を渡すと再正規化を避けられます
    - grams_a/grams_b: 2-gram 集合（frozenset など）を渡すと再構築を避けられます
    - 戻り値は [0.000, 1.000] に丸め（小数第3位）
    """
    # NFKC 済みを尊重（未指定なら正規化）
    sa = nfkc(sa if sa is not None else (a or ""))
    sb = nfkc(sb if sb is not None else (b or ""))

    # 完全一致は 1.0（basic）
    if sa == sb:
        return 1.0

    # 成分スコア
    jacc = _jaccard_2gram(sa, sb, grams_a, grams_b)
    levs = _lev_sim(sa, sb)
    raw  = (w_lev * levs) + (w_jacc * jacc)
    raw  = max(clamp[0], min(clamp[1], raw))

    # 軽い機械ペナルティ
    pen = _char_penalty(a, b)
    score = max(clamp[0], min(clamp[1], raw - pen))

    return round(float(score), 3)

# ===== Edge-only difference (first/last 1-char) =====
def classify_edge_diff(a: str, b: str) -> str:
    """
    端（前後）1文字だけの差で説明できるかを判定し、
    種別を返す（該当なしは "" を返す）。
      "prefix_add" / "prefix_change" / "suffix_add" / "suffix_change" / ""
    """
    if a is None or b is None:
        return ""
    a = str(a); b = str(b)
    if a == b:
        return ""
    la, lb = len(a), len(b)

    # 前: 有無（片方が1文字長くて、2文字目以降が完全一致）
    if la == lb + 1 and a[1:] == b:
        return "prefix_add"
    if lb == la + 1 and b[1:] == a:
        return "prefix_add"

    # 前: 違い（同長で、2文字目以降が一致）
    if la == lb and la >= 1 and a[1:] == b[1:]:
        return "prefix_change"

    # 後: 有無（片方が1文字長くて、末尾1文字を除けば一致）
    if la == lb + 1 and a[:-1] == b:
        return "suffix_add"
    if lb == la + 1 and b[:-1] == a:
        return "suffix_add"

    # 後: 違い（同長で、末尾以外が一致）
    if la == lb and la >= 1 and a[:-1] == b[:-1]:
        return "suffix_change"

    return ""

def classify_edge_diff_jp(a: str, b: str) -> str:
    """端差の日本語ラベル（表示・フィルタ用）。"""
    m = classify_edge_diff(a, b)
    return {
        "prefix_add":   "前1字有無",
        "prefix_change":"前1字違い",
        "suffix_add":   "後1字有無",
        "suffix_change":"後1字違い",
    }.get(m, "")

# ==== ベクトル化ユーティリティ（新規） =====================================
_DIGIT_RE = re.compile(r"\d", flags=re.UNICODE)

def _edge_labels_vectorized(df: pd.DataFrame) -> pd.Series:
    a = df["a"].astype("string")
    b = df["b"].astype("string")
    la = a.str.len()
    lb = b.str.len()
    res = pd.Series("", index=df.index, dtype="string")

    # 前1字有無
    m = (la == lb + 1) & (a.str.slice(1) == b)
    res = res.mask(m, "前1字有無")
    m = (lb == la + 1) & (b.str.slice(1) == a) & (res == "")
    res = res.mask(m, "前1字有無")

    # 前1字違い
    m = (la == lb) & (la >= 1) & (a.str.slice(1) == b.str.slice(1)) & (res == "")
    res = res.mask(m, "前1字違い")

    # 後1字有無
    m = (la == lb + 1) & (a.str.slice(0, -1) == b) & (res == "")
    res = res.mask(m, "後1字有無")
    m = (lb == la + 1) & (b.str.slice(0, -1) == a) & (res == "")
    res = res.mask(m, "後1字有無")

    # 後1字違い
    m = (la == lb) & (la >= 1) & (a.str.slice(0, -1) == b.str.slice(0, -1)) & (res == "")
    res = res.mask(m, "後1字違い")
    return res

# ============================================================================
# 数字以外一致（数字「まとまり」を無視して比較）
_NUMERIC_CHUNK_RE = re.compile(r"[+\-]?\d(?:[\d,.\s]*\d)?")

# 空白の正規化（半角スペース/全角スペースを1個の半角スペースに）
_WS_RE = re.compile(r"[\s\u3000]+")

def _strip_numbers_like(s: str) -> str:
    """文字列から『数字のまとまり』をすべて除去し、残りを空白正規化して返す。"""
    if s is None:
        s = ""
    s = str(s)
    # 内部のみ NFKC（全角→半角、全角記号→半角記号 等）
    nf = unicodedata.normalize("NFKC", s)
    # 数字まとまりを空文字へ（桁区切り・小数点・連続ドット等を含めて除去）
    t = _NUMERIC_CHUNK_RE.sub("", nf)
    # 余計な空白を整理（比較を安定化）
    t = _WS_RE.sub(" ", t).strip()
    return t

def is_numeric_only_diff(a: str, b: str) -> bool:
    if a is None and b is None:
        return False
    a_s, b_s = "" if a is None else str(a), "" if b is None else str(b)
    if a_s == b_s:
        return False
    # 少なくとも一方に数字があるか（NFKC後に判定）
    a_has = bool(re.search(r"\d", unicodedata.normalize("NFKC", a_s)))
    b_has = bool(re.search(r"\d", unicodedata.normalize("NFKC", b_s)))
    if not (a_has or b_has):
        return False
    # 数字のまとまりを除去して比較
    return _strip_numbers_like(a_s) == _strip_numbers_like(b_s)

def _numeric_only_label_vectorized(df: pd.DataFrame) -> pd.Series:
    a = df["a"].astype("string")
    b = df["b"].astype("string")

    # 事前条件: 元が同一は対象外
    diff_orig = (a != b)

    # “数字が含まれるか” の判定は NFKC 後に実施
    a_nf = a.map(lambda x: unicodedata.normalize("NFKC", "" if pd.isna(x) else str(x)))
    b_nf = b.map(lambda x: unicodedata.normalize("NFKC", "" if pd.isna(x) else str(x)))
    has_digit = a_nf.str.contains(r"\d") | b_nf.str.contains(r"\d")

    # 数字まとまりを除去して比較
    a_stripped = a.map(_strip_numbers_like)
    b_stripped = b.map(_strip_numbers_like)

    mask = diff_orig & has_digit & (a_stripped == b_stripped)
    out = pd.Series("", index=df.index, dtype="string")
    out[mask] = "数字以外一致"
    return out

def numeric_only_label(a: str, b: str) -> str:
    """表示用ラベル（該当時のみ '数字以外一致' を返す）"""
    return "数字以外一致" if is_numeric_only_diff(a, b) else ""

# ===== [ADD] ペア作成前の共通フィルタ関数(グローバルヘルパーとして追加) =====
# ※ is_numeric_only_diff() の直後あたりに配置するのが適切

def _should_skip_for_pairing(s: str) -> bool:
    """
    ペア作成時にスキップすべき文字列かどうかを判定
    - 数字のみ(ただし時刻/日付形式は除く)
    - アルファベット1文字のみ(全角含む)
    - 記号のみ
    
    戻り値:
        True: スキップすべき(候補から除外)
        False: ペア作成対象として保持
    """
    if not isinstance(s, str) or not s:
        return True
    
    s_strip = s.strip()
    if not s_strip:
        return True
    
    # === 1) 数字のみ(ただし時刻/日付は許容) ===
    # 既存の正規表現を利用
    try:
        NUM_RE = _NUM_ONLY_LINE_RE
    except NameError:
        NUM_DIG = r"[0-9\uFF10-\uFF19]"
        GROUP_SEP = r"[,\uFF0C \u00A0\u3000]"
        DEC_SEP = r"[.\uFF0E]"
        NUM_RE = re.compile(
            rf"^{NUM_DIG}+(?:{GROUP_SEP}{NUM_DIG}{{3}})*(?:{DEC_SEP}{NUM_DIG}+)?$",
            re.VERBOSE
        )
    
    # 時刻/日付判定(既存ロジックを再利用)
    TIME_RE = re.compile(
        r"""^\s*\d{1,2}[:\uff1a]\d{2}(?:[:\uff1a]\d{2})?
            (?:\s*(?:〜|~|\-|—|–)\s*\d{1,2}[:\uff1a]\d{2}(?:[:\uff1a]\d{2})?)?\s*$""",
        re.VERBOSE
    )
    DATE_RE = re.compile(r"^\s*\d{4}[-/]\d{1,2}[-/]\d{1,2}\s*$")
    
    def is_time_or_date(t: str) -> bool:
        if any(u in t for u in ("時", "分", "秒")):
            return True
        return bool(TIME_RE.fullmatch(t) or DATE_RE.fullmatch(t))
    
    # 数字のみ かつ 時刻/日付でない → スキップ
    if NUM_RE.fullmatch(s_strip) and not is_time_or_date(s_strip):
        return True
    
    # === 2) アルファベット1文字のみ(全角含む、NFKC後で判定) ===
    try:
        s_nfkc = nfkc(s_strip)
    except NameError:
        s_nfkc = unicodedata.normalize("NFKC", s_strip)
    
    if len(s_nfkc) == 1:
        if ('A' <= s_nfkc <= 'Z') or ('a' <= s_nfkc <= 'z'):
            return True
    
    # === 3) 記号のみ(Unicode Category 'P*' or 'S*' のみで構成) ===
    # 空白以外の文字が全て記号/句読点なら除外
    non_space_chars = [ch for ch in s_nfkc if not ch.isspace()]
    if non_space_chars:
        if all(unicodedata.category(ch).startswith(('P', 'S')) for ch in non_space_chars):
            return True
    
    return False
# ===== [/ADD] =====

# ------------------------------------------------------------
# Lexical（細粒度）語彙とペア
# ------------------------------------------------------------
# ===== [REPLACE] 語彙構築：MeCab 投入直前のテキストを保存 =====
def collect_lexicon_general(pages: List[PageData], top_k=4000, min_count=1) -> pd.DataFrame:
    ok, msg = ensure_mecab()
    if not ok:
        raise RuntimeError(msg)

    counts = Counter()
    readings_map: Dict[str, Counter] = defaultdict(Counter)
    lemma_map: Dict[str, Counter] = defaultdict(Counter)
    lemma_read_map: Dict[str, Counter] = defaultdict(Counter)
    pos_map: Dict[str, Counter] = defaultdict(Counter)

    for p in pages:
        try:
            save_nlp_input("mecab_lexicon", p.text_join, suffix=f"{p.pdf}_p{p.page}")
        except Exception:
            pass

        for surf, pos1, lemma, reading, ct, cf in tokenize_mecab(p.text_join, MECAB_TAGGER):
            if not pos1.startswith(("名詞", "動詞", "形容詞", "助動詞")):
                continue
            # ★ ここでフィルタ適用(早期除外)
            if _should_skip_for_pairing(surf):
                continue
            
            counts[surf] += 1
            r = normalize_kana(reading or "", drop_choon=True)
            if r:
                readings_map[surf][r] += 1
            if lemma:
                lemma_read_map[lemma][r] += 1
            if lemma:
                lemma_map[surf][lemma] += 1
            pos_map[surf][pos1] += 1

    def top1(cnt: Counter) -> str:
        return cnt.most_common(1)[0][0] if cnt else ""

    items = [(s, c) for s, c in counts.items() if c >= min_count]
    items.sort(key=lambda x: (-x[1], x[0]))
    if top_k and top_k > 0:
        items = items[:top_k]

    lemma2readnorm = {lem: top1(lemma_read_map[lem]) for lem in lemma_read_map}
    rows = []
    for s, c in items:
        lem = top1(lemma_map[s])
        rows.append({
            "surface": s,
            "count": int(c),
            "reading": top1(readings_map[s]),
            "lemma": lem,
            "lemma_read": lemma2readnorm.get(lem, ""),
            "pos1": top1(pos_map[s]),
        })
    return pd.DataFrame(rows)
# ===== [/REPLACE] =====


# ===== [REPLACE] build_synonym_pairs_general: 語彙を「名詞」「助・動詞」に分離 =====
def build_synonym_pairs_general(
    df_lex: pd.DataFrame,
    read_sim_th=0.90,
    char_sim_th=0.85,
    scope="語彙",          # ← 呼び出し互換のため残します（内部では無視）
    progress_cb=None
) -> pd.DataFrame:
    """
    語彙ペアの生成。scope を固定の「語彙」にせず、
    表層ごとの最頻 pos1 を用いて「名詞」「助・動詞」に二分して付与します。
    ・名詞 … pos1.startswith("名詞")
    ・助・動詞 … pos1.startswith(("動詞","助動詞","形容詞"))  # 形容詞は用言としてこちらに含める
    ペアの scope は「両方が名詞なら '名詞'、それ以外は '助・動詞'」に統一。
    """
    if df_lex.empty:
        return empty_pairs_df()

    # --- 表層ごとの粗カテゴリを作成（名詞 / 助・動詞） -------------------------
    def _coarse_group(pos1: str) -> str:
        p = str(pos1 or "")
        if p.startswith("名詞"):
            return "名詞"
        if p.startswith(("動詞", "助動詞", "形容詞")):
            return "助・動詞"
        # 未判定は名詞扱い（多くの語彙が名詞のため安全側）
        return "名詞"

    surf_series = df_lex["surface"].astype(str)
    pos1_series = df_lex.get("pos1", pd.Series([""] * len(df_lex)))
    surf2group: Dict[str, str] = {s: _coarse_group(p) for s, p in zip(surf_series, pos1_series)}

    # --- 既存ロジック（読み/文字類似によるペア化）は極力そのまま -------------------
    words = df_lex["surface"].tolist()
    counts = df_lex["count"].tolist()
    lemmas = df_lex["lemma"].fillna("").tolist()
    readings = df_lex["reading"].fillna("").tolist()

    n = len(words)
    norm_surface = [nfkc(w) for w in words]
    norm_read = [normalize_kana(r or "", True) for r in readings]
    ng_char = [_ngram_set(s) for s in norm_surface]

    rows = []
    step = max(1, n // 50)

    def try_reading_like(a, b, ra, rb, th):
        # 第1パス
        sim = reading_sim_with_penalty(a, b, ra, rb, th)
        if sim is not None:
            return sim
        # 第2パス（漢字↔カナなどは少し緩める）
        if _has_kanji(a) != _has_kanji(b):
            th2 = max(0.75, float(th) - 0.15)
            return reading_sim_with_penalty(a, b, ra, rb, th2)
        return None

    for i in range(n):
        for j in range(i + 1, n):
            a, b = words[i], words[j]
            ca, cb = int(counts[i]), int(counts[j])
            la, lb = lemmas[i], lemmas[j]
            ra, rb = norm_read[i], norm_read[j]
            sa, sb = norm_surface[i], norm_surface[j]

            reason = None
            score = 0.0

            # 🆕 更新頻度を上げる(元は100分割 → 200分割)
            if progress_cb:
                # より細かく進捗を報告(0.5%刻み)
                if (i % max(1, n // 50) == 0) or i == n - 1:
                    progress_cb(i + 1, n)

            # 1) lemma 優先（既存）
            if la and lb and la == lb and a != b:
                if (a == la) or (b == la):
                    reason, score = "lemma", 1.0
                else:
                    reason, score = "inflect", 0.95
            else:
                # 1.5) 和字コア一致のみ（英数記号差）は reading_eq
                if is_symbol_only_surface_diff(a, b) and a != b:
                    reason, score = "reading_eq", reading_eq_score(a, b)
                else:
                    # 2) 読み類似
                    sim_r = try_reading_like(a, b, ra, rb, read_sim_th)
                    if sim_r is not None:
                        reason, score = "reading", round(float(sim_r), 3)
                    else:
                        # 3) 文字類似（合成：2-gram Jaccard × Levenshtein + 軽ペナルティ）
                        inter = len(ng_char[i] & ng_char[j])
                        union = len(ng_char[i] | ng_char[j])
                        if union > 0 and (inter / union) >= 0.30:
                            sim_val = combined_char_similarity(
                                a, b,
                                sa=sa, sb=sb,
                                grams_a=ng_char[i], grams_b=ng_char[j],
                            )
                            if sim_val >= 0.999:
                                reason, score = "basic", 1.0
                            elif sim_val is not None and sim_val >= float(char_sim_th or 0.0):
                                reason, score = "char", sim_val

            if reason:
                # ★ scope を「名詞」「助・動詞」に二分
                ga = surf2group.get(a, "名詞")
                gb = surf2group.get(b, "名詞")
                scope_ab = "名詞" if (ga == "名詞" and gb == "名詞") else "活用語"

                rows.append({
                    "a": a, "b": b,
                    "a_count": ca, "b_count": cb,
                    "reason": reason, "score": score,
                    "scope": scope_ab,
                })

        if progress_cb and (i % step == 0 or i == n - 1):
            progress_cb(i + 1, n)

    if not rows:
        return empty_pairs_df()

    return pd.DataFrame(rows).sort_values(
        ["reason", "score", "a_count", "b_count"],
        ascending=[True, False, False, False]
    )
# ===== [/REPLACE] ==============================================================

# ===== 並列版 build_synonym_pairs_char_only（関数ごと差し替え） =====
from concurrent.futures import ProcessPoolExecutor, as_completed
import os
import math

# ===== [6-B2] 並列ワーカー差し替え：逆引きインデックスで候補前絞り =====
def _pairs_chunk_worker(args):
    """
    子プロセス側：i 範囲を担当し、逆引きインデックスで候補を前絞りしてから判定。
    文字類似は「2-gram Jaccard × Levenshtein の合成 + 軽ペナルティ」を用いる。
    """
    (i_start, i_end, words, counts, norm_char, read_norm,
     ngram_list, char_sim_th, read_sim_th, scope) = args

    rows = []
    n = len(words)
    JACC_TH = 0.30  # 粗フィルタ用（前段の足切り）

    # 逆引きインデックス（ローカル構築）
    from collections import defaultdict
    inv = defaultdict(list)
    for idx, grams in enumerate(ngram_list):
        for g in grams:
            inv[g].append(idx)
    for g, lst in inv.items():
        inv[g] = sorted(set(lst))

    # 読み類似（第1/第2パス）
    def try_reading_like(a, b, ra, rb, th):
        sim = reading_sim_with_penalty(a, b, ra, rb, th)
        if sim is not None:
            return sim
        if _has_kanji(a) != _has_kanji(b):
            th2 = max(0.75, float(th) - 0.15)
            return reading_sim_with_penalty(a, b, ra, rb, th2)
        return None

    for i in range(i_start, i_end):
        a  = words[i]
        sa = norm_char[i]
        Ai = ngram_list[i]

        # まず 2-gram の共起から候補集合
        pool = set()
        for g in Ai:
            pool.update(inv.get(g, ()))
        cand_js = [j for j in pool if j > i and j < n]

        for j in cand_js:
            b  = words[j]

            # 【読み一致（和字コア一致のみ）→ 既存ロジック】
            if read_sim_th is not None and is_symbol_only_surface_diff(a, b) and a != b:
                rows.append({
                    "a": a, "b": b,
                    "a_count": counts[a], "b_count": counts[b],
                    "reason": "reading_eq",
                    "score": reading_eq_score(a, b),
                    "scope": scope,
                })
                continue

            # 【読み類似】（既存）
            if read_sim_th is not None:
                ra = read_norm[i] if read_norm[i] is not None else ""
                rb = read_norm[j] if read_norm[j] is not None else ""
                sim_r = try_reading_like(a, b, ra, rb, read_sim_th)
                if sim_r is not None:
                    rows.append({
                        "a": a, "b": b,
                        "a_count": counts[a], "b_count": counts[b],
                        "reason": "reading",
                        "score": round(float(sim_r), 3),
                        "scope": scope,
                    })
                    continue

            # 【文字類似】ここを “合成スコア” に差し替え
            Aj = ngram_list[j]
            inter = len(Ai & Aj)
            union = len(Ai | Aj)
            if union == 0:
                continue
            jacc = inter / union
            if jacc < JACC_TH:  # 粗フィルタ
                continue

            sb = norm_char[j]
            sim_val = combined_char_similarity(
                a, b,
                sa=sa, sb=sb,
                grams_a=Ai, grams_b=Aj,
            )

            # 完全一致は "basic"、それ以外は "char"
            if sim_val >= 0.999:
                rows.append({
                    "a": a, "b": b,
                    "a_count": counts[a], "b_count": counts[b],
                    "reason": "basic", "score": 1.0,
                    "scope": scope,
                })
            elif sim_val >= float(char_sim_th or 0.0):
                rows.append({
                    "a": a, "b": b,
                    "a_count": counts[a], "b_count": counts[b],
                    "reason": "char", "score": sim_val,
                    "scope": scope,
                })

    return rows

def build_synonym_pairs_char_only(
    tokens: List[Tuple[str,int]],
    char_sim_th=0.90,
    read_sim_th=None,
    top_k=4000,
    scope="複合語",
    progress_cb=None,
    use_processes: bool = True,
    max_workers: Optional[int] = None
) -> pd.DataFrame:
    """
    並列対応・高速版：
      - 読み類似は（指定時）帯付きLevenshtein
      - 文字類似は n-gram Jaccard の粗フィルタ → 帯付きLevenshtein
      - i の範囲をチャンク分割し、ProcessPoolExecutor で分散
    """
    if not tokens:
        return empty_pairs_df()

    tokens = sorted(tokens, key=lambda x: (-x[1], x[0]))
    if top_k and top_k > 0:
        tokens = tokens[:top_k]

    words = [w for w, _ in tokens]
    counts = {w: int(c) for w, c in tokens}

    # 正規化文字列
    norm_char = [nfkc(w) for w in words]

    # 読み正規化（キャッシュ利用）
    use_reading = (read_sim_th is not None) and HAS_MECAB
    read_norm = [""] * len(words)
    if use_reading:
        ok, _ = ensure_mecab()
        if ok:
            read_norm = [phrase_reading_norm_cached(w) for w in words]

    # 2-gram の集合（frozenset にして子プロセスへシリアライズ負荷を軽減）
    ngram_list = [frozenset(_ngram_set(s)) for s in norm_char]

    n = len(words)
    if n < 2:
        return empty_pairs_df()

    # 逐次版（use_processes=False またはワーカ=1 の場合）
    if not use_processes:
        rows = _pairs_chunk_worker((0, n-1, words, counts, norm_char, read_norm,
                                    ngram_list, char_sim_th, read_sim_th, scope))
        if not rows:
            return empty_pairs_df()
        return pd.DataFrame(rows).sort_values(
            ["score", "a_count", "b_count"], ascending=[False, False, False]
        )

    # 並列版
    workers = max(1, (max_workers or os.cpu_count() or 1))
    if workers == 1 or n < 1024:
        # 要素数が少ない時はオーバーヘッド回避のため逐次
        rows = _pairs_chunk_worker((0, n-1, words, counts, norm_char, read_norm,
                                    ngram_list, char_sim_th, read_sim_th, scope))
        if not rows:
            return empty_pairs_df()
        return pd.DataFrame(rows).sort_values(
            ["score", "a_count", "b_count"], ascending=[False, False, False]
        )

    # i の分割（i は 0..n-2 を走査）
    # チャンク数は workers * 4 目安（小さすぎると負荷分散が偏る）
    total_i = n - 1
    chunks = max(workers * 4, 1)
    step = math.ceil(total_i / chunks)
    tasks = []
    futures = []
    args_common = (words, counts, norm_char, read_norm, ngram_list, char_sim_th, read_sim_th, scope)

    with ProcessPoolExecutor(max_workers=workers) as ex:
        for s in range(0, total_i, step):
            e = min(total_i, s + step)
            # 子プロセスにはタプルで引数一括渡し
            args = (s, e, *args_common)
            futures.append(ex.submit(_pairs_chunk_worker, args))
            tasks.append((s, e))

        rows_all = []
        # 進捗：チャンクの「終端 i」を元の n に写像して報告（おおよそでOK）
        for (s, e), fut in zip(tasks, as_completed(futures)):
            try:
                rows_all.extend(fut.result())
            except Exception:
                pass
            if progress_cb:
                # 🆕 チャンク完了ごとに進捗を報告(より細かく)
                progress_cb(min(e, n-1), n)

    if not rows_all:
        return empty_pairs_df()

    return pd.DataFrame(rows_all).sort_values(
        ["score", "a_count", "b_count"], ascending=[False, False, False]
    )
# ===== 並列版 ここまで =====

# ------------------------------------------------------------
# 統合（★「読み類似」を最優先）
# ------------------------------------------------------------
REASON_PRIORITY = {
    "basic": 7.2,          # 基本一致（NFKC同一）
    "reading_eq": 6.8,     # 読み一致（英数記号除く）
    "reading_same": 6.5,   # ★ 追加：読み一致（表記違い）
    "reading": 6.0,        # 読み類似
    "lemma": 5.0,
    "inflect": 4.5,
    "lemma_read": 4.0,
    "char": 2.0,
}

# --- REPLACE: unify_pairs (簡素化) ---
def unify_pairs(*dfs: pd.DataFrame) -> pd.DataFrame:
    frames = []
    for d in dfs:
        if d is None or d.empty:
            continue
        t = d.copy()
        for col in PAIR_COLUMNS:
            if col not in t.columns:
                t[col] = pd.NA
        frames.append(t[PAIR_COLUMNS].copy())
    if not frames:
        return empty_pairs_df()

    df = pd.concat(frames, ignore_index=True)

    # 正規化キー（a<=b で並べ替えたペア）
    def _norm_row(r):
        a, b = r["a"], r["b"]
        if pd.isna(a) or pd.isna(b):
            return a, b
        return (a, b) if str(a) <= str(b) else (b, a)

    df[["a_n", "b_n"]] = df.apply(_norm_row, axis=1, result_type="expand")

    # 優先度と合算頻度を付与
    df["prio"] = df["reason"].map(REASON_PRIORITY).fillna(1.0)
    df["sum_count"] = (
        pd.to_numeric(df["a_count"], errors="coerce").fillna(0).astype(int) +
        pd.to_numeric(df["b_count"], errors="coerce").fillna(0).astype(int)
    )
    df["score"] = pd.to_numeric(df["score"], errors="coerce").fillna(0.0)

    # （a_n,b_n）グループで最良（prio→score→sum_count の降順）を1件選抜
    df = df.sort_values(
        ["a_n", "b_n", "prio", "score", "sum_count"],
        ascending=[True, True, False, False, False]
    ).drop_duplicates(["a_n", "b_n"], keep="first")

    # 表示用の最終順
    df = df.drop(columns=["a_n", "b_n", "prio", "sum_count"], errors="ignore")
    return df.sort_values(["reason", "score", "a", "b"], ascending=[True, False, True, True]).reset_index(drop=True)
# --- /REPLACE ---

# ------------------------------------------------------------
# 変種のグループ化（lemma＋連結成分）→ gid 付与に利用
# ------------------------------------------------------------
def build_variation_groups(df_unified: pd.DataFrame, df_lex: Optional[pd.DataFrame] = None):
    from collections import defaultdict, Counter, deque
    g = defaultdict(set)

    if df_unified is not None and not df_unified.empty:
        for a, b in df_unified[['a','b']].itertuples(index=False):
            if isinstance(a, str) and isinstance(b, str) and a and b:
                g[a].add(b); g[b].add(a)

    surf2lemma = {}
    if df_lex is not None and not df_lex.empty:
        for s, l in df_lex[['surface','lemma']].itertuples(index=False):
            if isinstance(s, str) and s:
                surf2lemma[s] = l if isinstance(l, str) and l else s

    lemma_groups = defaultdict(list)
    for s, l in surf2lemma.items():
        lemma_groups[l].append(s)
    for mems in lemma_groups.values():
        for i in range(len(mems)):
            for j in range(i+1, len(mems)):
                a, b = mems[i], mems[j]
                g[a].add(b); g[b].add(a)

    visited = set()
    groups = []
    for node in list(g.keys()):
        if node in visited:
            continue
        q = deque([node]); visited.add(node)
        comp = []
        while q:
            u = q.popleft(); comp.append(u)
            for v in g[u]:
                if v not in visited:
                    visited.add(v); q.append(v)
        groups.append(sorted(set(comp)))

    df_groups_rows = []
    surf2gid: Dict[str, int] = {}
    gid2members: Dict[int, List[str]] = {}

    def best_label(members):
        if surf2lemma:
            cc = Counter(surf2lemma.get(s, s) for s in members)
            lab, cnt = cc.most_common(1)[0]
            if lab:
                return lab
        return sorted(members, key=lambda x: (len(x), x))[0]

    for gid, mems in enumerate(sorted(groups, key=lambda ms: (-len(ms), ms[0] if ms else "")), start=1):
        label = best_label(mems) if mems else ""
        for s in mems:
            surf2gid[s] = gid
        gid2members[gid] = mems
        df_groups_rows.append({'gid': gid, 'label': label, 'size': len(mems), 'members': ", ".join(mems)})

    df_groups = pd.DataFrame(df_groups_rows) if df_groups_rows else pd.DataFrame(
        columns=['gid','label','size','members']
    )
    return df_groups, surf2gid, gid2members

# ==== ここから差し替え：2→3分類に拡張（新カテゴリを表示側に追加） ====
REASON_TO_GROUP = {
    "basic": "basic",
    "reading_eq": "reading_eq",
    "reading_same": "reading_same",  # ★ 追加
    "lemma": "reading",
    "lemma_read": "reading",
    "reading": "reading",
    "inflect": "reading",
    "char": "char",
}

REASON_GROUP_JA = {
    "basic": "基本一致",
    "reading_eq": "読み一致（英数記号除く）",
    "reading_same": "読み一致（表記違い）",  # ★ 追加
    "reading": "読み類似",
    "char": "文字類似",
}

def apply_reason_ja(df: pd.DataFrame) -> pd.DataFrame:
    """
    reason列（英）→ 「一致要因」（日本語）の3分類に変換。
    - 未知のreasonは安全側で「文字類似」にフォールバック。
    - 最終的に「reason」列は削除。
    """
    if df is None or df.empty or "reason" not in df.columns:
        return df
    df = df.copy()
    grp_key = df["reason"].map(REASON_TO_GROUP).fillna("char")
    df["一致要因"] = grp_key.map(REASON_GROUP_JA).fillna("文字類似")
    df = df.drop(columns=["reason"])
    return df

# ------------------------------------------------------------
# Diff（インライン差分のみ）
# ------------------------------------------------------------
import html

def _esc_html(s: str) -> str:
    return "" if s is None else html.escape(str(s), quote=False)

# ===== [REPLACE] html_inline_diff（クラス付与版：eq/tok-del/tok-ins） =====
def html_inline_diff(a: str, b: str) -> str:
    """
    インライン差分を HTML で返す。
    等しい部分: <span class='eq'>…</span>
    削除      : <span class='tok-del'>…</span>
    追加      : <span class='tok-ins'>…</span>
    """
    sm = difflib.SequenceMatcher(a=a or "", b=b or "")
    out = []
    A = a or ""; B = b or ""
    for op, i1, i2, j1, j2 in sm.get_opcodes():
        if op == "equal":
            seg = _esc_html(A[i1:i2])
            if seg:
                out.append(f"<span class='eq'>{seg}</span>")
        elif op == "delete":
            seg = _esc_html(A[i1:i2])
            if seg:
                out.append(f"<span class='tok-del'>{seg}</span>")
        elif op == "insert":
            seg = _esc_html(B[j1:j2])
            if seg:
                out.append(f"<span class='tok-ins'>{seg}</span>")
        elif op == "replace":
            del_seg = _esc_html(A[i1:i2])
            ins_seg = _esc_html(B[j1:j2])
            if del_seg:
                out.append(f"<span class='tok-del'>{del_seg}</span>")
            if ins_seg:
                out.append(f"<span class='tok-ins'>{ins_seg}</span>")
    return "".join(out)
# ===== [/REPLACE] =====

# ===== [REPLACE] 左プレビュー: 統合インライン差分の見た目（薄い未変更＋淡背景） =====
def build_unified_inline_diff_embed(a: str, b: str) -> str:
    a = "" if a is None else str(a)
    b = "" if b is None else str(b)
    body = html_inline_diff(a, b)

    css = """
    <style>
      html, body { margin:0; padding:0; background:#fff; }
      body {
        font-family: 'Yu Gothic UI','Noto Sans JP',sans-serif;
        font-size: 16px;                /* お好みで 15〜18px */
        line-height: 1.5; letter-spacing: .2px;
      }
      .box {
        border: 1px solid #dee2e6; border-radius: 6px;
        padding: 8px 10px; background: #fff;
        white-space: pre-wrap; word-break: break-word;
      }
      /* 等しい部分：薄めのグレー（前回よりさらに薄く） */
      .eq       { color:#94a3b8; }       /* 例: #94a3b8 / #9aa0a6 / #a0a7b1 */
      /* 追加：青字＋淡い青背景＋下線を少し太く */
      .tok-ins  {
        color:#1c7ed6; background:#d0ebff;
        text-decoration: underline;
        text-decoration-thickness: 2px; text-underline-offset: 1px;
        border-radius: 2px; padding: 0 1px;
      }
      /* 削除：赤字＋淡い赤背景＋取り消し線を少し太く */
      .tok-del  {
        color:#c92a2a; background:#ffe3e3;
        text-decoration: line-through;
        text-decoration-thickness: 2px;
        border-radius: 2px; padding: 0 1px;
      }
    </style>
    """
    return f"<html><head>{css}</head><body><div class='box'>{body}</div></body></html>"


# ===== [REPLACE] 差分プレビュー用：縦並び（見出しなし・A/B表記なし） =====
def build_vertical_diff_html_embed(a: str, b: str) -> str:
    """
    簡易プレビュー向けの縦並び差分（見出しなし/A・B表記なし）。
    上段=旧テキスト相当、下段=新テキスト相当 ですが文言は出しません。
    """
    import difflib, html as _html

    def esc(s: str) -> str:
        return _html.escape(s or "", quote=False)

    def tok_a(a_line: str, b_line: str) -> str:
        sm2 = difflib.SequenceMatcher(a=a_line or "", b=b_line or "")
        parts = []
        for op, i1, i2, j1, j2 in sm2.get_opcodes():
            a_seg = (a_line or "")[i1:i2]
            if op == "equal":
                parts.append(esc(a_seg))
            elif op in ("delete", "replace"):
                if a_seg:
                    parts.append(f"<span class='tok-del'>{esc(a_seg)}</span>")
        return "".join(parts)

    def tok_b(a_line: str, b_line: str) -> str:
        sm2 = difflib.SequenceMatcher(a=a_line or "", b=b_line or "")
        parts = []
        for op, i1, i2, j1, j2 in sm2.get_opcodes():
            b_seg = (b_line or "")[j1:j2]
            if op == "equal":
                parts.append(esc(b_seg))
            elif op in ("insert", "replace"):
                if b_seg:
                    parts.append(f"<span class='tok-ins'>{esc(b_seg)}</span>")
        return "".join(parts)

    a_lines = (a or "").splitlines()
    b_lines = (b or "").splitlines()
    sm = difflib.SequenceMatcher(a=a_lines, b=b_lines)
    ops = sm.get_opcodes()

    css = """
    <style>
      body { background:#fff; color:#111; font-family:'Yu Gothic UI','Noto Sans JP',sans-serif;
             font-size:13px; line-height:1.6; margin:0; }
      .sec { margin:0 0 6px 0; }
      .box { border:1px solid #dee2e6; border-radius:6px; padding:6px 8px; background:#fff; }
      .line { white-space:pre-wrap; padding:1px 0; }
      .eq  {}
      .add { background:#e7f5ff; }
      .del { background:#fff0f0; }
      .rep { background:#fff7e6; }
      .tok-ins { color:#1c7ed6; background:#d0ebff; border-radius:2px; }
      .tok-del { color:#c92a2a; background:#ffe3e3; border-radius:2px; }
    </style>
    """

    a_rows = ["<div class='sec'><div class='box'>"]
    b_rows = ["<div class='sec'><div class='box'>"]

    for tag, i1, i2, j1, j2 in ops:
        if tag == "equal":
            for k in range(i2 - i1):
                text = esc(a_lines[i1 + k])
                a_rows.append(f"<div class='line eq'>{text}</div>")
                b_rows.append(f"<div class='line eq'>{text}</div>")
        elif tag == "delete":
            for k in range(i2 - i1):
                a_line = a_lines[i1 + k]
                a_rows.append(f"<div class='line del'>{tok_a(a_line, '')}</div>")
            for _ in range(i2 - i1):
                b_rows.append("<div class='line del'></div>")
        elif tag == "insert":
            for k in range(j2 - j1):
                b_line = b_lines[j1 + k]
                b_rows.append(f"<div class='line add'>{tok_b('', b_line)}</div>")
            for _ in range(j2 - j1):
                a_rows.append("<div class='line add'></div>")
        elif tag == "replace":
            h = max(i2 - i1, j2 - j1)
            for k in range(h):
                a_line = a_lines[i1 + k] if (i1 + k) < i2 else ""
                b_line = b_lines[j1 + k] if (j1 + k) < j2 else ""
                a_rows.append(f"<div class='line rep'>{tok_a(a_line, b_line) if a_line != '' else ''}</div>")
                b_rows.append(f"<div class='line rep'>{tok_b(a_line, b_line) if b_line != '' else ''}</div>")

    a_rows.append("</div></div>")
    b_rows.append("</div></div>")

    return f"<html><head>{css}</head><body>{''.join(a_rows)}{''.join(b_rows)}</body></html>"
# ===== [/REPLACE] ==============================================================

# ===== 置換ブロック: 縦並びサイド（A 上 / B 下）の差分ダイアログ =====
class DiffDialog(QDialog):
    def __init__(self, a_text: str, b_text: str, parent=None):
        super().__init__(parent)
        self.setWindowTitle("差分表示")
        self.resize(600, 300)

        # ルートレイアウト
        root = QVBoxLayout(self)
        root.setContentsMargins(8, 8, 8, 8)
        root.setSpacing(6)

        # タイトル + 区切り罫
        title = QLabel("差分（a → b）")
        title.setStyleSheet("font-weight:700; font-size:14px; color:#111;")
        root.addWidget(title)

        divider = QFrame()
        divider.setFrameShape(QFrame.HLine)
        divider.setStyleSheet("color:#dee2e6;")
        root.addWidget(divider)

        # 単一ビュー（上下スクロール一体）
        self.view = QTextEdit()
        self.view.setReadOnly(True)
        self.view.setFrameShape(QFrame.NoFrame)
        self.view.setContentsMargins(0, 0, 0, 0)
        self.view.setViewportMargins(0, 0, 0, 0)
        self.view.document().setDocumentMargin(0)
        self.view.setStyleSheet("QTextEdit { font-size:16px; }")
        root.addWidget(self.view, 1)

        # OK のみ
        btns = QDialogButtonBox(QDialogButtonBox.Ok, parent=self)
        btns.accepted.connect(self.accept)
        root.addWidget(btns)

        # HTML 生成して描画
        html = self._build_vertical_html(a_text or "", b_text or "")
        self.view.setHtml(html)

    # --- REPLACE: DiffDialog._build_vertical_html (HTMLエスケープ修正 & 最小化) ---
    def _build_vertical_html(self, a: str, b: str) -> str:
        import difflib, html
        a_lines = (a or "").splitlines()
        b_lines = (b or "").splitlines()
        sm = difflib.SequenceMatcher(a=a_lines, b=b_lines)
        ops = sm.get_opcodes()

        css = """
        <style>
        html, body, div, p { margin:0; padding:0; }
        body {
        font-family:'Yu Gothic UI','Noto Sans JP',sans-serif;
        color:#111; font-size:16px; line-height:1.80; letter-spacing:0.2px;
        }
        .page { padding-top:6px; padding-bottom:0; }
        .sec { margin:0 0 6px 0; } .sec:last-child { margin-bottom:0; }
        .sec-b { margin-top:12px; }
        .sec-head { font-weight:600; margin:0; line-height:0; }
        .sec-a .sec-head { color:#d9480f; }
        .sec-b .sec-head { color:#1c7ed6; }
        .diff { line-height:1.00; }
        table,.wrap { border-collapse:collapse; border-spacing:0; margin:0; }
        .wrap { width:100%; }
        .rail { width:18px; }
        .rail-a { background:#fa5252; } .rail-b { background:#4dabf7; }
        .body { padding:4px 10px 6px 10px; }
        .line { white-space:pre-wrap; padding:2px 4px; }
        .eq{} .add{ background:#e7f5ff; } .del{ background:#fff0f0; } .rep{ background:#fff7e6; }
        .tok-ins{ color:#1c7ed6; background:#d0ebff; border-radius:2px; }
        .tok-del{ color:#c92a2a; background:#ffe3e3; border-radius:2px; }
        </style>
        """

        def esc(s: str) -> str:
            return html.escape(s or "", quote=False)

        def tok_a(a_line: str, b_line: str) -> str:
            sm2 = difflib.SequenceMatcher(a=a_line or "", b=b_line or "")
            parts = []
            for op, i1, i2, j1, j2 in sm2.get_opcodes():
                a_seg = (a_line or "")[i1:i2]
                b_seg = (b_line or "")[j1:j2]
                if op == "equal":
                    parts.append(esc(a_seg))
                elif op in ("delete", "replace"):
                    if a_seg:
                        parts.append(f"<span class='tok-del'>{esc(a_seg)}</span>")
            return "".join(parts)

        def tok_b(a_line: str, b_line: str) -> str:
            sm2 = difflib.SequenceMatcher(a=a_line or "", b=b_line or "")
            parts = []
            for op, i1, i2, j1, j2 in sm2.get_opcodes():
                a_seg = (a_line or "")[i1:i2]
                b_seg = (b_line or "")[j1:j2]
                if op == "equal":
                    parts.append(esc(b_seg))
                elif op in ("insert", "replace"):
                    if b_seg:
                        parts.append(f"<span class='tok-ins'>{esc(b_seg)}</span>")
            return "".join(parts)

        a_rows = [
            "<div class='page sec sec-a'>",
            "<div class='diff'>",
            "<table class='wrap' cellspacing='0' cellpadding='0'><tr>",
            "<td class='rail rail-a'></td>",
            "<td class='body'>"
        ]
        b_rows = [
            "<div class='page sec sec-b'>",
            "<div class='diff'>",
            "<table class='wrap' cellspacing='0' cellpadding='0'><tr>",
            "<td class='rail rail-b'></td>",
            "<td class='body'>"
        ]

        for tag, i1, i2, j1, j2 in ops:
            if tag == "equal":
                for k in range(i2 - i1):
                    text = esc(a_lines[i1 + k])
                    a_rows.append(f"<div class='line eq'>{text}</div>")
                    b_rows.append(f"<div class='line eq'>{text}</div>")
            elif tag == "delete":
                for k in range(i2 - i1):
                    a_line = a_lines[i1 + k]
                    a_rows.append(f"<div class='line del'>{tok_a(a_line, '')}</div>")
                for _ in range(i2 - i1):
                    b_rows.append("<div class='line del'></div>")
            elif tag == "insert":
                for k in range(j2 - j1):
                    b_line = b_lines[j1 + k]
                    b_rows.append(f"<div class='line add'>{tok_b('', b_line)}</div>")
                for _ in range(j2 - j1):
                    a_rows.append("<div class='line add'></div>")
            elif tag == "replace":
                h = max(i2 - i1, j2 - j1)
                for k in range(h):
                    a_line = a_lines[i1 + k] if (i1 + k) < i2 else ""
                    b_line = b_lines[j1 + k] if (j1 + k) < j2 else ""
                    a_rows.append(f"<div class='line rep'>{tok_a(a_line, b_line) if a_line != '' else ''}</div>")
                    b_rows.append(f"<div class='line rep'>{tok_b(a_line, b_line) if b_line != '' else ''}</div>")

        a_rows += ["</td></tr></table>", "</div>", "</div>"]
        b_rows += ["</td></tr></table>", "</div>", "</div>"]
        return f"<html><head>{css}</head><body>{''.join(a_rows)}{''.join(b_rows)}</body></html>"
    # --- /REPLACE ---
# ------------------------------------------------------------
# Qt Models / Delegate / Style / Proxy
# ------------------------------------------------------------
# ===== ツールチップ遅延表示フィルタ（完全版：差し替え用） =====
from PySide6.QtCore import QObject, QTimer, QPoint, QEvent, Qt, QModelIndex
from PySide6.QtWidgets import QToolTip, QTableView
from PySide6.QtGui import QHelpEvent

class ToolTipDebouncer(QObject):
    """
    QTableView の ToolTip を少し遅らせて表示し、ホバー移動中の無駄な
    ToolTipRole 計算（重いHTML生成）を抑制します。
    """
    def __init__(self, view: QTableView, delay_ms: int = 200, parent=None):
        super().__init__(parent)
        self.view = view
        self.delay_ms = int(max(0, delay_ms))
        self.timer = QTimer(self)
        self.timer.setSingleShot(True)
        self._pending_index: Optional[QModelIndex] = None
        self._pending_global_pos: Optional[QPoint] = None
        self.timer.timeout.connect(self._show_tooltip)

    def _show_tooltip(self):
        idx = self._pending_index
        self._pending_index = None  # 先にクリア
        if not idx or not idx.isValid():
            return
        model = self.view.model()
        if not model:
            return
        try:
            text = model.data(idx, Qt.ToolTipRole)
        except Exception:
            text = None
        if text:
            pos = self._pending_global_pos or self.view.mapToGlobal(QPoint(0, 0))
            QToolTip.showText(pos, text, self.view)
        self._pending_global_pos = None

    def eventFilter(self, obj, ev):
        # ToolTip イベントだけ捕捉して遅延表示
        if ev.type() == QEvent.ToolTip and isinstance(ev, QHelpEvent):
            # viewport 座標で index を取得し、グローバル座標も保存
            idx = self.view.indexAt(ev.pos())
            self._pending_index = idx if idx.isValid() else None
            self._pending_global_pos = ev.globalPos()
            # 高速移動中は何度も再起動され、実行は発火時（＝静止時）のみ
            self.timer.start(self.delay_ms)
            return True  # 既定の即時ツールチップは抑止
        # ビューから離れたらキャンセル
        if ev.type() in (QEvent.Leave, QEvent.Hide):
            self.timer.stop()
            QToolTip.hideText()
            self._pending_index = None
            self._pending_global_pos = None
        return False
# ===== ここまで =====

class UnifiedModel(QAbstractTableModel):
    def __init__(self, df=pd.DataFrame(), parent=None):
        super().__init__(parent)
        self._df = df.copy()
        self._checks: List[bool] = [False] * len(self._df)
        self._col_idx = {}
        self._row_palette = [QColor("#D6E9FF"), QColor("#DFF5D8"), QColor("#FFF1C2"), QColor("#F1D4FF")]
        self._row_alpha = 36

    def setDataFrame(self, df: pd.DataFrame, copy: bool = True):
        self.beginResetModel()
        self._df = df.copy() if copy else df
        # 既存ロジック
        self._checks = [False] * len(self._df)
        self._col_idx = {name: i for i, name in enumerate(self._df.columns)}
        self.endResetModel()

    def _row_color_for_gid(self, gid: Optional[int]) -> Optional[QColor]:
        if gid is None or pd.isna(gid):
            return None
        gid = int(gid)
        c = self._row_palette[(gid - 1) % len(self._row_palette)]
        qc = QColor(c)
        qc.setAlpha(self._row_alpha)
        return qc

    def rowCount(self, parent=QModelIndex()):
        return len(self._df)

    def columnCount(self, parent=QModelIndex()):
        return 1 + len(self._df.columns)

    def headerData(self, section, orientation, role=Qt.DisplayRole):
        if role != Qt.DisplayRole:
            return None
        if orientation == Qt.Horizontal:
            if section == 0:
                return "✓"
            name = str(self._df.columns[section - 1])
            if name == "a_count":
                return "a数"
            if name == "b_count":
                return "b数"
            if name == "score":
                return "類似度"
            if name.lower() == "scope":
                return "対象"
            return "一致要因" if name == "一致要因" else name
        return str(section + 1)

    def flags(self, index: QModelIndex):
        if not index.isValid():
            return Qt.NoItemFlags
        flags = Qt.ItemIsEnabled | Qt.ItemIsSelectable
        if index.column() == 0:
            flags |= Qt.ItemIsUserCheckable
        return Qt.ItemFlags(flags)

    def data(self, index: QModelIndex, role=Qt.DisplayRole):
        if not index.isValid():
            return None
        r, c = index.row(), index.column()

        # 0列はチェックボックス
        if role == Qt.CheckStateRole and c == 0:
            return Qt.Checked if self._checks[r] else Qt.Unchecked

        # 表示テキスト
        if role == Qt.DisplayRole:
            if c == 0:
                return None
            v = self._df.iat[r, c - 1]
            if pd.isna(v):
                return ""
            # ★ score 列だけは小数点第3位で固定表示
            try:
                col_name = self._df.columns[c - 1]
            except Exception:
                col_name = ""
            if col_name == "score":
                try:
                    return f"{float(v):.3f}"
                except Exception:
                    return str(v)
            return str(v)

        # 行背景（gidによる着色）
        if role == Qt.BackgroundRole:
            try:
                if "gid" in self._col_idx:
                    gid_val = self._df.iat[r, self._col_idx["gid"]]
                    qc = self._row_color_for_gid(gid_val)
                    if qc:
                        return qc
            except Exception:
                pass

        # 数値系は右寄せ
        if role == Qt.TextAlignmentRole:
            col_name = self._df.columns[c - 1] if c > 0 else ""
            if col_name in ("gid", "字数", "a_count", "b_count", "score", "a数", "b数"):
                return Qt.AlignRight | Qt.AlignVCenter

        return None

    def setData(self, index: QModelIndex, value, role=Qt.EditRole):
        if not index.isValid():
            return False
        if index.column() == 0 and role == Qt.CheckStateRole:
            self._checks[index.row()] = (value == Qt.Checked)
            self.dataChanged.emit(index, index, [Qt.CheckStateRole])
            return True
        return False

    def check_all(self, checked: bool):
        if not self._checks:
            return
        for i in range(len(self._checks)):
            self._checks[i] = checked
        self.dataChanged.emit(self.index(0,0), self.index(len(self._checks)-1,0), [Qt.CheckStateRole])

# ===== [REPLACE] CheckBoxDelegate（センター描画 + sizeHint + __init__） =====
class CheckBoxDelegate(QStyledItemDelegate):
    def __init__(self, parent=None):
        super().__init__(parent)

    def sizeHint(self, option, index):
        # セルの推奨サイズ（高さに追随しつつ 16px 基準）
        h = max(16, option.rect.height())
        return QSize(h, h)

    def paint(self, painter: QPainter, option, index):
        if index.column() != 0:
            return super().paint(painter, option, index)

        state = index.model().data(index, Qt.CheckStateRole)
        rect = option.rect
        size = min(rect.height(), 16)  # 16px を目安
        x = rect.x() + (rect.width() - size) // 2
        y = rect.y() + (rect.height() - size) // 2
        r = QRect(x, y, size, size)

        painter.save()
        painter.setRenderHint(QPainter.Antialiasing, True)
        # 枠
        pen = QPen(QColor("#111")); pen.setWidth(1)
        painter.setPen(pen); painter.setBrush(QColor("#fff"))
        painter.drawRoundedRect(r, 2, 2)
        # チェック
        if state == Qt.Checked:
            pen = QPen(QColor("#111")); pen.setWidth(2)
            painter.setPen(pen)
            x1 = r.x() + int(size*0.22); y1 = r.y() + int(size*0.56)
            x2 = r.x() + int(size*0.46); y2 = r.y() + int(size*0.80)
            x3 = r.x() + int(size*0.80); y3 = r.y() + int(size*0.28)
            painter.drawLine(x1,y1,x2,y2); painter.drawLine(x2,y2,x3,y3)
        painter.restore()
# ===== [/REPLACE] ===========================================================

class HighContrastCheckboxStyle(QProxyStyle):
    def pixelMetric(self, metric, option=None, widget=None):
        if metric in (QStyle.PM_IndicatorWidth, QStyle.PM_IndicatorHeight):
            return 16
        return super().pixelMetric(metric, option, widget)

    def drawPrimitive(self, elem, opt, painter: QPainter, widget=None):
        if elem == QStyle.PE_IndicatorCheckBox:
            r = opt.rect; size = min(r.height(), 16)
            x = r.x() + (r.width() - size) // 2; y = r.y() + (r.height() - size) // 2
            box = QRect(x,y,size,size)
            painter.save(); painter.setRenderHint(QPainter.Antialiasing, True)
            pen = QPen(QColor("#111")); pen.setWidth(1); painter.setPen(pen); painter.setBrush(QColor("#fff"))
            painter.drawRoundedRect(box, 2, 2)
            if (opt.state & QStyle.State_On) or (opt.state & QStyle.State_NoChange):
                pen = QPen(QColor("#111")); pen.setWidth(2); painter.setPen(pen)
                x1 = box.x()+int(size*0.22); y1 = box.y()+int(size*0.56)
                x2 = box.x()+int(size*0.46); y2 = box.y()+int(size*0.80)
                x3 = box.x()+int(size*0.80); y3 = box.y()+int(size*0.28)
                painter.drawLine(x1,y1,x2,y2); painter.drawLine(x2,y2,x3,y3)
            painter.restore(); return
        super().drawPrimitive(elem, opt, painter, widget)

class PandasModel(QAbstractTableModel):
    def __init__(self, df=pd.DataFrame(), parent=None):
        super().__init__(parent)
        self._df = df.copy()
    def setDataFrame(self, df: pd.DataFrame, copy: bool = True):
        self.beginResetModel()
        self._df = df.copy() if copy else df
        self.endResetModel()
    def rowCount(self, parent=QModelIndex()):
        return len(self._df)
    def columnCount(self, parent=QModelIndex()):
        return len(self._df.columns)
    def data(self, index: QModelIndex, role=Qt.DisplayRole):
        if not index.isValid():
            return None
        if role == Qt.DisplayRole:
            v = self._df.iat[index.row(), index.column()]
            return "" if pd.isna(v) else str(v)
        return None
    def headerData(self, section, orientation, role=Qt.DisplayRole):
        if role != Qt.DisplayRole:
            return None
        return str(self._df.columns[section]) if orientation == Qt.Horizontal else str(section + 1)

class UnifiedFilterProxy(QSortFilterProxyModel):
    def __init__(self):
        super().__init__()
        self._text = ""
        self._reasons: Set[str] = set()
        self._scopes: Set[str] = set()
        self._hide_prefix_edge = False
        self._hide_suffix_edge = False
        self._hide_numeric_only = False
        self._hide_contains = False   # 内包関係除外
        # ▼ 新規：短文（aの文字数）フィルタ
        self._hide_short_enabled = False
        self._hide_short_len = 0

        self._collator = QCollator()
        self._collator.setNumericMode(True)
        self._collator.setCaseSensitivity(Qt.CaseInsensitive)

        # 列キャッシュ
        self._cols: Dict[str, int] = {}

    # ---- 外部API（既存） ----
    def setTextFilter(self, text: str):
        self._text = text or ""
        self.invalidateFilter()

    def setReasonFilter(self, reasons: Set[str]):
        self._reasons = set(reasons or [])
        self.invalidateFilter()

    def setScopeFilter(self, scopes: Set[str]):
        self._scopes = set(scopes or [])
        self.invalidateFilter()

    def setHidePrefixEdge(self, hide: bool):
        self._hide_prefix_edge = bool(hide)
        self.invalidateFilter()

    def setHideSuffixEdge(self, hide: bool):
        self._hide_suffix_edge = bool(hide)
        self.invalidateFilter()

    def setHideNumericOnly(self, hide: bool):
        self._hide_numeric_only = bool(hide)
        self.invalidateFilter()

    # ---- 新規API：短文フィルタ ----
    def setHideShortEnabled(self, enabled: bool):
        self._hide_short_enabled = bool(enabled)
        self.invalidateFilter()

    def setHideShortLength(self, n: int):
        try:
            n_int = int(n)
        except Exception:
            n_int = 0
        self._hide_short_len = max(0, n_int)
        self.invalidateFilter()

    # ---- 列キャッシュ構築 ----
    def setSourceModel(self, model):
        super().setSourceModel(model)
        if model:
            model.modelReset.connect(self._rebuild_cols)
            model.layoutChanged.connect(self._rebuild_cols)
        self._rebuild_cols()

    def _rebuild_cols(self):
        self._cols.clear()
        m = self.sourceModel()
        if not m:
            return
        for c in range(m.columnCount()):
            h = m.headerData(c, Qt.Horizontal, Qt.DisplayRole)
            if isinstance(h, str):
                name = h.strip()
                # ▼ “字数” ほかに __contains__ を追加
                if name in {
                    "a", "b", "一致要因", "理由", "reason",
                    "対象", "scope", "target",
                    "端差", "数字以外一致", "字数",
                    "__contains__",   # ← これを追加
                }:
                    self._cols[name] = c

    def _col(self, *names: str) -> int:
        for nm in names:
            if nm in self._cols:
                return self._cols[nm]
        return -1

    # ---- フィルタ本体 ----
    def filterAcceptsRow(self, row, parent):
        m = self.sourceModel()
        if not m:
            return True

        # 1) scope
        if self._scopes:
            c_scope = self._col("対象", "scope", "target")
            if c_scope >= 0:
                v = m.data(m.index(row, c_scope, parent), Qt.DisplayRole) or ""
                if str(v) not in self._scopes:
                    return False

        # 2) 全文フィルタ
        if self._text:
            needle = self._text.lower()
            cols = [
                self._col("a"), self._col("b"),
                self._col("一致要因", "理由", "reason"),
                self._col("対象", "scope", "target")
            ]
            cols = [c for c in cols if c >= 0]
            if not cols:
                cols = list(range(m.columnCount()))
            hit = False
            for c in cols:
                v = m.data(m.index(row, c, parent), Qt.DisplayRole)
                if v is not None and needle in str(v).lower():
                    hit = True
                    break
            if not hit:
                return False

        # 3) 端差（前/後1字）
        if self._hide_prefix_edge or self._hide_suffix_edge:
            c_edge = self._col("端差")
            edge_val = ""
            if c_edge >= 0:
                edge_val = (m.data(m.index(row, c_edge, parent), Qt.DisplayRole) or "").strip()
            else:
                ca, cb = self._col("a"), self._col("b")
                if ca >= 0 and cb >= 0:
                    a_txt = m.data(m.index(row, ca, parent), Qt.DisplayRole) or ""
                    b_txt = m.data(m.index(row, cb, parent), Qt.DisplayRole) or ""
                    try:
                        edge_val = classify_edge_diff_jp(str(a_txt), str(b_txt))
                    except Exception:
                        edge_val = ""
            if self._hide_prefix_edge and edge_val in ("前1字有無", "前1字違い"):
                return False
            if self._hide_suffix_edge and edge_val in ("後1字有無", "後1字違い"):
                return False

        # 4) 数字以外一致
        if self._hide_numeric_only:
            c_num = self._col("数字以外一致")
            is_num_only = False
            if c_num >= 0:
                v = (m.data(m.index(row, c_num, parent), Qt.DisplayRole) or "").strip()
                is_num_only = (v == "数字以外一致")
            else:
                ca, cb = self._col("a"), self._col("b")
                if ca >= 0 and cb >= 0:
                    a_txt = m.data(m.index(row, ca, parent), Qt.DisplayRole) or ""
                    b_txt = m.data(m.index(row, cb, parent), Qt.DisplayRole) or ""
                    try:
                        is_num_only = is_numeric_only_diff(str(a_txt), str(b_txt))
                    except Exception:
                        is_num_only = False
            if is_num_only:
                return False

        # 4.5) ▼ 新規：短文フィルタ（a の文字数が閾値以下なら非表示）
        if self._hide_short_enabled and self._hide_short_len > 0:
            # まず “字数” 列があればそれを使う（Workerで付与済み想定）
            c_len = self._col("字数")
            a_len = None
            if c_len >= 0:
                try:
                    v = m.data(m.index(row, c_len, parent), Qt.DisplayRole)
                    a_len = int(str(v)) if v is not None else None
                except Exception:
                    a_len = None
            # なければ a 列から長さを算出
            if a_len is None:
                ca = self._col("a")
                if ca >= 0:
                    a_txt = m.data(m.index(row, ca, parent), Qt.DisplayRole) or ""
                    a_len = len(str(a_txt))
            if a_len is not None and a_len <= self._hide_short_len:
                return False

        # 4.7) ▼ 内包関係除外（a ∈ b もしくは b ∈ a）
        if self._hide_contains:
            # まずは前計算列があればそれを使う（高速）
            c_flag = self._col("__contains__")
            if c_flag >= 0:
                v = m.data(m.index(row, c_flag, parent), Qt.DisplayRole)
                # pandasのbool/NaN/文字列True/Falseを素朴に判定
                sv = ("" if v is None else str(v)).strip().lower()
                if sv in ("true", "1"):
                    return False
            else:
                # フォールバック：従来の“その場計算”（列が無い場合のみ実行）
                ca, cb = self._col("a"), self._col("b")
                if ca >= 0 and cb >= 0:
                    a_txt = m.data(m.index(row, ca, parent), Qt.DisplayRole) or ""
                    b_txt = m.data(m.index(row, cb, parent), Qt.DisplayRole) or ""
                    try:
                        sa = nfkc(str(a_txt)).lower()
                        sb = nfkc(str(b_txt)).lower()
                        # 完全一致は内包扱いにしない
                        if sa and sb and sa != sb and (sa.find(sb) != -1 or sb.find(sa) != -1):
                            return False
                    except Exception:
                        pass


        # 5) 一致要因
        if self._reasons:
            c_reason = self._col("一致要因", "理由", "reason")
            if c_reason >= 0:
                v = m.data(m.index(row, c_reason, parent), Qt.DisplayRole) or ""
                if str(v) not in self._reasons:
                    return False

        return True

    # ---- 並び替え（既存） ----
    def lessThan(self, left: QModelIndex, right: QModelIndex) -> bool:
        m = self.sourceModel()
        if not m:
            return super().lessThan(left, right)

        v1 = m.data(left, Qt.DisplayRole)
        v2 = m.data(right, Qt.DisplayRole)

        def to_num(x):
            if x is None:
                return None
            s = str(x).strip()
            try:
                return float(s)
            except Exception:
                return None

        n1, n2 = to_num(v1), to_num(v2)
        if n1 is not None and n2 is not None:
            return n1 < n2

        s1 = "" if v1 is None else str(v1)
        s2 = "" if v2 is None else str(v2)
        return self._collator.compare(s1, s2) < 0

    # --- 追加: 内包関係除外のオン/オフ ---
    def setHideContainment(self, hide: bool):
        self._hide_contains = bool(hide)
        self.invalidateFilter()


# ------------------------------------------------------------
# Worker（★進捗配分を後半重めに再設計）
# ------------------------------------------------------------
# ========== AnalyzerWorker（_subprogress_factory を削除した版） ==========
class AnalyzerWorker(QObject):
    finished = Signal(dict, str)
    progress = Signal(int)
    progress_text = Signal(str)
    
    def __init__(self, files, use_mecab, read_th, char_th):
        super().__init__()
        self.files = files
        self.use_mecab = use_mecab
        self.read_th = read_th
        self.char_th = char_th
        self.top_k_lex = 4000
        self.min_count_lex = 1
        self.min_len = 1
        self.max_len = 120
        
        # 進捗のデバウンス（過剰なシグナル発火抑制）
        self._current_progress = 0
        self._last_emit_time = 0
        self._min_emit_interval = 0.10  # 100ms以上空ける（過剰更新防止）

    def _emit(self, v: int):
        """進捗を発火（重複＆過剰更新抑制付き）"""
        import time
        now = time.time()
        new_val = max(0, min(100, int(v)))
        # 同じ値または短時間での連続発火を抑制
        if new_val == self._current_progress and (now - self._last_emit_time) < self._min_emit_interval:
            return
        self._current_progress = new_val
        self._last_emit_time = now
        self.progress.emit(new_val)

    # ===== run() 全体で「progress_cb」を使わず、直接 _emit() する =====
    def run(self):
        try:
            # 0) PDF抽出（0–12%）
            pages = []
            n_files = max(1, len(self.files))
            for i, f in enumerate(self.files, 1):
                self.progress_text.emit("PDF抽出中…")
                pages += extract_pages(f, flatten_for_nlp=True, v_strategy="strict_y")
                self._emit(int(12 * i / n_files))
            if not pages:
                raise RuntimeError("PDFからテキストが取得できませんでした。")

            # 1) 細粒度トークン参照 / 語彙構築（12–20%）
            self.progress_text.emit("細粒度トークン収集（参照）中…")
            if self.use_mecab and HAS_MECAB:
                tokens_fine = extract_candidates_regex(
                    pages, self.min_len, self.max_len, self.min_count_lex
                )
            else:
                raise RuntimeError('MeCab (fugashi/unidic-lite) が必要です。 pip install "fugashi[unidic-lite]"')
            self._emit(16)

            self.progress_text.emit("細粒度語彙の構築中…")
            df_lex = collect_lexicon_general(
                pages, top_k=self.top_k_lex, min_count=self.min_count_lex
            )
            self._emit(20)

            # 2) 細粒度ペア生成（語彙ペア）（20–45%）
            # progress_cb を渡さずに直接実行
            df_pairs_lex_general = build_synonym_pairs_general(
                df_lex,
                read_sim_th=self.read_th,
                char_sim_th=self.char_th,
                scope="語彙",
            )
            self._emit(45)

            # 3) 複合語候補抽出（45–55%）
            self.progress_text.emit("複合語候補抽出中…")
            tokens_compound_main = extract_candidates_compound_alljoin(
                pages,
                min_len=self.min_len, max_len=self.max_len,
                min_count=self.min_count_lex, top_k=0, use_mecab=True
            )
            tokens_compound_fb = extract_candidates_compound_alljoin(
                pages,
                min_len=self.min_len, max_len=self.max_len,
                min_count=self.min_count_lex, top_k=0, use_mecab=False
            )
            c_all = Counter()
            for w, c in tokens_compound_main:
                c_all[w] += c
            for w, c in tokens_compound_fb:
                c_all[w] += c
            tokens_compound = sorted(c_all.items(), key=lambda x: (-x[1], x[0]))
            self._emit(55)

            # 4) 複合語ペア生成（55–78%）
            df_pairs_compound = build_synonym_pairs_char_only(
                tokens_compound,
                char_sim_th=self.char_th,
                top_k=self.top_k_lex, scope="複合語",
                read_sim_th=self.read_th,
                use_processes=False,       # 並列を無効化
                max_workers=1              # 念のためワーカ数も1に
            )

            self._emit(78)

            # 5) 文・文節候補抽出（GiNZA 1パス）（78–83%）
            self.progress_text.emit("文・文節候補抽出中…")
            tokens_bunsetsu, tokens_sentence = extract_candidates_sentence_bunsetsu_onepass(
                pages,
                bun_min_len=self.min_len, bun_max_len=self.max_len,
                sent_min_len=self.min_len, sent_max_len=max(self.max_len, 300),
                min_count=self.min_count_lex, top_k=0
            )
            self._emit(83)

            # 6) 文節ペア生成（83–92%）
            df_pairs_bunsetsu = build_synonym_pairs_char_only(
                tokens_bunsetsu,
                char_sim_th=self.char_th,
                top_k=self.top_k_lex, scope="文節",
                read_sim_th=self.read_th,
            )
            self._emit(92)

            # 7) 文章ペア生成（92–97%）
            df_pairs_sentence = build_synonym_pairs_char_only(
                tokens_sentence,
                char_sim_th=self.char_th,
                top_k=self.top_k_lex, scope="文章",
                read_sim_th=self.read_th
            )
            self._emit(97)

            # 8) 統合・再採点・仕上げ（97–100%）
            self.progress_text.emit("候補統合中…")
            df_unified = unify_pairs(
                df_pairs_lex_general, df_pairs_compound, df_pairs_bunsetsu, df_pairs_sentence
            )
            self._emit(97)

            # 読みキャッシュのプリウォーム
            surfaces = []
            if not df_unified.empty and {"a", "b"}.issubset(df_unified.columns):
                surfaces = pd.unique(pd.concat([df_unified["a"], df_unified["b"]]).astype("string")).tolist()
            self.progress_text.emit("読みキャッシュ準備中…")
            prewarm_reading_caches(surfaces, keepchoon=True)

            # 再採点・再分類
            df_unified = recalibrate_reading_like_scores(df_unified, read_th=self.read_th)
            df_unified = enforce_combined_similarity_score(df_unified, keep_backup=False, drop_existing_backup=True)
            df_unified = reclassify_basic_for_reading_eq(df_unified, eps=0.0005)
            df_unified = reclassify_reading_equal_formdiff(df_unified)
            df_unified = sanitize_reading_same(df_unified)

            # スコア丸め・単独仮名除去
            if df_unified is not None and not df_unified.empty and "score" in df_unified.columns:
                df_unified["score"] = pd.to_numeric(df_unified["score"], errors="coerce").round(3)
            if not df_unified.empty and "a" in df_unified.columns:
                df_unified = df_unified[~df_unified["a"].apply(is_single_kana_char)].reset_index(drop=True)
            self._emit(98)

            # トークン一覧整形
            parts = []
            if len(tokens_fine) > 0:
                df_fine = pd.DataFrame(tokens_fine, columns=["token", "count"]); df_fine["type"] = "fine"; parts.append(df_fine)
            if len(tokens_compound) > 0:
                df_comp = pd.DataFrame(tokens_compound, columns=["token", "count"]); df_comp["type"] = "compound"; parts.append(df_comp)
            if len(tokens_bunsetsu) > 0:
                df_bun  = pd.DataFrame(tokens_bunsetsu, columns=["token", "count"]); df_bun["type"] = "bunsetsu"; parts.append(df_bun)
            if len(tokens_sentence) > 0:
                df_sent = pd.DataFrame(tokens_sentence, columns=["token", "count"]); df_sent["type"] = "sentence"; parts.append(df_sent)
            df_tokens = pd.concat(parts, ignore_index=True) if parts else pd.DataFrame(columns=["type","token","count"])
            if not df_tokens.empty:
                df_tokens = df_tokens["type token count".split()].sort_values(["type","count","token"], ascending=[True, False, True])

            # 表示ラベル変換・各種前計算（端差/数字以外一致/内包/字数）
            df_unified = apply_reason_ja(df_unified)
            try:
                if df_unified is not None and not df_unified.empty and "a" in df_unified.columns and "b" in df_unified.columns:
                    df_unified["端差"] = _edge_labels_vectorized(df_unified)
                    df_unified["数字以外一致"] = _numeric_only_label_vectorized(df_unified)
            except Exception:
                pass

            try:
                if df_unified is not None and not df_unified.empty and {"a","b"}.issubset(df_unified.columns):
                    sa = df_unified["a"].astype("string").map(lambda x: nfkc(x or "").lower())
                    sb = df_unified["b"].astype("string").map(lambda x: nfkc(x or "").lower())
                    def _contains_pair(a: str, b: str) -> bool:
                        if not a or not b:
                            return False
                        if a == b:
                            return False
                        return (a.find(b) != -1) or (b.find(a) != -1)
                    df_unified["__contains__"] = [
                        _contains_pair(aa, bb) for aa, bb in zip(sa.tolist(), sb.tolist())
                    ]
                else:
                    if df_unified is not None and not df_unified.empty:
                        df_unified["__contains__"] = False
            except Exception:
                if df_unified is not None and not df_unified.empty:
                    df_unified["__contains__"] = False

            if not df_unified.empty and "a" in df_unified.columns:
                try:
                    df_unified["字数"] = df_unified["a"].astype("string").str.len().fillna(0).astype(int)
                except Exception:
                    df_unified["字数"] = df_unified["a"].astype(str).str.len().fillna(0).astype(int)

            cols = list(df_unified.columns)
            pref = [c for c in ["字数", "a", "b"] if c in cols]
            rest = [c for c in cols if c not in pref]
            df_unified = df_unified[pref + rest]

            self.progress_text.emit("グループ割当中…")
            df_groups, surf2gid, gid2members = build_variation_groups(df_unified, df_lex)
            if not df_unified.empty and "a" in df_unified.columns:
                df_unified["gid"] = df_unified["a"].map(
                    lambda x: surf2gid.get(x) if isinstance(x, str) and x else None
                )

            # 完了
            self._emit(100)
            self.progress_text.emit("仕上げ中…")
            self.finished.emit(
                {"unified": df_unified, "tokens": df_tokens, "groups": df_groups,
                 "surf2gid": surf2gid, "gid2members": gid2members},
                ""
            )

        except Exception as e:
            self.finished.emit({}, str(e))


# ------------------------------------------------------------
# Theme / Drag helpers / GUI
# ------------------------------------------------------------
def apply_light_theme(app: QApplication, base_font_size=11):
    family = "Yu Gothic UI"
    app.setFont(QFont(family, base_font_size))

    pal = QPalette()
    pal.setColor(QPalette.Window, QColor("#fff"))
    pal.setColor(QPalette.Base, QColor("#fff"))
    pal.setColor(QPalette.AlternateBase, QColor("#f8f9fa"))
    pal.setColor(QPalette.Text, QColor("#111"))
    pal.setColor(QPalette.WindowText, QColor("#111"))
    pal.setColor(QPalette.Button, QColor("#fff"))
    pal.setColor(QPalette.ButtonText, QColor("#111"))
    pal.setColor(QPalette.Highlight, QColor("#e7f5ff"))
    pal.setColor(QPalette.HighlightedText, QColor("#0b7285"))

    # ★★★ 追加：ツールチップの白背景・文字色をパレットで固定
    pal.setColor(QPalette.ToolTipBase, QColor("#ffffff"))
    pal.setColor(QPalette.ToolTipText, QColor("#111111"))

    app.setPalette(pal)
    app.setStyle(HighContrastCheckboxStyle(app.style()))

    # ★★★ 追加：QToolTip の背景・枠線・文字色をスタイルで固定
    # 既存の app.setStyleSheet に QToolTip のルールを追記しています
    app.setStyleSheet("""
    QMainWindow, QWidget { background:#fff; color:#111; }
    QLabel, QGroupBox, QCheckBox { color:#111; }

    /* GroupBoxの余白を詰める */
    QGroupBox {
        border:1px solid #e9ecef; border-radius:8px;
        margin-top:12px;
    }
    QGroupBox::title {
        subcontrol-origin: margin; subcontrol-position: top left;
        padding:2px 6px; font-weight:600;
    }

    /* 入力のパディング縮小 */
    QLineEdit {
        background:#fff; color:#111;
        border:1px solid #dee2e6; border-radius:6px;
        padding:4px 6px; min-height: 24px;
    }
    QLineEdit:focus { border:1px solid #4dabf7; }

    QPushButton {
        background:#4dabf7; color:#fff; border-radius:8px;
        padding:6px 10px; border:1px solid #339af0;
    }
    QPushButton:hover { background:#339af0; }
    QPushButton:disabled { background:#cfd4da; color:#888; border-color:#cfd4da; }
    #btnRun {
        background:#f59f00; border-color:#f08c00;
        font-weight:700; font-size:13px; padding:8px 12px; min-height:36px;
    }
    #btnRun:hover { background:#f08c00; }

    QTabWidget::pane { border:1px solid #dee2e6; background:#fff; }
    QTabBar::tab {
        background:#f1f3f5; color:#111; padding:4px 8px; margin:2px; border-radius:6px;
    }
    QTabBar::tab:selected { background:#4dabf7; color:#fff; }

    QTableView {
        background:#fff; color:#111; gridline-color:#dee2e6;
        alternate-background-color:#f8f9fa;
        selection-background-color:#e7f5ff; selection-color:#0b7285;
    }
    QHeaderView::section {
        background:#f1f3f5; color:#212529;
        padding:4px 6px; border:1px solid #dee2e6; font-weight:600;
    }
    QTableView QTableCornerButton::section { background:#f1f3f5; border:1px solid #dee2e6; }

    QProgressBar { border:1px solid #dee2e6; border-radius:6px; text-align:center; height:16px; }
    QProgressBar::chunk { background:#4dabf7; }

    /* コンパクト化用：行間を詰める */
    QCheckBox { spacing:6px; }

    /* ★★★ ここが重要：ツールチップを白背景＋淡い枠に固定（ダークテーマでも白に）
       必要なら背景を #f8f9fa に変えると淡いグレーになります。 */
    QToolTip {
        background-color: #ffffff;
        color: #111111;
        border: 1px solid #dee2e6;
        padding: 6px 8px;
        border-radius: 6px;
        font-family: 'Yu Gothic UI','Noto Sans JP',sans-serif;
        font-size: 12px;
    }
    """)

def extract_pdf_paths_from_urls(urls) -> List[str]:
    out = []
    for u in urls:
        p = u.toLocalFile()
        if not p:
            continue
        if os.path.isdir(p):
            for root, _, files in os.walk(p):
                for fn in files:
                    if fn.lower().endswith(".pdf"):
                        out.append(os.path.join(root, fn))
        elif p.lower().endswith(".pdf"):
            out.append(p)
    return sorted(dict.fromkeys(out))

# ===== [REPLACE] DropArea（高さ固定を撤回：固定高さ指定なし） =====
class DropArea(QTextEdit):
    filesChanged = Signal(list)
    def __init__(self):
        super().__init__(); self.setAcceptDrops(True); self.setReadOnly(True)
        self.setPlaceholderText("ここにPDFをドラッグ＆ドロップしてください（複数可）")
        # 見た目はそのまま／高さは固定しない（必要なら最小高さだけ）
        self.setStyleSheet(
            "QTextEdit{border:2px dashed #4dabf7; border-radius:10px; "
            "background:#fff; padding:10px; color:#111; min-height:100px;}"
        )

    def dragEnterEvent(self, e):
        e.acceptProposedAction() if e.mimeData().hasUrls() else e.ignore()
    def dragMoveEvent(self, e):
        e.acceptProposedAction() if e.mimeData().hasUrls() else e.ignore()
    def dropEvent(self, e):
        if not e.mimeData().hasUrls():
            e.ignore(); return
        files = extract_pdf_paths_from_urls(e.mimeData().urls())
        if files:
            self.setText("\n".join(files))
            self.filesChanged.emit(files)
        e.acceptProposedAction()
# ===== [/REPLACE] =====================================================

# =========================================================
# MainWindow（まるごと差し替え）
# =========================================================
class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("PDF 表記ゆれチェック [ver.1.61]")
        self.resize(1000, 700)

        # ---- レイアウト骨格 ----
        central = QWidget()
        self.setCentralWidget(central)
        root = QHBoxLayout(central)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        # ========== 左パネル ==========
        left_panel = QWidget()
        left_panel.setFixedWidth(280)
        left = QVBoxLayout(left_panel)
        left.setContentsMargins(8, 8, 8, 8)
        left.setSpacing(8)

        # 入力PDF
        gb_files = QGroupBox("入力PDF")
        v_files = QVBoxLayout(gb_files)
        # ===== [PATCH] 入力PDFレイアウトの下余白を小さく =====
        v_files.setContentsMargins(8, 8, 8, 4)  # 左,上,右,下
        v_files.setSpacing(6)
        # ===== [/PATCH] =================================
        self.drop = DropArea()
        v_files.addWidget(self.drop)
        row = QHBoxLayout()
        self.btn_browse = QPushButton("ファイルを選択")
        self.btn_clear = QPushButton("クリア")
        row.addWidget(self.btn_browse)
        row.addWidget(self.btn_clear)
        v_files.addLayout(row)
        # ===== [PATCH] ファイル一覧リストを “3行表示” に固定して縦幅を狭く =====
        self.list_files = QListWidget()
        self.list_files.setAlternatingRowColors(True)
        self.list_files.setUniformItemSizes(True)  # 行高の計算を安定化
        self.list_files.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)

        # 3行ぶんの高さに抑える（フォントに応じて自動計算）
        visible_rows = 3
        fm = self.list_files.fontMetrics()
        row_h = max(18, fm.height() + 6)  # 行の目安（字高 + パディング）
        frame = self.list_files.frameWidth() * 2
        # 上下の余白を少なめに見積もって調整
        self.list_files.setFixedHeight(visible_rows * row_h + frame + 2)

        # 件数ラベルも余白を詰めてコンパクトに
        self.lbl_count = QLabel("選択ファイル：0 件")
        self.lbl_count.setStyleSheet("color:#666; margin-top:2px;")
        # ===== [/PATCH] =======================================================
        v_files.addWidget(self.list_files)
        v_files.addWidget(self.lbl_count)

        # 設定
        gb_params = QGroupBox("設定")
        form = QFormLayout(gb_params)
        self.dsb_read = QDoubleSpinBox()
        self.dsb_read.setRange(0.0, 1.0)
        self.dsb_read.setSingleStep(0.05)
        self.dsb_read.setValue(0.90)
        self.dsb_char = QDoubleSpinBox()
        self.dsb_char.setRange(0.0, 1.0)
        self.dsb_char.setSingleStep(0.05)
        self.dsb_char.setValue(0.85)
        form.addRow("読み 類似しきい値", self.dsb_read)
        form.addRow("文字 類似しきい値", self.dsb_char)

        # ===== [PATCH] 設定：NLPテキスト保存トグルを追加 =====
        #self.chk_save_nlp = QCheckBox("NLPテキスト保存")
        #self.chk_save_nlp.setChecked(False)
        #self.chk_save_nlp.setToolTip("MeCab / GiNZA に渡す直前のテキストを .txt で保存します（ON時）。")
        #form.addRow("テキスト保存", self.chk_save_nlp)
        #self.chk_save_nlp.stateChanged.connect(self.on_save_nlp_changed)


        # 実行・進捗
        self.btn_run = QPushButton("解析開始")
        self.btn_run.setObjectName("btnRun")
        self.progress = QProgressBar()
        self.progress.setRange(0, 100)
        self.progress.setValue(0)
        self.lbl_stage = QLabel("待機中")
        self.lbl_elapsed = QLabel("経過 00:00")

        # 🆕 進捗バーの滑らか化用タイマー
        self._progress_smooth_timer = QTimer(self)
        self._progress_smooth_timer.setInterval(50)  # 50ms = 20fps
        self._progress_smooth_timer.timeout.connect(self._smooth_progress_update)
        
        self._target_progress = 0  # 目標進捗値
        self._current_smooth_progress = 0.0  # 現在の滑らか化された進捗
        self._progress_smooth_speed = 0.3  # 追従速度(0-1の間、大きいほど速い)

        left.addWidget(gb_files)
        left.addWidget(gb_params)
        # ===== [PATCH] 差分プレビュー領域を少し狭く（高さ/余白/パディング） =====
        gb_diff = QGroupBox("簡易差分")
        v_diff = QVBoxLayout(gb_diff)
        # 余白と間隔を詰める（下方向をやや強め）
        v_diff.setContentsMargins(8, 4, 8, 6)  # L, T, R, B
        v_diff.setSpacing(4)

        self.diff_view = QTextBrowser()
        self.diff_view.setOpenExternalLinks(False)
        self.diff_view.setReadOnly(True)

        # パディングを小さく・枠は維持、背景は白
        self.diff_view.setStyleSheet(
            "QTextBrowser{background:#fff; border:0px solid #dee2e6; "
            "border-radius:6px; padding:10px 6px;}"  # ← 6px/8px → 4px/6px に縮小
        )

        # 高さは“上限”を付けてコンパクトに（必要なら数値を微調整）
        self.diff_view.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Maximum)
        self.diff_view.setMaximumHeight(180)  # ← 140〜180でお好み調整

        # スクロールバーは必要時のみ表示（高さ上限内でスクロール）
        self.diff_view.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self.diff_view.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)

        # 初期メッセージ（そのまま or 省略済み）
        self.diff_view.setHtml(
            "<html><body style='font-family:Yu Gothic UI, Noto Sans JP, sans-serif; "
            "color:#666; font-size:12px; margin:4px 6px;'>"
            "表のセルを選択すると、簡易差分をここに表示します。"
            "</body></html>"
        )

        v_diff.addWidget(self.diff_view)

        # GroupBox 自体も“高さを取りすぎない”ように最大方針
        gb_diff.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Maximum)

        # 追加位置は従来どおり
        left.addWidget(gb_diff)
        # ===== [/PATCH] ======================================================

        left.addStretch(1)
        left.addWidget(self.btn_run)
        left.addWidget(self.progress)
        left.addWidget(self.lbl_stage)
        left.addWidget(self.lbl_elapsed)

        # ========== 右パネル ==========
        right_panel = QWidget()
        right = QVBoxLayout(right_panel)
        right.setContentsMargins(6, 6, 6, 6)
        right.setSpacing(6)

        # --- フィルタ群（統合） ---
        gb_filters = QGroupBox("フィルタ")
        v_filters = QVBoxLayout(gb_filters)
        v_filters.setContentsMargins(8, 8, 8, 8)
        v_filters.setSpacing(4)

        # 1) 全文 ＋ 短文フィルタ
        row_full = QHBoxLayout()
        row_full.setSpacing(8)
        row_full.addWidget(QLabel("全文:"))
        self.ed_filter = QLineEdit()
        self.ed_filter.setPlaceholderText("絞り込みたい語や表現を入力")
        row_full.addWidget(self.ed_filter, 1)

        self.chk_shortlen = QCheckBox("字数")
        self.sb_shortlen = QSpinBox()
        self.sb_shortlen.setRange(1, 120)
        self.sb_shortlen.setValue(3)
        lbl_short_suffix = QLabel("以下を隠す")
        self.chk_shortlen.setToolTip("a 列の文字数が指定値以下の行を非表示にします。")
        self.sb_shortlen.setToolTip("非表示にする最大文字数（a 列の長さ、1〜120）")

        row_full.addSpacing(6)
        row_full.addWidget(self.chk_shortlen)
        row_full.addWidget(self.sb_shortlen)
        row_full.addWidget(lbl_short_suffix)
        v_filters.addLayout(row_full)

        sep1 = QFrame(); sep1.setFrameShape(QFrame.HLine); sep1.setFrameShadow(QFrame.Sunken)
        v_filters.addWidget(sep1)

        # 2) 対象（語彙/複合語/文節）
        row_scope = QHBoxLayout(); row_scope.setSpacing(8)
        row_scope.addWidget(QLabel("対　　象:"))
        self.scope_labels = ["名詞", "活用語", "複合語", "文節", "文章"]
        self.chk_scopes: Dict[str, QCheckBox] = {}
        for s in self.scope_labels:
            cb = QCheckBox(s); cb.setChecked(True); self.chk_scopes[s] = cb
            row_scope.addWidget(cb)
        row_scope.addStretch(1)
        v_filters.addLayout(row_scope)

        sep2 = QFrame(); sep2.setFrameShape(QFrame.HLine); sep2.setFrameShadow(QFrame.Sunken)
        v_filters.addWidget(sep2)

        # 3) 一致要因
        row_reason = QHBoxLayout(); row_reason.setSpacing(8)
        row_reason.addWidget(QLabel("一致要因:"))
        self.reason_labels = [
            "基本一致",
            "読み一致（英数記号除く）",
            "読み一致（表記違い）",
            "読み類似",
            "文字類似",
        ]
        self.chk_reasons: Dict[str, QCheckBox] = {}
        for r in self.reason_labels:
            cb = QCheckBox(r); cb.setChecked(True); self.chk_reasons[r] = cb
            row_reason.addWidget(cb)
        row_reason.addStretch(1)
        v_filters.addLayout(row_reason)

        sep3 = QFrame(); sep3.setFrameShape(QFrame.HLine); sep3.setFrameShadow(QFrame.Sunken)
        v_filters.addWidget(sep3)

        # 4) 端差/数字以外一致
        row_edge = QHBoxLayout(); row_edge.setSpacing(8)
        row_edge.addWidget(QLabel("特殊絞込:"))
        self.chk_edge_prefix = QCheckBox("前1文字差を隠す")
        self.chk_edge_suffix = QCheckBox("後1文字差を隠す")
        self.chk_num_only   = QCheckBox("数字以外一致を隠す")
        
        self.chk_contains    = QCheckBox("内包関係除外")
        self.chk_contains.setChecked(False)
        self.chk_contains.setToolTip("a/b のどちらかがもう一方を含む行を非表示にします。")

        self.chk_edge_prefix.setChecked(False)
        self.chk_edge_suffix.setChecked(False)
        self.chk_num_only.setChecked(False)
        self.chk_edge_prefix.setToolTip("「前1字有無」「前1字違い」の候補を表から除外します。")
        self.chk_edge_suffix.setToolTip("「後1字有無」「後1字違い」の候補を表から除外します。")
        self.chk_num_only.setToolTip("数字だけが異なる候補（例: 1時間30分 ↔ 1時間5分）を表から除外します。")
        row_edge.addWidget(self.chk_edge_prefix)
        row_edge.addWidget(self.chk_edge_suffix)
        row_edge.addWidget(self.chk_num_only)
        row_edge.addWidget(self.chk_contains)
        row_edge.addStretch(1)
        v_filters.addLayout(row_edge)

        right.addWidget(gb_filters)

        # --- タブ＆テーブル ---
        self.tabs = QTabWidget()

        self.view_unified = self._make_table()
        # 行高固定・折り返しなし・省略表記は右側（ElideRight）
        vh = self.view_unified.verticalHeader()
        vh.setSectionResizeMode(QHeaderView.Fixed)
        vh.setDefaultSectionSize(28)
        self.view_unified.setWordWrap(False)
        self.view_unified.setTextElideMode(Qt.ElideRight)  # ← ElideNoneは使わない

        # 右クリックメニュー（差分表示）
        self.view_unified.setContextMenuPolicy(Qt.CustomContextMenu)
        self.view_unified.customContextMenuRequested.connect(self.on_unified_context_menu)

        self.view_tokens = self._make_table_simple()

        self.tabs.addTab(self.view_unified, "表記ゆれ候補")
        self.tabs.addTab(self.view_tokens, "候補（トークン）")
        right.addWidget(self.tabs, 1)

        # --- 下部操作 ---
        ops = QHBoxLayout()
        self.btn_select_all = QPushButton("すべて選択")
        self.btn_clear_sel  = QPushButton("すべて解除")
        self.btn_export     = QPushButton("Excelエクスポート")
        self.btn_mark       = QPushButton("PDFにマーキング")
        # リンク風ヘルプボタン
        help_url = (
            "https://itpcojp-my.sharepoint.com/:b:/g/personal/masahiro_tanaka_itp_co_jp/"
            "Ec84iGbfkvRFl4ZHhK-XR9YBcjDRoWvi6cf3XX59l2sBzg?e=qvFOZP"
        )
        self.btn_help = QPushButton("ヘルプ")
        self.btn_help.setFlat(True)
        self.btn_help.setCursor(Qt.PointingHandCursor)
        self.btn_help.setStyleSheet("""
            QPushButton { color:#1c7ed6; background:transparent; border:none; padding:0 4px; }
            QPushButton:hover { text-decoration: underline; }
        """)
        self.btn_help.clicked.connect(lambda: QDesktopServices.openUrl(QUrl(help_url)))

        ops.addWidget(self.btn_select_all)
        ops.addWidget(self.btn_clear_sel)
        ops.addStretch(1)
        ops.addWidget(self.btn_help)
        ops.addSpacing(8)
        ops.addWidget(self.btn_export)
        ops.addWidget(self.btn_mark)
        right.addLayout(ops)

        # ルートへ追加
        root.addWidget(left_panel)
        root.addWidget(right_panel)
        root.setStretch(0, 0)
        root.setStretch(1, 1)

        # ---- Model / Proxy ----
        self.model_unified = UnifiedModel()
        self.model_tokens  = PandasModel()
        self.proxy_unified = UnifiedFilterProxy()
        self.proxy_unified.setSourceModel(self.model_unified)
        self.proxy_tokens = QSortFilterProxyModel()
        self.proxy_tokens.setSourceModel(self.model_tokens)
        self.proxy_tokens.setFilterCaseSensitivity(Qt.CaseInsensitive)
        self.proxy_tokens.setFilterKeyColumn(1)  # tokens: 0:type, 1:token, 2:count

        self.view_unified.setModel(self.proxy_unified)
        # ===== [PATCH] MainWindow.__init__ の「モデル/ビュー接続」直後に追記 =====
        self.view_unified.setModel(self.proxy_unified)
        # チェック列（0列）の描画をセンター揃えデリゲートに
        self.view_unified.setItemDelegateForColumn(0, CheckBoxDelegate(self.view_unified))
        # ヘッダは中央寄せ（✓ ヘッダも中央に）
        self.view_unified.horizontalHeader().setDefaultAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
        # ===== [/PATCH] =============================================================
        # ===== [PATCH] 選択変更で差分プレビュー更新 =====
        self.view_unified.selectionModel().selectionChanged.connect(self.on_unified_selection_changed)
        # ===== [/PATCH] =================================
        self.view_tokens.setModel(self.proxy_tokens)

        # ---- Signals ----
        self.drop.filesChanged.connect(self.on_files_changed)
        self.btn_browse.clicked.connect(self.on_browse)
        self.btn_clear.clicked.connect(self.on_clear_files)
        self.btn_run.clicked.connect(self.on_run)

        self.ed_filter.textChanged.connect(self.on_text_filter_changed)
        self.chk_shortlen.stateChanged.connect(self.on_shortlen_filter_changed)
        self.sb_shortlen.valueChanged.connect(self.on_shortlen_filter_changed)
        for cb in self.chk_reasons.values():
            cb.stateChanged.connect(self.on_reason_changed)
        for cb in self.chk_scopes.values():
            cb.stateChanged.connect(self.on_scope_changed)
        self.chk_edge_prefix.stateChanged.connect(self.on_edge_filter_changed)
        self.chk_edge_suffix.stateChanged.connect(self.on_edge_filter_changed)
        self.chk_num_only.stateChanged.connect(self.on_edge_filter_changed)
        self.chk_contains.stateChanged.connect(self.on_edge_filter_changed)

        self.btn_select_all.clicked.connect(lambda: self.model_unified.check_all(True))
        self.btn_clear_sel.clicked.connect(lambda: self.model_unified.check_all(False))
        self.view_unified.clicked.connect(self.on_unified_clicked)
        self.btn_mark.clicked.connect(self.on_mark_pdf)
        self.btn_export.clicked.connect(self.on_export)

        # ---- 状態系 ----
        self.files: List[str] = []
        self._t0: Optional[float] = None
        self._timer = QTimer(self)
        self._timer.setInterval(250)
        self._timer.timeout.connect(self._update_elapsed)

        self.df_groups = pd.DataFrame()
        self.surf2gid: Dict[str, int] = {}
        self.gid2members: Dict[int, List[str]] = {}

        # 初期フィルタ状態（重複初期化を排除）
        self._init_filter_state_once()
        self._refresh_candidate_count()

    # =========================================================
    # 小ユーティリティ
    # =========================================================
    # ===== [REPLACE] MainWindow._make_table（Stretch無効 + ヘッダ中央 + スクロール） =====
    def _make_table(self) -> QTableView:
        tv = QTableView()
        tv.setSortingEnabled(True)
        tv.setAlternatingRowColors(True)
        tv.setSelectionBehavior(QAbstractItemView.SelectRows)
        tv.setSelectionMode(QAbstractItemView.SingleSelection)
        tv.verticalHeader().setVisible(False)
        tv.verticalHeader().setDefaultSectionSize(28)
        # ★ ここを False に（最終列が余白で勝手に広がらない）
        tv.horizontalHeader().setStretchLastSection(False)
        tv.horizontalHeader().setHighlightSections(False)
        # ★ ヘッダの文字揃えを中央へ
        tv.horizontalHeader().setDefaultAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
        # スクロールの手触り
        tv.setHorizontalScrollMode(QAbstractItemView.ScrollPerPixel)
        tv.setVerticalScrollMode(QAbstractItemView.ScrollPerPixel)
        # テキストは折り返さず、長い時は右側省略
        tv.setWordWrap(False)
        tv.setTextElideMode(Qt.ElideRight)
        return tv
    # ===== [/REPLACE] ===========================================================

    def _make_table_simple(self) -> QTableView:
        return self._make_table()

    def _init_filter_state_once(self):
        # Proxy 初期状態の一括反映
        self.proxy_unified.setScopeFilter(self._selected_scopes())
        self.proxy_unified.setReasonFilter(self._selected_reasons())
        self.proxy_unified.setHidePrefixEdge(self.chk_edge_prefix.isChecked())
        self.proxy_unified.setHideSuffixEdge(self.chk_edge_suffix.isChecked())
        self.proxy_unified.setHideNumericOnly(self.chk_num_only.isChecked())
        self.proxy_unified.setHideShortEnabled(self.chk_shortlen.isChecked())
        self.proxy_unified.setHideShortLength(self.sb_shortlen.value())
        self.proxy_unified.setHideContainment(self.chk_contains.isChecked())

    def _selected_scopes(self) -> Set[str]:
        return {k for k, cb in self.chk_scopes.items() if cb.isChecked()}

    def _selected_reasons(self) -> Set[str]:
        # GUI表示ラベル → apply_reason_ja() 後の「一致要因」列の値と一致
        return {k for k, cb in self.chk_reasons.items() if cb.isChecked()}

    def _update_elapsed(self):
        if self._t0 is None:
            self.lbl_elapsed.setText("経過 00:00")
            return
        sec = max(0, int(time.monotonic() - self._t0))
        h, rem = divmod(sec, 3600)
        m, s = divmod(rem, 60)
        fmt = f"{h:02d}:{m:02d}:{s:02d}" if h else f"{m:02d}:{s:02d}"
        self.lbl_elapsed.setText(f"経過 {fmt}")

    def _update_elapsed(self):
        if self._t0 is None:
            self.lbl_elapsed.setText("経過 00:00")
            return
        sec = max(0, int(time.monotonic() - self._t0))
        h, rem = divmod(sec, 3600)
        m, s = divmod(rem, 60)
        fmt = f"{h:02d}:{m:02d}:{s:02d}" if h else f"{m:02d}:{s:02d}"
        self.lbl_elapsed.setText(f"経過 {fmt}")

    def _smooth_progress_update(self):
        """進捗バーを滑らかに更新(目標値へ徐々に近づける)"""
        if abs(self._target_progress - self._current_smooth_progress) < 0.5:
            # ほぼ到達したら即座に合わせる
            self._current_smooth_progress = float(self._target_progress)
            self.progress.setValue(int(self._current_smooth_progress))
            return
        
        # 指数関数的に目標値へ近づける(イージング効果)
        diff = self._target_progress - self._current_smooth_progress
        self._current_smooth_progress += diff * self._progress_smooth_speed
        
        self.progress.setValue(int(round(self._current_smooth_progress)))

    def _on_worker_progress(self, value: int):
        """ワーカーからの進捗を受け取り、目標値として設定"""
        self._target_progress = max(0, min(100, int(value)))

    def _refresh_candidate_count(self):
        try:
            df_u = getattr(self.model_unified, "_df", pd.DataFrame())
            total = int(df_u.shape[0]) if isinstance(df_u, pd.DataFrame) else 0
            visible = int(self.proxy_unified.rowCount())
            self.tabs.setTabText(0, f"表記ゆれ候補[{visible:,}/{total:,}]")
        except Exception:
            self.tabs.setTabText(0, "表記ゆれ候補[0/0]")

    def _cap_ab_widths(self, min_a=220, max_a=320, min_b=220, max_b=320):
        # a/b 列の幅を上限クリップ
        df = getattr(self.model_unified, "_df", pd.DataFrame())
        if df.empty:
            return
        cols = list(df.columns)
        try:
            idx_a = cols.index("a")
            idx_b = cols.index("b")
        except ValueError:
            return
        va = 1 + idx_a  # 先頭のチェック列があるので +1
        vb = 1 + idx_b
        hdr = self.view_unified.horizontalHeader()
        hdr.setSectionResizeMode(va, QHeaderView.Interactive)
        hdr.setSectionResizeMode(vb, QHeaderView.Interactive)
        w = self.view_unified.columnWidth
        self.view_unified.setColumnWidth(va, min(max_a, max(min_a, w(va))))
        self.view_unified.setColumnWidth(vb, min(max_b, max(min_b, w(vb))))

    def _hide_aux_columns(self):
        # 表示を軽くするため補助列を隠す（存在チェック付き）
        view = self.view_unified
        model = view.model()  # proxy
        if not model:
            return
        names_hide = {"端差", "数字以外一致", "__contains__"}  # 必要に応じて追加
        hdr = view.horizontalHeader()
        for c in range(model.columnCount()):
            name = (model.headerData(c, Qt.Horizontal, Qt.DisplayRole) or "").strip()
            if name in names_hide:
                hdr.setSectionResizeMode(c, QHeaderView.Fixed)
                view.setColumnHidden(c, True)

    # ===== [REPLACE] MainWindow._compact_columns（gid基準の同幅＋一致要因だけStretch） =====
    def _compact_columns(self):
        view = self.view_unified
        model = view.model()
        if not model:
            return

        hdr = view.horizontalHeader()
        hdr.setMinimumSectionSize(40)

        # 表示ヘッダ名から列番号を探すヘルパ
        def col_by_name(*names: str) -> int:
            for c in range(model.columnCount()):
                name = (model.headerData(c, Qt.Horizontal, Qt.DisplayRole) or "").strip()
                if name in names:
                    return c
            return -1

        # 2) 「a数」「b数」「類似度」「対象」を gid と同幅の Fixed にする
        for label in ("gid", "字数", "a数", "b数", "類似度", "対象"):
            c = col_by_name(label)
            if c >= 0:
                hdr.setSectionResizeMode(c, QHeaderView.Fixed)
                view.setColumnWidth(c, 72)

        # 3) 「一致要因」だけは Stretch（ウィンドウに合わせて伸縮）
        c_reason = col_by_name("一致要因")
        if c_reason >= 0:
            hdr.setSectionResizeMode(c_reason, QHeaderView.Stretch)
            # ★ヘッダの“全列共通”最小幅は設定しない！
            # 初期表示が狭すぎる場合だけ、起動直後に少し広げる
            view.setColumnWidth(c_reason, max(view.columnWidth(c_reason), 200))


        # 4) それ以外の列
        #    - a/b はユーザー操作しやすい Interactive
        #    - 残りは Fixed + 上限幅キャップでコンパクトに
        NUMERIC = {"字数"}  # （a数/b数/類似度 はすでに gid 同幅にしているので除外）
        LABELS  = {"理由", "scope", "target"}  # 「一致要因」「対象」は既に処理済み

        for c in range(model.columnCount()):
            name = (model.headerData(c, Qt.Horizontal, Qt.DisplayRole) or "").strip()

            # 既に個別ルールを与えた列はスキップ
            if name in {"gid", "a数", "b数", "類似度", "対象", "一致要因"}:
                continue

            if name in {"a", "b"}:
                hdr.setSectionResizeMode(c, QHeaderView.Interactive)
                continue

            # その他は Fixed + 上限で抑える
            if name in NUMERIC:
                max_w = 72
            elif name in LABELS:
                max_w = 120
            else:
                max_w = 140

            cur = view.columnWidth(c)
            view.setColumnWidth(c, min(max_w, cur if cur > 0 else max_w))
            hdr.setSectionResizeMode(c, QHeaderView.Fixed)
    # ===== [/REPLACE] =========================================================

    # =========================================================
    # DnD & 入出力
    # =========================================================
    def on_files_changed(self, files: List[str]):
        self.set_files(files)

    def set_files(self, files: List[str]):
        self.files = files
        self.list_files.clear()
        for f in files:
            self.list_files.addItem(f)
        self.lbl_count.setText(f"選択ファイル：{len(files)} 件")

    def on_browse(self):
        paths, _ = QFileDialog.getOpenFileNames(self, "PDFを選択", "", "PDF Files (*.pdf)")
        if paths:
            self.set_files(paths)

    def on_clear_files(self):
        self.set_files([])
        self.drop.clear()

        # 🆕 進捗バーの滑らか化用タイマー
        self._progress_smooth_timer = QTimer(self)
        self._progress_smooth_timer.setInterval(50)  # 50ms = 20fps で滑らか
        self._progress_smooth_timer.timeout.connect(self._smooth_progress_update)
        
        self._target_progress = 0  # 目標進捗値
        self._current_smooth_progress = 0  # 現在の滑らか化された進捗
        self._progress_smooth_speed = 2.0  # 追従速度(値が大きいほど速い)

    def _smooth_progress_update(self):
        """進捗バーを滑らかに更新(目標値へ徐々に近づける)"""
        if abs(self._target_progress - self._current_smooth_progress) < 0.5:
            # ほぼ到達したら即座に合わせる
            self._current_smooth_progress = self._target_progress
            self.progress.setValue(int(self._current_smooth_progress))
            return
        
        # 指数関数的に目標値へ近づける
        diff = self._target_progress - self._current_smooth_progress
        self._current_smooth_progress += diff * self._progress_smooth_speed * 0.05
        
        self.progress.setValue(int(self._current_smooth_progress))

    def on_run(self):
        if not self.files:
            QMessageBox.warning(self, "注意", "PDFを指定してください。")
            return
        if not HAS_MECAB:
            QMessageBox.critical(
                self, "エラー",
                'MeCab (fugashi/unidic-lite) が必要です。\n'
                'インストール例: pip install "fugashi[unidic-lite]"'
            )
            return

        self.btn_run.setEnabled(False)
        self.progress.setValue(0)
        
        # 🆕 滑らか化の初期化
        self._target_progress = 0
        self._current_smooth_progress = 0.0
        self._progress_smooth_timer.start()
        
        self._t0 = time.monotonic()
        self._update_elapsed()
        self._timer.start()
        self.lbl_stage.setText("開始")

        self.thread = QThread()
        self.worker = AnalyzerWorker(self.files, True, self.dsb_read.value(), self.dsb_char.value())
        self.worker.moveToThread(self.thread)
        self.thread.started.connect(self.worker.run)
        
        # 🆕 進捗シグナルを滑らか化機構に接続
        self.worker.progress.connect(self._on_worker_progress)
        
        self.worker.progress_text.connect(self.lbl_stage.setText)
        self.worker.finished.connect(self.on_finished)
        self.worker.finished.connect(lambda *_: self.thread.quit())
        self.thread.finished.connect(lambda: self.btn_run.setEnabled(True))
        self.thread.finished.connect(self.thread.deleteLater)
        
        # 🆕 完了時に滑らか化タイマーを停止
        self.thread.finished.connect(self._progress_smooth_timer.stop)
        
        self.thread.start()

    def _on_worker_progress(self, value: int):
        """ワーカーからの進捗を受け取り、目標値として設定"""
        self._target_progress = max(0, min(100, int(value)))


    def on_finished(self, results: dict, error: str):
        # 🆕 完了時は即座に100%へ
        self._target_progress = 100
        self._current_smooth_progress = 100.0
        self.progress.setValue(100)
        self._progress_smooth_timer.stop()
        
        if error:
            QMessageBox.critical(self, "エラー", error)
            self.lbl_stage.setText("エラー")
            if self._timer.isActive(): self._timer.stop()
            return

        self.lbl_stage.setText("集計完了（表示中…）")

        self._target_progress = 100
        self._current_smooth_progress = 100
        self.progress.setValue(100)
        self._progress_smooth_timer.stop()

        df_unified = results.get("unified", pd.DataFrame())
        self.df_groups = results.get("groups", pd.DataFrame())
        self.surf2gid = results.get("surf2gid", {}) or {}
        self.gid2members = results.get("gid2members", {}) or {}

        if not df_unified.empty:
            df_u = df_unified.copy()
            cols = list(df_u.columns)
            pref = [c for c in ["gid", "字数", "a", "b"] if c in cols]
            rest = [c for c in cols if c not in pref]
            df_u = df_u[pref + rest]
        else:
            df_u = df_unified

        self.model_unified.setDataFrame(df_u)
        self.model_tokens.setDataFrame(results.get("tokens", pd.DataFrame()))

        # テーブル調整（再描画負荷を抑えつつ）
        self.view_unified.setUpdatesEnabled(False)
        self.view_unified.setSortingEnabled(False)
        self._cap_ab_widths(min_a=220, max_a=320, min_b=220, max_b=320)
        self._hide_aux_columns()
        self._ensure_ab_visible(min_a=220, min_b=220)
        self._compact_columns()
        self.view_unified.setSortingEnabled(True)
        self.view_unified.setUpdatesEnabled(True)

        self._refresh_candidate_count()
        QApplication.processEvents()
        self.progress.setValue(100)
        self.lbl_stage.setText("完了")
        if self._timer.isActive():
            self._timer.stop()

    def _ensure_ab_visible(self, min_a=220, min_b=220):
        df = getattr(self.model_unified, "_df", pd.DataFrame())
        if df.empty:
            return
        cols = list(df.columns)
        try:
            idx_a = cols.index("a")
            idx_b = cols.index("b")
        except ValueError:
            return
        va = 1 + idx_a  # 先頭チェック列があるため +1
        vb = 1 + idx_b
        self.view_unified.setColumnWidth(va, max(self.view_unified.columnWidth(va), min_a))
        self.view_unified.setColumnWidth(vb, max(self.view_unified.columnWidth(vb), min_b))

    # =========================================================
    # テーブル操作・メニュー
    # =========================================================
    def on_unified_clicked(self, proxy_index: QModelIndex):
        if not proxy_index.isValid() or proxy_index.column() != 0:
            return
        src_index = self.proxy_unified.mapToSource(proxy_index)
        current = self.model_unified.data(src_index, Qt.CheckStateRole)
        new_state = Qt.Unchecked if current == Qt.Checked else Qt.Checked
        self.model_unified.setData(src_index, new_state, Qt.CheckStateRole)

    def on_unified_context_menu(self, pos: QPoint):
        index = self.view_unified.indexAt(pos)
        if not index.isValid():
            return
        proxy_row = index.row()
        src_index = self.proxy_unified.mapToSource(self.proxy_unified.index(proxy_row, 1))
        row = src_index.row()
        df = getattr(self.model_unified, "_df", pd.DataFrame())
        if df.empty or row < 0 or row >= len(df):
            return
        a = df.at[row, "a"] if "a" in df.columns else ""
        b = df.at[row, "b"] if "b" in df.columns else ""

        menu = QMenu(self)
        act_diff = QAction("差分を表示…", self)

        def _show():
            dlg = DiffDialog(str(a) if a is not None else "", str(b) if b is not None else "", self)
            dlg.exec()

        act_diff.triggered.connect(_show)
        menu.addAction(act_diff)
        menu.exec(self.view_unified.viewport().mapToGlobal(pos))

    # =========================================================
    # フィルタ変更ハンドラ
    # =========================================================
    def on_text_filter_changed(self, text: str):
        self.proxy_unified.setTextFilter(text)
        self.proxy_tokens.setFilterFixedString(text)
        self._refresh_candidate_count()

    def on_shortlen_filter_changed(self, *_):
        self.proxy_unified.setHideShortEnabled(self.chk_shortlen.isChecked())
        self.proxy_unified.setHideShortLength(self.sb_shortlen.value())
        self._refresh_candidate_count()

    def on_reason_changed(self, *_):
        self.proxy_unified.setReasonFilter(self._selected_reasons())
        self._refresh_candidate_count()

    def on_scope_changed(self, *_):
        self.proxy_unified.setScopeFilter(self._selected_scopes())
        self._refresh_candidate_count()

    def on_edge_filter_changed(self, *_):
        self.proxy_unified.setHidePrefixEdge(self.chk_edge_prefix.isChecked())
        self.proxy_unified.setHideSuffixEdge(self.chk_edge_suffix.isChecked())
        self.proxy_unified.setHideNumericOnly(self.chk_num_only.isChecked())
        self.proxy_unified.setHideContainment(self.chk_contains.isChecked())
        self._refresh_candidate_count()

    # =========================================================
    # エクスポート / PDFマーキング
    # =========================================================
    def _build_unified_df_with_selection_all(self) -> pd.DataFrame:
        """GUIの選択状態を含めて、全件の DataFrame を返す（Excel用）"""
        src_df = getattr(self.model_unified, "_df", pd.DataFrame()).copy()
        checks = list(getattr(self.model_unified, "_checks", []))
        if src_df.empty:
            return src_df
        if len(checks) != len(src_df):
            checks = [False] * len(src_df)
        df_all = src_df.reset_index(drop=True)
        df_all.insert(0, "選択", checks)
        if "字数" not in df_all.columns and "a" in df_all.columns:
            df_all["字数"] = df_all["a"].map(lambda x: len(str(x)) if x is not None else 0).astype(int)
        return df_all

    # ===== [REPLACE] Excelエクスポート（ヘッダ候補でのフィルタ解決 版） =====
    def on_export(self):
        """Excel（全件 + GUIの初期フィルタ状態）を出力"""
        import math as _math
        import json
        import pandas as pd
        from PySide6.QtWidgets import QMessageBox, QFileDialog
        from openpyxl.utils import get_column_letter, column_index_from_string
        from openpyxl.worksheet.filters import CustomFilters, CustomFilter, FilterColumn

        # 1) 全件データの取得
        try:
            df_all = self._build_unified_df_with_selection_all()
        except Exception:
            src_df = getattr(self.model_unified, "_df", pd.DataFrame()).copy()
            checks = list(getattr(self.model_unified, "_checks", []))
            if src_df.empty:
                QMessageBox.information(self, "情報", "エクスポートするデータがありません。")
                return
            if len(checks) != len(src_df):
                checks = [False] * len(src_df)
            df_all = src_df.reset_index(drop=True)
            df_all.insert(0, "選択", checks)
            if "字数" not in df_all.columns and "a" in df_all.columns:
                df_all["字数"] = df_all["a"].map(lambda x: len(str(x)) if x is not None else 0).astype(int)

        if df_all.empty:
            QMessageBox.information(self, "情報", "エクスポートするデータがありません。")
            return

        # 2) 保存先
        path, _ = QFileDialog.getSaveFileName(
            self, "Excelの保存先", "variants_unified.xlsx", "Excel ファイル (*.xlsx)"
        )
        if not path:
            return

        # 3) GUI フィルタ状態
        scopes_on = sorted(list(self._selected_scopes()))
        reasons_on = sorted(list(self._selected_reasons()))
        hide_prefix = self.chk_edge_prefix.isChecked()
        hide_suffix = self.chk_edge_suffix.isChecked()
        hide_numonly = self.chk_num_only.isChecked()
        hide_contains = self.chk_contains.isChecked()

        short_on = self.chk_shortlen.isChecked()
        short_n = int(self.sb_shortlen.value()) if short_on else None

        fulltext = (self.ed_filter.text() or "").strip()
        need_full = bool(fulltext)

        # 4) 書き出し用の補助列
        if "字数" not in df_all.columns and "a" in df_all.columns:
            df_all["字数"] = df_all["a"].astype("string").fillna("").str.len().astype(int)

        if hide_contains and {"a", "b"}.issubset(df_all.columns):
            def _contains_row(a, b):
                sa = nfkc(str(a or "")).lower()
                sb = nfkc(str(b or "")).lower()
                return bool(sa and sb and sa != sb and (sa.find(sb) != -1 or sb.find(sa) != -1))
            df_all["__contains__"] = [_contains_row(a, b) for a, b in zip(df_all["a"], df_all["b"])]

        if need_full:
            cols_for_full = [c for c in ["a", "b", "一致要因", "対象", "scope", "target"] if c in df_all.columns]
            if cols_for_full:
                df_all["全文__concat"] = df_all[cols_for_full].astype(str).fillna("").agg(" / ".join, axis=1)
            else:
                need_full = False

        # 5) pandas → openpyxl
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            sheet_name = "表記ゆれ候補"

            # --- [ADD] Excel 禁止文字のクリーンアップ（to_excel の直前） ---
            from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
            import pandas as pd
            import numpy as np

            def _clean_illegal_excel_chars(df: pd.DataFrame) -> pd.DataFrame:
                # 文字列系の列だけ対象（数値や日付の型を壊さない）
                str_like = df.select_dtypes(include=["object", "string"]).columns
                if len(str_like) == 0:
                    return df
                def _clean_one(v):
                    if isinstance(v, str):
                        # 制御文字を除去
                        return ILLEGAL_CHARACTERS_RE.sub("", v)
                    return v
                # 列ごとに map（型崩れ最小化）
                for c in str_like:
                    df[c] = df[c].map(_clean_one)
                return df

            df_all = _clean_illegal_excel_chars(df_all)
            # --- [/ADD] ---

            df_all.to_excel(writer, sheet_name=sheet_name, index=False)
            wb = writer.book
            ws = writer.sheets[sheet_name]

            # ===== [ADD] Excelヘッダだけ上書き（出力専用の表示名差し替え） =====
            # 既存の1行目ヘッダを読み出し
            headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]

            # 出力専用の表示名マップ（左=元の列名 / 右=Excelで見せたい見出し）
            # ※ DataFrame 側の列名は変更しません。Excel表示だけ変えます。
            header_map_excel_only = {
                "scope": "対象",          # 英名→日本語
                "target": "対象",         # 念のため
                "__contains__": "内包関係", # 内部列→日本語
                "a_count": "a数",    # 実テーブルの表示名に合わせてここを変えてください
                "b_count": "b数",
                "score":   "類似度",

                # 必要ならここに追加できます:
                # "a": "A", "b": "B", など
            }

            # 既存の日本語名をそのまま使うための保護
            # 例: すでに「対象」見出しがあるのに "target" を「対象」にすると重複する場合があるため、
            #     その場合は「対象(2)」のように重複回避します（Excelは重複でも動きますが念のため）。
            def _unique_header_name(desired: str, used: set[str]) -> str:
                if desired not in used:
                    return desired
                k = 2
                while f"{desired}({k})" in used:
                    k += 1
                return f"{desired}({k})"

            used = set(h for h in headers if h)
            for col_idx, old in enumerate(headers, start=1):
                if old in header_map_excel_only:
                    new_name = _unique_header_name(header_map_excel_only[old], used)
                    ws.cell(row=1, column=col_idx, value=new_name)
                    used.add(new_name)

            # 以降のフィルタ列解決は「正規化名＋旧名の候補」で行えるように、
            # 先にヘッダを取り直しておくと安心
            headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
            # ===== [/ADD] =====


            # ヘッダ固定・AutoFilter 範囲設定
            ws.freeze_panes = "A2"  # 1行目固定（見やすさ）[3](https://dnmtechs.com/freezing-header-row-in-openpyxl/)
            ws.auto_filter.ref = ws.dimensions  # シート全体にフィルタ[4](https://stackoverflow.com/questions/51566349/openpyxl-how-to-add-filters-to-all-columns)

            # === 列解決ヘルパ（複数候補を許容） ===
            def __find_col(candidates):
                """
                candidates: ヘッダ候補（タプル/リスト/文字列）を上から順に探す
                戻り値: 1起点の列番号 or None
                """
                if not isinstance(candidates, (tuple, list)):
                    candidates = (candidates,)
                # 1行目を走査
                header_texts = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[1]]
                # 完全一致優先
                for cand in candidates:
                    for idx, name in enumerate(header_texts, start=1):
                        if name == cand:
                            return idx
                # 大文字小文字/空白差などを吸収するゆるめ一致
                norm = lambda s: nfkc(str(s or "")).strip().lower()
                cand_norm = [norm(c) for c in candidates]
                for idx, name in enumerate(header_texts, start=1):
                    if norm(name) in cand_norm:
                        return idx
                return None

            # 0起点 col_id を得るために左端列を取得
            left_ref = ws.auto_filter.ref.split(":")[0]
            start_col_idx = column_index_from_string("".join(ch for ch in left_ref if ch.isalpha()))

            def _add_value_filter(header_names, values, allow_blank=False):
                """値リストによる表示フィルタ（header_names は候補群に対応）"""
                values = [v for v in (values or []) if (v is not None and str(v) != "")]
                if not values and not allow_blank:
                    return
                col = __find_col(header_names)
                if not col:
                    return
                col_id = (col - start_col_idx)  # 0起点
                ws.auto_filter.add_filter_column(
                    col_id,
                    list(dict.fromkeys(map(str, values))),
                    blank=bool(allow_blank)
                )
                # add_filter_column の使い方（値リスト/ブランク許容）[1](https://openpyxl.pages.heptapod.net/openpyxl/api/openpyxl.worksheet.filters.html)

            def _add_number_threshold_filter(header_names, operator: str, val: int):
                """数値カスタムフィルタ（>=, <= など）"""
                col = __find_col(header_names)
                if not col:
                    return
                col_id = (col - start_col_idx)
                cf = CustomFilters()
                cf.customFilter = [CustomFilter(operator=operator, val=str(val))]
                ws.auto_filter.filterColumn.append(FilterColumn(colId=col_id, customFilters=cf))
                # カスタムフィルタの低レベルAPI[1](https://openpyxl.pages.heptapod.net/openpyxl/api/openpyxl.worksheet.filters.html)

            # (a) 対象（= scope 表示列）
            #     ← ここが修正ポイント：("対象","scope","target") の順で探索
            if scopes_on:
                _add_value_filter(("対象", "scope", "target"), scopes_on, allow_blank=False)

            # (b) 一致要因
            if reasons_on and ("一致要因" in df_all.columns):
                _add_value_filter(("一致要因",), reasons_on, allow_blank=False)

            # (c) 端差（前/後1字を隠す → 許容値のみ残す）
            if "端差" in df_all.columns and (hide_prefix or hide_suffix):
                series = df_all["端差"].astype(str)
                uniq = set(s.strip() for s in series.fillna(""))
                deny = set()
                if hide_prefix: deny |= {"前1字有無", "前1字違い"}
                if hide_suffix: deny |= {"後1字有無", "後1字違い"}
                allow_blank = ("" in uniq)
                allowed = [v for v in sorted(uniq) if v and v not in deny]
                _add_value_filter(("端差",), allowed, allow_blank=allow_blank)

            # (d) 数字以外一致（隠す）
            if "数字以外一致" in df_all.columns and hide_numonly:
                series = df_all["数字以外一致"].astype(str)
                uniq = set(s.strip() for s in series.fillna(""))
                allow_blank = ("" in uniq)
                allowed = [v for v in sorted(uniq) if v and v != "数字以外一致"]
                _add_value_filter(("数字以外一致",), allowed, allow_blank=allow_blank)

            # (e) 短文（字数 <= N を隠す → 字数 >= N+1 を残す）
            if short_on and isinstance(short_n, int) and short_n > 0:
                _add_number_threshold_filter(("字数",), operator="greaterThanOrEqual", val=int(short_n) + 1)

            # (f) 内包関係除外（True を除外 → False のみ許容）
            # hide_contains が True のとき、False のみ表示（= True を除外）
            if hide_contains and ("__contains__" in df_all.columns or "内包関係" in headers):
                _add_value_filter(("内包関係", "__contains__"), [False], allow_blank=False)

            # (g) 全文（部分一致を代替：該当レコードの値リストでホワイトリスト化）
            if need_full and "全文__concat" in df_all.columns:
                vals = df_all.loc[df_all["全文__concat"].str.contains(fulltext, case=False, na=False), "全文__concat"]
                if len(vals) > 0:
                    _add_value_filter(("全文__concat",), list(pd.unique(vals.astype(str))), allow_blank=False)

            ws.auto_filter.ref = ws.dimensions  # 念のため同期

            # --- A) 非マッチ行を隠す（例：「対象/scope」が scopes_on のみ表示） ---
            from openpyxl.utils import column_index_from_string

            ws = writer.sheets[sheet_name]  # pandas.ExcelWriter から取り出したシート

            # ---- [ADD at 保存直前ブロック] 数値以外一致の行を先に隠す ----
            headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]

            def _find_col(cands):
                if not isinstance(cands, (tuple, list)): cands = (cands,)
                # 完全一致
                for nm in cands:
                    if nm in headers: return headers.index(nm) + 1
                # ゆるめ一致（NFKC を使わず簡易に）
                def _norm(s): return str(s or "").strip().lower()
                H = [_norm(h) for h in headers]
                for nm in cands:
                    nn = _norm(nm)
                    if nn in H: return H.index(nn) + 1
                return None

            def _is_truthy(v):
                s = str(v).strip().lower()
                return (s in {"true","1","t","y","yes","はい","有","○","◯"}
                        or str(v).strip() in {"数値以外一致","数字以外一致"})

            if hide_numonly:
                col_num = _find_col(("数値以外一致","数字以外一致","non_numeric_match"))
                if col_num:
                    for r in range(2, ws.max_row + 1):
                        v = ws.cell(row=r, column=col_num).value
                        if _is_truthy(v):  # ←「数値以外一致」と判定された行は非表示に
                            ws.row_dimensions[r].hidden = True
            # ---- [/ADD] ----


            headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]

            def _find_col(cands):
                if not isinstance(cands, (tuple, list)): cands = (cands,)
                # 完全一致
                for nm in cands:
                    if nm in headers: return headers.index(nm) + 1
                # ゆるめ（NFKC・小文字）
                def norm(s): return nfkc(str(s or "")).strip().lower()
                H = [norm(h) for h in headers]
                for nm in cands:
                    if norm(nm) in H: return H.index(norm(nm)) + 1
                return None

            col_scope = _find_col(("対象", "scope", "target"))

            if col_scope and scopes_on:
                show_set = set(map(str, scopes_on))
                # データ行（2行目〜最終行）を評価
                for r in range(2, ws.max_row + 1):
                    v = ws.cell(row=r, column=col_scope).value
                    if str(v) not in show_set:
                        ws.row_dimensions[r].hidden = True  # ← ここで非表示にする

            # --- AutoFilter は定義として残す（ドロップダウン＆条件表示用） ---
            ws.auto_filter.ref = ws.dimensions
            left_ref = ws.auto_filter.ref.split(":")[0]
            start_col_idx = column_index_from_string("".join(ch for ch in left_ref if ch.isalpha()))
            col_id = (col_scope - start_col_idx) if col_scope else None
            if col_id is not None:
                ws.auto_filter.add_filter_column(col_id, list(show_set), blank=False)

        self.statusBar().showMessage("Excelを出力しました", 5000)
        QMessageBox.information(
            self, "完了",
            "Excelエクスポートが完了しました。"
        )
    # ===== [/REPLACE] =====


    def on_mark_pdf(self):
        """チェック済みの行に行番号(1,2,...)を振り、行内の a/b に 1-a, 1-b ... の注釈を付ける"""
        import os
        import unicodedata
        import pandas as pd
        import fitz  # PyMuPDF

        # 入力チェック
        if not self.files:
            QMessageBox.warning(self, "注意", "PDFを指定してください。")
            return

        df = getattr(self.model_unified, "_df", pd.DataFrame())
        if df.empty:
            QMessageBox.information(self, "情報", "マーキング対象がありません。")
            return

        # カラム位置
        col_a = df.columns.get_indexer(["a"])[0] if "a" in df.columns else -1
        col_b = df.columns.get_indexer(["b"])[0] if "b" in df.columns else -1

        # ===== 1) チェック済み行を a/b のペアとして束ね、行順で通し番号を振る土台を作成 =====
        # rows: [ [("a", a_text?), ("b", b_text?)] , ... ]  ※aまたはbが欠ける場合もあり
        rows = []
        for r_proxy in range(self.proxy_unified.rowCount()):
            src_idx = self.proxy_unified.mapToSource(self.proxy_unified.index(r_proxy, 0))
            i = src_idx.row()
            if i < 0 or i >= len(self.model_unified._checks):
                continue
            if not self.model_unified._checks[i]:
                continue

            pair = []
            if col_a >= 0:
                a = df.iat[i, col_a]
                if isinstance(a, str) and a:
                    pair.append(("a", a))
            if col_b >= 0:
                b = df.iat[i, col_b]
                if isinstance(b, str) and b:
                    pair.append(("b", b))

            if pair:
                # 念のため a→b の順を保証
                pair.sort(key=lambda t: t[0])
                rows.append(pair)

        if not rows:
            QMessageBox.information(self, "情報", "マーキング対象が選択されていません。")
            return

        # ===== 2) 色設定（従来踏襲） =====
        COLOR_YELLOW = (0.98, 0.90, 0.25)  # a
        COLOR_CYAN   = (0.00, 0.90, 1.00)  # b
        HIGHLIGHT_COLORS = {"a": COLOR_YELLOW, "b": COLOR_CYAN}

        # 出力フォルダ
        out_dir = QFileDialog.getExistingDirectory(self, "出力フォルダを選択")
        if not out_dir:
            return

        # ===== 3) 検索ユーティリティ（元実装の堅牢検索をそのまま使用） =====
        CONN_CHARS = "ー\\-－–—・/／\\_＿"
        SPACE_CHARS = {" ", "\u00A0", "\u3000"}

        def nfkc(s: str) -> str:
            return unicodedata.normalize("NFKC", s)

        def hira_to_kata(s: str) -> str:
            out = []
            for ch in s:
                o = ord(ch)
                out.append(chr(o + 0x60) if 0x3041 <= o <= 0x3096 else ch)
            return "".join(out)

        def kana_norm_variants(s: str):
            base = nfkc(s or ""); k = hira_to_kata(base)
            return {k, k.replace("ー", "")}

        def base_variants(s: str):
            vs = set(); s0 = s or ""; s1 = nfkc(s0)
            for cand in (s0, s1):
                if not cand:
                    continue
                vs.add(cand)
                vs.add(cand.replace(" ", ""))
                vs.add(cand.replace(" ", "\u00A0"))
                vs.add(cand.replace(" ", "\u3000"))
                for ch in list(CONN_CHARS):
                    if ch in cand:
                        vs.add(cand.replace(ch, ""))
                        vs.add(cand.replace(ch, " "))
                        vs.add(cand.replace(ch, "\u00A0"))
            add = set()
            for c in vs:
                add |= kana_norm_variants(c)
            vs |= add
            return {v for v in vs if v}

        def norm_stream_chars(page):
            rd = page.get_text("rawdict")
            chars = []
            line_counter = 0
            used_equal_split = False
            blocks = rd.get("blocks", []) if isinstance(rd, dict) else []
            for b in blocks:
                for line in b.get("lines", []):
                    for span in line.get("spans", []):
                        if "chars" in span and isinstance(span["chars"], list) and span["chars"]:
                            for ch in span["chars"]:
                                c = ch.get("c", ""); bbox = ch.get("bbox", None)
                                if not c or not bbox:
                                    continue
                                r = fitz.Rect(bbox)
                                chars.append({"orig": c, "rect": r, "line": line_counter})
                        else:
                            text = span.get("text", ""); bbox = span.get("bbox", None)
                            if not text or not bbox:
                                continue
                            used_equal_split = True
                            r = fitz.Rect(bbox); n = len(text)
                            if n <= 0:
                                continue
                            x0, y0, x1, y1 = r
                            dx, dy = 1.0, 0.0
                            d = span.get("dir", None)
                            if isinstance(d, (list, tuple)) and len(d) >= 2:
                                try:
                                    dx, dy = float(d[0] or 0.0), float(d[1] or 0.0)
                                except Exception:
                                    dx, dy = 1.0, 0.0
                            is_vertical = abs(dy) > abs(dx)
                            if is_vertical:
                                ch_h = max(1e-3, (y1 - y0) / n)
                                for idx, c in enumerate(text):
                                    rr = fitz.Rect(x0, y0 + idx*ch_h, x1, y0 + (idx+1)*ch_h)
                                    chars.append({"orig": c, "rect": rr, "line": line_counter})
                            else:
                                ch_w = max(1e-3, (x1 - x0) / n)
                                for idx, c in enumerate(text):
                                    rr = fitz.Rect(x0 + idx*ch_w, y0, x0 + (idx+1)*ch_w, y1)
                                    chars.append({"orig": c, "rect": rr, "line": line_counter})
                    line_counter += 1

            def norm_keep(c: str) -> str:
                if c in SPACE_CHARS or c in CONN_CHARS:
                    return ""
                c = hira_to_kata(nfkc(c))
                if c in SPACE_CHARS or c in CONN_CHARS:
                    return ""
                return c

            def norm_nochoon(c: str) -> str:
                s = norm_keep(c)
                if s == "ー":
                    return ""
                return s

            keep_list, keep_idxmap = [], []
            noch_list, noch_idxmap = [], []
            for i, it in enumerate(chars):
                c = it["orig"]
                nk = norm_keep(c)
                if nk:
                    keep_list.append(nk); keep_idxmap.append(i)
                nn = norm_nochoon(c)
                if nn:
                    noch_list.append(nn); noch_idxmap.append(i)
            stream_keep = "".join(keep_list)
            stream_noch = "".join(noch_list)
            return chars, (stream_keep, keep_idxmap), (stream_noch, noch_idxmap), used_equal_split

        def search_stream(stream_text: str, idxmap: list, qnorm: str):
            pos = 0; spans = []; Lq = len(qnorm)
            if not qnorm or not stream_text:
                return spans
            while True:
                k = stream_text.find(qnorm, pos)
                if k < 0:
                    break
                s_idx = k; e_idx = k + Lq - 1
                spans.append((idxmap[s_idx], idxmap[e_idx]))
                pos = k + 1
            return spans

        def merge_to_line_rects(chars, idx_s: int, idx_e: int):
            from collections import defaultdict
            items = chars[idx_s:idx_e+1]
            if not items:
                return []
            by_line = defaultdict(list)
            for it in items:
                by_line[it["line"]].append(it["rect"])
            rects = []
            for ln in sorted(by_line.keys()):
                rs = by_line[ln]
                x0 = min(r.x0 for r in rs); y0 = min(r.y0 for r in rs)
                x1 = max(r.x1 for r in rs); y1 = max(r.y1 for r in rs)
                rects.append(fitz.Rect(x0, y0, x1, y1))
            return rects

        def dedup(rects):
            uniq, seen = [], set()
            for rr in rects:
                key = (round(rr.x0, 2), round(rr.y0, 2), round(rr.x1, 2), round(rr.y1, 2))
                if key not in seen:
                    seen.add(key); uniq.append(rr)
            return uniq

        def robust_find_on_page(page, queries: list):
            chars, (stream_keep, keep_idxmap), (stream_noch, noch_idxmap), _ = norm_stream_chars(page)
            qnorm_keep_set, qnorm_noch_set = set(), set()
            for q in queries:
                for v in base_variants(q):
                    v_keep = hira_to_kata(nfkc(v))
                    v_keep = "".join(ch for ch in v_keep if ch not in SPACE_CHARS and ch not in CONN_CHARS)
                    if v_keep:
                        qnorm_keep_set.add(v_keep)
                    v_noch = v_keep.replace("ー", "")
                    if v_noch:
                        qnorm_noch_set.add(v_noch)

            stream_hits = []
            for qk in qnorm_keep_set:
                for s_idx, e_idx in search_stream(stream_keep, keep_idxmap, qk):
                    stream_hits.extend(merge_to_line_rects(chars, s_idx, e_idx))
            for qn in qnorm_noch_set:
                for s_idx, e_idx in search_stream(stream_noch, noch_idxmap, qn):
                    stream_hits.extend(merge_to_line_rects(chars, s_idx, e_idx))
            stream_hits = dedup(stream_hits)

            search_hits = []
            vset = set()
            for q in queries:
                vset |= base_variants(q)
            for cand in vset:
                try:
                    search_hits += page.search_for(cand)
                except Exception:
                    pass
            search_hits = dedup(search_hits)
            rect_hits = search_hits if search_hits else stream_hits
            return dedup(rect_hits)

        def suppress_overlap_hits(hits, iou_th=0.60):
            def area(r): return max(0.0, (r.x1 - r.x0) * (r.y1 - r.y0))
            def inter(r1, r2):
                x0 = max(r1.x0, r2.x0); y0 = max(r1.y0, r2.y0)
                x1 = min(r1.x1, r2.x1); y1 = min(r1.y1, r2.y1)
                if x1 <= x0 or y1 <= y0: return 0.0
                return (x1 - x0) * (y1 - y0)
            kept = []
            hits_sorted = sorted(hits, key=lambda r: (-area(r), r.y0, r.x0))
            for r in hits_sorted:
                drop = False
                for rk in kept:
                    inter_a = inter(r, rk)
                    if inter_a <= 0: 
                        continue
                    iou = inter_a / (area(r) + area(rk) - inter_a + 1e-6)
                    if iou >= iou_th:
                        drop = True; break
                    if (r.x0 >= rk.x0 - 0.5 and r.y0 >= rk.y0 - 0.5 and
                        r.x1 <= rk.x1 + 0.5 and r.y1 <= rk.y1 + 0.5):
                        drop = True; break
                if not drop:
                    kept.append(r)
            return kept

        # ===== 4) PDF へマーキング（行番号は rows の並び順で 1,2,3,... 固定） =====
        try:
            for src in self.files:
                doc = None
                try:
                    doc = fitz.open(src)
                    for page in doc:
                        page_rects = []
                        # row_no はページ/ファイルをまたいでも同じ行に対して一定
                        for row_no, pair in enumerate(rows, start=1):
                            for kind, s in pair:
                                rects = robust_find_on_page(page, [s])
                                rects = suppress_overlap_hits(rects)
                                for r in rects:
                                    page_rects.append((r, s, row_no, kind))

                        # アノテーション反映（title: "行-種別", subject: "#行"）
                        for r, text, row_no, kind in page_rects:
                            ann = page.add_highlight_annot(r)
                            width = 3
                            ann.set_info({
                                "title":   f"{row_no:0{width}d}-{kind}",
                                "subject": f"#{row_no:0{width}d}",
                                "content": text
                            })
                            color = HIGHLIGHT_COLORS.get(kind, COLOR_YELLOW)
                            ann.set_colors(stroke=color)
                            ann.update()

                    base = os.path.splitext(os.path.basename(src))[0]
                    out_path = os.path.join(out_dir, f"{base}_marked.pdf")
                    doc.save(out_path)
                finally:
                    if doc is not None:
                        doc.close()
            QMessageBox.information(self, "完了", "PDFへマーキングしました。")
        except Exception as e:
            QMessageBox.critical(self, "エラー", str(e))


    # ===== [NEW] プレビュー更新: 選択インデックスから a/b を拾って差分を描画 =====
    def _show_inline_diff_for_index(self, proxy_index: QModelIndex):
        try:
            if not proxy_index or not proxy_index.isValid():
                return
            # Proxy -> Source にマップ
            src_index = self.proxy_unified.mapToSource(proxy_index)
            row = src_index.row()

            # UnifiedModel 側の DataFrame から a/b を取得
            df = self.model_unified._df  # 既存実装に倣い内部DFへアクセス
            if df is None or df.empty:
                return
            if "a" not in df.columns or "b" not in df.columns:
                return

            a_txt = "" if pd.isna(df.at[row, "a"]) else str(df.at[row, "a"])
            b_txt = "" if pd.isna(df.at[row, "b"]) else str(df.at[row, "b"])

            html = build_vertical_diff_html_embed(a_txt, b_txt)
            self.diff_view.setHtml(html)
        except Exception:
            pass

    # ===== 差し替え: 左プレビューを統合インライン差分に切り替え =====
    def on_unified_selection_changed(self, selected, deselected):
        try:
            sel = self.view_unified.selectionModel().selectedRows()
        except Exception:
            sel = []
        if not sel:
            # 何も選択されていないときのプレースホルダ
            self.diff_view.setHtml(
                "<html><body style='font-family:Yu Gothic UI, Noto Sans JP, sans-serif;"
                "color:#666; font-size:12px; margin:4px 6px;'>"
                "表のセルを選択すると、<b>統合インライン差分</b>をここに表示します。"
                "</body></html>"
            )
            return

        proxy_index = sel[0]
        src_index = self.proxy_unified.mapToSource(proxy_index)
        row = src_index.row()

        df = getattr(self.model_unified, "_df", pd.DataFrame())
        if df.empty or row < 0 or row >= len(df):
            return

        a = df.at[row, "a"] if "a" in df.columns else ""
        b = df.at[row, "b"] if "b" in df.columns else ""

        html = build_unified_inline_diff_embed(
            "" if a is None else str(a),
            "" if b is None else str(b)
        )
        self.diff_view.setHtml(html)
    # ===== 差し替えここまで =====

    # ===== [ADD] MainWindow: NLPテキスト保存のON/OFF反映 =====
    def on_save_nlp_changed(self, state):
        # UI のチェック状態をグローバルトグルへ反映
        set_save_nlp(bool(state))
    # ===== [/ADD] =====



def main():
    app = QApplication(sys.argv)
    apply_light_theme(app, base_font_size=10)
    win = MainWindow(); win.show()
    sys.exit(app.exec())

if __name__ == "__main__":
    import multiprocessing as mp
    mp.freeze_support()  # ★ 追加：凍結exeでの子プロセス起動対策
    main()

